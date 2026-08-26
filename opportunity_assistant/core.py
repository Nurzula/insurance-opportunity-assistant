"""乙方宝商机导出文件的离线解析与确定性业务规则。

本模块刻意不依赖 Streamlit，也不执行网络请求。所有公开函数都接收 bytes、
DataFrame 或普通标量并返回新对象，便于桌面版 UI、单元测试和未来的 FastAPI
接口复用。
"""

from __future__ import annotations

from datetime import date, datetime, timedelta
from hashlib import sha1
from io import BytesIO
import math
import re
from typing import Any, Mapping
from zipfile import BadZipFile, ZipFile

import pandas as pd


OLE_BIFF_MAGIC = bytes.fromhex("D0CF11E0A1B11AE1")
ZIP_MAGICS = (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")

STANDARD_COLUMNS = [
    "关键词",
    "项目名称",
    "信息发布时间",
    "项目编号",
    "发布省份",
    "发布市级",
    "发布区级",
    "招标阶段",
    "报名截止时间",
    "投标截止时间",
    "招标金额（元）",
    "招标单位",
    "招标单位联系人",
    "招标单位联系人电话",
    "代理单位",
    "代理单位联系人",
    "代理单位联系人电话",
    "官网查看地址",
]

DERIVED_COLUMNS = [
    "来源类型",
    "源文件行号",
    "标准金额",
    "金额状态",
    "发布日期",
    "报名截止日期",
    "投标截止日期",
    "险种分类",
    "商机分类",
    "判定状态",
    "判定理由",
    "是否纳入",
    "需人工复核",
    "项目去重键",
    "公告去重键",
    "是否重复",
    "区域大类",
    "区域归属",
    "复核意见",
    "推送备注",
]

EVIDENCE_COLUMNS = [
    "数据来源",
    "来源平台",
    "官方来源标识",
    "来源分类",
    "公告正文",
    "内容摘要",
    "证据摘录",
    "金额依据",
    "金额口径",
    "金额提取依据",
    "报名截止原文",
    "投标截止原文",
    "正文取证状态",
    "AI理由",
    "AI判定",
    "AI置信度",
    "AI复核模型",
]

OUTPUT_COLUMNS = STANDARD_COLUMNS + DERIVED_COLUMNS + EVIDENCE_COLUMNS

INSURANCE_TYPES = (
    "工程险",
    "货运险",
    "意外险",
    "健康险",
    "责任险",
    "企财险",
    "信用险",
    "保证险",
)

NO_SERVICE_DISTRICTS = frozenset(
    {
        "成华区",
        "锦江区",
        "高新区",
        "成都高新区",
        "天府新区",
        "四川天府新区",
    }
)

DEFAULT_ENGINEERING_MIN_AMOUNT = 10_000_000.0
MAX_REASONABLE_AMOUNT = 1_000_000_000_000.0


class OpportunityWorkbookError(ValueError):
    """商机工作簿无法识别、解析或不符合乙方宝导出结构。"""


def detect_excel_format(data: bytes) -> str:
    """根据文件内容识别 BIFF ``.xls`` 或 OOXML ``.xlsx``。

    文件名后缀不会参与判断，因此用户即使手工改错后缀也不会误读。普通 ZIP
    文件不会被误认成 xlsx；必须同时存在 OOXML 的内容类型和工作簿部件。
    """

    if not isinstance(data, (bytes, bytearray, memoryview)):
        raise TypeError("data 必须是 Excel 文件的二进制内容")
    payload = bytes(data)
    if payload.startswith(OLE_BIFF_MAGIC):
        return "xls"
    if payload.startswith(ZIP_MAGICS):
        try:
            with ZipFile(BytesIO(payload)) as archive:
                names = set(archive.namelist())
        except (BadZipFile, OSError) as exc:
            raise OpportunityWorkbookError("文件具有 ZIP 标记，但不是有效的 xlsx") from exc
        if "[Content_Types].xml" in names and "xl/workbook.xml" in names:
            return "xlsx"
    raise OpportunityWorkbookError("无法识别文件格式：仅支持真实的 .xls 或 .xlsx")


def _read_xls_rows(data: bytes) -> tuple[list[list[Any]], int]:
    try:
        import xlrd

        workbook = xlrd.open_workbook(file_contents=data, on_demand=True)
        sheet = (
            workbook.sheet_by_name("商机信息导出")
            if "商机信息导出" in workbook.sheet_names()
            else workbook.sheet_by_index(0)
        )
        rows = [sheet.row_values(index) for index in range(sheet.nrows)]
        datemode = int(workbook.datemode)
        workbook.release_resources()
        return rows, datemode
    except OpportunityWorkbookError:
        raise
    except Exception as exc:  # xlrd 会抛出多种 BIFF 专用异常
        raise OpportunityWorkbookError(f"无法解析 .xls 文件：{exc}") from exc


def _read_xlsx_rows(data: bytes) -> tuple[list[list[Any]], int]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        sheet = (
            workbook["商机信息导出"]
            if "商机信息导出" in workbook.sheetnames
            else workbook[workbook.sheetnames[0]]
        )
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        workbook.close()
        return rows, 0
    except Exception as exc:
        raise OpportunityWorkbookError(f"无法解析 .xlsx 文件：{exc}") from exc


def _plain_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if bool(pd.isna(value)):
            return ""
    except (TypeError, ValueError):
        # 容器不是合法单元格值；交给 str 形成可诊断文本。
        pass
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).strip())


def _is_placeholder(value: Any) -> bool:
    return _plain_text(value).strip().lower() in {
        "",
        "--",
        "-",
        "/",
        "暂无",
        "暂未公布",
        "未公布",
        "none",
        "nan",
        "null",
    }


def clean_date(value: Any, *, datemode: int = 0) -> date | None:
    """清洗 Excel 日期、datetime/date 或常见中文日期字符串。"""

    if value is None or _is_placeholder(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if not math.isfinite(float(value)) or float(value) <= 0:
            return None
        try:
            import xlrd

            return xlrd.xldate_as_datetime(float(value), datemode).date()
        except Exception:
            # OOXML 的默认 1900 日期系统；仅作为没有 xlrd 时的安全回退。
            try:
                return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
            except (OverflowError, ValueError):
                return None

    text = _plain_text(value)
    text = re.sub(r"\s+", " ", text)
    formats = (
        "%Y/%m/%d",
        "%Y-%m-%d",
        "%Y.%m.%d",
        "%Y年%m月%d日",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d %H:%M",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        parsed = pd.to_datetime(text, errors="coerce")
        return None if pd.isna(parsed) else parsed.date()
    except (TypeError, ValueError, OverflowError):
        return None


def clean_amount(value: Any) -> float | None:
    """把数值或含 元/万/亿 的金额文本统一为人民币元。"""

    if value is None or _is_placeholder(value) or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None

    text = _plain_text(value).replace(",", "").replace("，", "")
    text = text.replace("人民币", "").replace("RMB", "").replace("￥", "").replace("¥", "")
    match = re.search(r"[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?", text)
    if not match:
        return None
    try:
        number = float(match.group(0))
    except ValueError:
        return None
    if "亿" in text:
        number *= 1_000_000_000_000 if "万亿" in text else 100_000_000
    elif "万" in text:
        number *= 10_000
    return number if math.isfinite(number) else None


def _find_header_index(rows: list[list[Any]]) -> int:
    candidates: list[int] = []
    for index, row in enumerate(rows[:10]):
        values = [_plain_text(value) for value in row]
        if len(values) >= 11 and values[0] == "关键词" and values[1] == "项目名称":
            if "招标金额" in values[10]:
                candidates.append(index)
    if not candidates:
        raise OpportunityWorkbookError("未找到乙方宝双层表头（关键词/项目名称/招标金额）")
    # 双层表头的第二行包含完整联系人字段；取最后一个连续候选行。
    return candidates[-1]


def parse_yifangbao_excel(data: bytes, filename: str | None = None) -> pd.DataFrame:
    """从乙方宝导出的 xls/xlsx bytes 解析出标准 18 列 DataFrame。

    ``filename`` 只用于错误消息，绝不会用于格式判断。日期被转换为 ``date``，
    金额转换为浮点元，空占位符转换为空字符串/None，并额外保留 Excel 源行号。
    """

    payload = bytes(data)
    file_format = detect_excel_format(payload)
    rows, datemode = _read_xls_rows(payload) if file_format == "xls" else _read_xlsx_rows(payload)
    if not rows:
        raise OpportunityWorkbookError("工作簿没有可读取的数据")
    header_index = _find_header_index(rows)

    records: list[dict[str, Any]] = []
    for row_index, raw_row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        values = list(raw_row[: len(STANDARD_COLUMNS)])
        values.extend([None] * (len(STANDARD_COLUMNS) - len(values)))
        if _is_placeholder(values[1]):
            continue
        record: dict[str, Any] = {}
        for column, value in zip(STANDARD_COLUMNS, values):
            if column in {"信息发布时间", "报名截止时间", "投标截止时间"}:
                record[column] = clean_date(value, datemode=datemode)
            elif column == "招标金额（元）":
                record[column] = clean_amount(value)
            else:
                record[column] = "" if _is_placeholder(value) else _plain_text(value)
        record["源文件行号"] = row_index
        records.append(record)

    if not records:
        suffix = f"（{filename}）" if filename else ""
        raise OpportunityWorkbookError(f"未读取到商机明细{suffix}")
    frame = pd.DataFrame.from_records(records)
    return frame.reindex(columns=STANDARD_COLUMNS + ["源文件行号"])


_INSURANCE_TYPE_PATTERNS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("工程险", ("建筑工程一切险", "安装工程一切险", "工程保险", "工程险")),
    ("货运险", ("货物运输保险", "货运保险", "货运险")),
    (
        "意外险",
        (
            "意外伤害险",
            "意外伤害保险",
            "团体意外险",
            "团体意外保险",
            "人身意外团体保险",
            "员工意外",
        ),
    ),
    ("健康险", ("健康保险", "健康险", "补充医疗保险", "医疗保险")),
    (
        "责任险",
        (
            "责任险",
            "责任保险",
            "雇主责任",
            "安全生产责任",
            "公众责任",
            "第三者责任",
        ),
    ),
    ("企财险", ("企业财产保险", "财产一切险", "财产综合险", "财产保险", "企财险")),
    ("信用险", ("信用保险", "信用险")),
    (
        "保证险",
        (
            "保证保险",
            "保证险",
            "履约保证金保险",
            "投标保证金保险",
            "工资支付履约保证保险",
        ),
    ),
)

_INSURANCE_SERVICE_EXCLUSION = re.compile(
    r"(?:保险|财险|人寿|学平险).*(?:宣传|广告|营销|活动|软件|系统|小程序|培训|装修|家具|加装|电梯|设备|职场)"
    r"|(?:宣传|广告|营销|软件|小程序|培训|装修|家具|加装|电梯|设备).*(?:保险|财险|人寿|学平险)",
    re.IGNORECASE,
)
_INSURANCE_IRRELEVANT = re.compile(
    r"保险孔|排危除险|避险(?:工程|搬迁|广场)|抢险|病险(?:水库|工程)|除险加固|风险|危险|社会保险|医疗保障|保险经办人员",
    re.IGNORECASE,
)
_NON_TARGET_INSURANCE = re.compile(
    r"(?:车辆保险|机动车保险|车险|农业保险|种植保险|养殖保险|养老保险)(?!.*责任)",
    re.IGNORECASE,
)
_OBSOLETE_PATTERN = re.compile(r"该信息已更新即将删除|已失效|信息已删除")


def classify_insurance_title(title: Any) -> dict[str, Any]:
    """按八大险种对一个标题作可审计的确定性分类。"""

    text = _plain_text(title)
    if not text:
        return {
            "险种分类": "",
            "商机分类": "待复核",
            "判定状态": "review",
            "判定理由": "项目名称为空",
        }
    if _OBSOLETE_PATTERN.search(text):
        return {
            "险种分类": "",
            "商机分类": "非保险",
            "判定状态": "excluded",
            "判定理由": "乙方宝标记为已更新即将删除",
        }
    if "雇员忠诚险" in text or "现金险" in text:
        return {
            "险种分类": "企财险（候选）",
            "商机分类": "待复核",
            "判定状态": "review",
            "判定理由": "忠诚险/现金险与企财险口径存在歧义，需业务老师确认",
        }
    if _INSURANCE_SERVICE_EXCLUSION.search(text):
        return {
            "险种分类": "",
            "商机分类": "非保险",
            "判定状态": "excluded",
            "判定理由": "采购对象是保险机构的宣传、系统、培训或其他货物服务，而非保单",
        }

    matched = [
        insurance_type
        for insurance_type, patterns in _INSURANCE_TYPE_PATTERNS
        if any(pattern.casefold() in text.casefold() for pattern in patterns)
    ]
    # “医疗责任保险”只归责任险，避免因“医疗保险”子串产生双重误判。
    if "责任险" in matched and "健康险" in matched and re.search(r"医疗责任(?:险|保险)", text):
        matched.remove("健康险")
    if matched:
        return {
            "险种分类": "、".join(dict.fromkeys(matched)),
            "商机分类": "保险商机",
            "判定状态": "accepted",
            "判定理由": f"标题明确命中八大险种：{'、'.join(dict.fromkeys(matched))}",
        }
    if _NON_TARGET_INSURANCE.search(text):
        return {
            "险种分类": "",
            "商机分类": "非保险",
            "判定状态": "excluded",
            "判定理由": "标题属于当前八大险种之外的保险品类",
        }
    if _INSURANCE_IRRELEVANT.search(text):
        return {
            "险种分类": "",
            "商机分类": "非保险",
            "判定状态": "excluded",
            "判定理由": "“险”用于风险、危险、抢险、避险或除险等非保险语义",
        }
    if "保险" in text or "险" in text:
        return {
            "险种分类": "未确定",
            "商机分类": "待复核",
            "判定状态": "review",
            "判定理由": "标题涉及保险，但无法高置信映射到八大险种",
        }
    return {
        "险种分类": "",
        "商机分类": "非保险",
        "判定状态": "excluded",
        "判定理由": "标题未体现保险采购",
    }


_ENGINEERING_NON_PROJECT = re.compile(
    r"经营权出让|直线加速器|医疗设备|设备采购|办公家具|车辆维护|普通货物.*运输|零散运输|搬迁运输|"
    r"环境影响评价文件受理|宣传服务|软件开发|系统开发|试剂|耗材|保险服务|产品采购",
    re.IGNORECASE,
)
_ENGINEERING_REVIEW = re.compile(
    r"设备服务|制作招标|采购项目(?:\(.*\))?(?:采购)?(?:更正|变更)?公告",
    re.IGNORECASE,
)
_ENGINEERING_EARLY = re.compile(
    r"招标计划|招标预告|提前公示|预公示|规划手续|勘察设计|设计服务|造价咨询|预算审核|"
    r"监理(?:服务)?|检测(?:服务)?|可行性研究|咨询服务|审批项目",
    re.IGNORECASE,
)
_ENGINEERING_DIRECT = re.compile(
    r"施工|设计施工总承包|工程总承包|EPC|专业分包|劳务(?:分包|工程|合作)|土石方|弱电|消防分包|"
    r"道路(?:建设|改造|工程)|桥梁|管网|水库|航道整治|基础设施|房建|装修装饰|变电站|安防工程|"
    r"智慧医院工程|灌溉|排水防涝|水生态保护修复|边坡治理|灾害治理|钻井工程|改造提升|建设工程",
    re.IGNORECASE,
)


def amount_status(amount: Any, *, min_amount: float | None = None) -> str:
    """返回金额的 UI 状态：正常、缺失、异常或低于门槛。"""

    number = clean_amount(amount)
    if number is None:
        return "缺失"
    if number < 0 or number > MAX_REASONABLE_AMOUNT:
        return "异常"
    if min_amount is not None and number < float(min_amount):
        return "低于门槛"
    return "正常"


def classify_engineering_opportunity(
    title: Any,
    amount: Any,
    *,
    stage: Any = "",
    min_amount: float = DEFAULT_ENGINEERING_MIN_AMOUNT,
) -> dict[str, Any]:
    """分类工程商机并执行一千万元门槛与异常金额保护。"""

    text = _plain_text(title)
    stage_text = _plain_text(stage)
    money_state = amount_status(amount, min_amount=min_amount)
    if _OBSOLETE_PATTERN.search(text):
        return {
            "险种分类": "工程险",
            "商机分类": "非工程",
            "判定状态": "excluded",
            "判定理由": "乙方宝标记为已更新即将删除",
            "金额状态": money_state,
        }
    if money_state == "异常":
        return {
            "险种分类": "工程险",
            "商机分类": "待复核",
            "判定状态": "review",
            "判定理由": "金额明显超出合理范围，可能是源数据格式错误",
            "金额状态": money_state,
        }
    if money_state == "缺失":
        return {
            "险种分类": "工程险",
            "商机分类": "待复核",
            "判定状态": "review",
            "判定理由": "缺少招标金额，无法执行一千万元门槛",
            "金额状态": money_state,
        }

    if _ENGINEERING_EARLY.search(text) or stage_text in {"招标预告", "审批项目"}:
        category = "前期线索"
    elif _ENGINEERING_DIRECT.search(text):
        category = "直接施工"
    elif _ENGINEERING_NON_PROJECT.search(text):
        category = "非工程"
    elif _ENGINEERING_REVIEW.search(text) or "工程" in text or "项目" in text:
        category = "待复核"
    else:
        # 没有足够语义时不能仅凭短标题静默排除，交给业务老师复核。
        category = "待复核"

    if money_state == "低于门槛":
        return {
            "险种分类": "工程险",
            "商机分类": category,
            "判定状态": "excluded",
            "判定理由": f"招标金额低于{min_amount:,.0f}元门槛",
            "金额状态": money_state,
        }
    if category == "非工程":
        return {
            "险种分类": "工程险",
            "商机分类": category,
            "判定状态": "excluded",
            "判定理由": "采购对象属于货物、运输、维护或其他非工程内容",
            "金额状态": money_state,
        }
    if category == "待复核":
        return {
            "险种分类": "工程险",
            "商机分类": category,
            "判定状态": "review",
            "判定理由": "金额达标，但标题不足以确认是否形成工程险机会",
            "金额状态": money_state,
        }
    return {
        "险种分类": "工程险",
        "商机分类": category,
        "判定状态": "accepted",
        "判定理由": "金额达标且标题体现直接施工" if category == "直接施工" else "金额达标，作为工程前期线索保留",
        "金额状态": money_state,
    }


def _canonical_district(value: Any) -> str:
    text = _plain_text(value)
    return "" if _is_placeholder(text) else text.replace(" ", "")


def assign_region(province: Any, city: Any, district: Any) -> dict[str, str]:
    """按部门口径返回区域大类和区域归属。"""

    province_text = _plain_text(province).replace("省", "")
    city_text = _plain_text(city).replace("市", "")
    district_text = _canonical_district(district)
    if city_text == "成都":
        if not district_text:
            return {"区域大类": "成都地区", "区域归属": "地区未明确"}
        if district_text in NO_SERVICE_DISTRICTS or any(
            marker in district_text for marker in ("高新区", "天府新区")
        ):
            return {"区域大类": "成都地区", "区域归属": "无区域类"}
        return {"区域大类": "成都地区", "区域归属": district_text}
    if province_text == "四川" or (not province_text and city_text):
        if city_text:
            return {"区域大类": "川内其他地区", "区域归属": f"{city_text}市"}
        return {"区域大类": "川内其他地区", "区域归属": "地区未明确"}
    if not province_text and not city_text:
        return {"区域大类": "地区未明确", "区域归属": "地区未明确"}
    return {"区域大类": "省外", "区域归属": _plain_text(city) or _plain_text(province)}


def _normalize_project_title(title: Any) -> str:
    text = _plain_text(title)
    text = _OBSOLETE_PATTERN.sub("", text)
    text = re.sub(r"^\s*(?:\[[^\]]+\]\s*)+", "", text)
    text = re.sub(r"[（(](?:第)?[一二三四五六七八九十\d]+次[）)]", "", text)
    lifecycle_patterns = (
        r"变更[（(]?补遗[）)]?公告\d*",
        r"(?:第\d+次)?(?:更正|变更|澄清|补遗|答疑)公告.*$",
        r"(?:公开)?(?:招标|竞价|比选|询价|评选|采购|谈判|磋商)(?:采购)?公告.*$",
        r"竞争性(?:磋商|谈判)(?:采购)?(?:邀请|公告).*$",
        r"招标文件(?:提前)?公示.*$",
    )
    for pattern in lifecycle_patterns:
        text = re.sub(pattern, "", text, flags=re.IGNORECASE)
    text = re.sub(r"[\s\-—_/·，,。；;：:（）()\[\]【】﹒]+", "", text)
    return text.casefold()


def make_project_key(row: Mapping[str, Any] | pd.Series) -> str:
    """生成跨变更/补遗公告稳定的项目去重键。"""

    title = _normalize_project_title(row.get("项目名称", ""))
    province = _plain_text(row.get("发布省份", ""))
    city = _plain_text(row.get("发布市级", ""))
    material = f"{province}|{city}|{title}"
    return f"PRJ-{sha1(material.encode('utf-8')).hexdigest()[:16].upper()}"


def make_announcement_key(row: Mapping[str, Any] | pd.Series) -> str:
    """生成公告级去重键；乙方宝 infoDetail ID 优先。"""

    official_id = _plain_text(row.get("官方来源标识", ""))
    if official_id:
        return f"ANN-{sha1(official_id.encode('utf-8')).hexdigest()[:16].upper()}"
    url = _plain_text(row.get("官网查看地址", ""))
    match = re.search(r"/infoDetail/(\d+)", url)
    if match:
        return f"ANN-{match.group(1)}"
    title = _plain_text(row.get("项目名称", "")).casefold()
    project_number = _plain_text(row.get("项目编号", ""))
    published = row.get("信息发布时间", "")
    material = f"{project_number}|{published}|{title}"
    return f"ANN-{sha1(material.encode('utf-8')).hexdigest()[:16].upper()}"


def _prepare_base_frame(frame: pd.DataFrame, source_type: str) -> pd.DataFrame:
    result = frame.copy(deep=True)
    for column in STANDARD_COLUMNS:
        if column not in result:
            result[column] = None if column == "招标金额（元）" else ""
    if "源文件行号" not in result:
        result["源文件行号"] = range(2, len(result) + 2)

    result["招标金额（元）"] = result["招标金额（元）"].map(clean_amount)
    result["标准金额"] = result["招标金额（元）"]
    for source, target, raw_column in (
        ("信息发布时间", "发布日期", ""),
        ("报名截止时间", "报名截止日期", "报名截止原文"),
        ("投标截止时间", "投标截止日期", "投标截止原文"),
    ):
        if raw_column and raw_column not in result:
            result[raw_column] = result[source].map(
                lambda value: "" if value is None or str(value).strip() in {"", "nan", "NaT"} else str(value).strip()
            )
        result[source] = result[source].map(clean_date)
        result[target] = result[source]
    result["来源类型"] = source_type
    if "数据来源" not in result:
        result["数据来源"] = "会员Excel导入"
    else:
        result["数据来源"] = result["数据来源"].fillna("").astype(str)
        result.loc[result["数据来源"].str.strip().eq(""), "数据来源"] = "会员Excel导入"
    for column in EVIDENCE_COLUMNS:
        if column not in result:
            result[column] = ""
    result["复核意见"] = ""
    result["推送备注"] = ""
    return result


def _finish_frame(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy(deep=True)
    region_rows = [
        assign_region(row["发布省份"], row["发布市级"], row["发布区级"])
        for _, row in result.iterrows()
    ]
    result["区域大类"] = [item["区域大类"] for item in region_rows]
    result["区域归属"] = [item["区域归属"] for item in region_rows]
    result["项目去重键"] = [make_project_key(row) for _, row in result.iterrows()]
    result["公告去重键"] = [make_announcement_key(row) for _, row in result.iterrows()]
    result["是否重复"] = result["项目去重键"].duplicated(keep=False)
    result["是否纳入"] = result["判定状态"].eq("accepted")
    result["需人工复核"] = result["判定状态"].eq("review")
    for column in OUTPUT_COLUMNS:
        if column not in result:
            result[column] = ""
    return result.reindex(columns=OUTPUT_COLUMNS).reset_index(drop=True)


def classify_insurance_dataframe(frame: pd.DataFrame) -> pd.DataFrame:
    """返回带八险种判定、区域和去重信息的保险结果副本。"""

    result = _prepare_base_frame(frame, "保险")
    decisions = [classify_insurance_title(title) for title in result["项目名称"]]
    for column in ("险种分类", "商机分类", "判定状态", "判定理由"):
        result[column] = [decision[column] for decision in decisions]
    result["金额状态"] = result["标准金额"].map(lambda value: amount_status(value))
    return _finish_frame(result)


def classify_engineering_dataframe(
    frame: pd.DataFrame,
    min_amount: float = DEFAULT_ENGINEERING_MIN_AMOUNT,
) -> pd.DataFrame:
    """返回带千万级门槛、工程阶段、区域和去重信息的工程结果副本。"""

    if min_amount < 0:
        raise ValueError("min_amount 不能为负数")
    result = _prepare_base_frame(frame, "工程")
    decisions = [
        classify_engineering_opportunity(
            row["项目名称"],
            row["标准金额"],
            stage=row["招标阶段"],
            min_amount=min_amount,
        )
        for _, row in result.iterrows()
    ]
    for column in ("险种分类", "商机分类", "判定状态", "判定理由", "金额状态"):
        result[column] = [decision[column] for decision in decisions]
    return _finish_frame(result)


def assign_regions(frame: pd.DataFrame) -> pd.DataFrame:
    """只重新计算区域字段，不改变输入 DataFrame。"""

    result = frame.copy(deep=True)
    regions = [
        assign_region(row.get("发布省份", ""), row.get("发布市级", ""), row.get("发布区级", ""))
        for _, row in result.iterrows()
    ]
    result["区域大类"] = [item["区域大类"] for item in regions]
    result["区域归属"] = [item["区域归属"] for item in regions]
    return result


def process_uploaded_workbook(
    data: bytes,
    source_type: str,
    filename: str | None = None,
    min_amount: float = DEFAULT_ENGINEERING_MIN_AMOUNT,
) -> pd.DataFrame:
    """解析并处理一份上传文件，是 Streamlit 最常用的单入口。"""

    normalized_type = _plain_text(source_type).casefold()
    frame = parse_yifangbao_excel(data, filename=filename)
    if normalized_type in {"保险", "险", "insurance"}:
        return classify_insurance_dataframe(frame)
    if normalized_type in {"工程", "engineering", "construction"}:
        return classify_engineering_dataframe(frame, min_amount=min_amount)
    raise ValueError("source_type 只能是保险/insurance 或工程/engineering")


def process_opportunity_files(
    insurance_data: bytes,
    engineering_data: bytes,
    *,
    insurance_filename: str | None = None,
    engineering_filename: str | None = None,
    min_amount: float = DEFAULT_ENGINEERING_MIN_AMOUNT,
) -> dict[str, pd.DataFrame]:
    """一次处理两份文件并返回保险、工程及合并结果。"""

    insurance = process_uploaded_workbook(
        insurance_data,
        "保险",
        filename=insurance_filename,
        min_amount=min_amount,
    )
    engineering = process_uploaded_workbook(
        engineering_data,
        "工程",
        filename=engineering_filename,
        min_amount=min_amount,
    )
    combined = pd.concat([insurance, engineering], ignore_index=True)
    return {"insurance": insurance, "engineering": engineering, "combined": combined}


def split_results(frame: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """按固定状态拆成适合三个 data_editor/tab 的结果。"""

    if "判定状态" not in frame:
        raise ValueError("DataFrame 缺少判定状态列")
    return {
        status: frame.loc[frame["判定状态"].eq(status)].copy().reset_index(drop=True)
        for status in ("accepted", "review", "excluded")
    }


def build_region_summary(frame: pd.DataFrame) -> pd.DataFrame:
    """汇总已纳入商机的区域、保险/工程项目数和金额。"""

    accepted = frame.loc[frame["判定状态"].eq("accepted")].copy()
    columns = ["区域归属", "保险项目数", "保险金额（元）", "工程项目数", "工程金额（元）"]
    if accepted.empty:
        return pd.DataFrame(columns=columns)
    accepted["标准金额"] = pd.to_numeric(accepted["标准金额"], errors="coerce").fillna(0.0)
    records: list[dict[str, Any]] = []
    for region, group in accepted.groupby("区域归属", dropna=False, sort=True):
        insurance = group.loc[group["来源类型"].eq("保险")]
        engineering = group.loc[group["来源类型"].eq("工程")]
        records.append(
            {
                "区域归属": region or "地区未明确",
                "保险项目数": int(len(insurance)),
                "保险金额（元）": round(float(insurance["标准金额"].sum()), 2),
                "工程项目数": int(len(engineering)),
                "工程金额（元）": round(float(engineering["标准金额"].sum()), 2),
            }
        )
    total = {
        "区域归属": "总计",
        "保险项目数": sum(item["保险项目数"] for item in records),
        "保险金额（元）": round(sum(item["保险金额（元）"] for item in records), 2),
        "工程项目数": sum(item["工程项目数"] for item in records),
        "工程金额（元）": round(sum(item["工程金额（元）"] for item in records), 2),
    }
    return pd.DataFrame.from_records(records + [total], columns=columns)


def summarize_opportunities(frame: pd.DataFrame) -> dict[str, Any]:
    """返回可直接展示为指标卡的 JSON 友好摘要。"""

    if "判定状态" not in frame:
        raise ValueError("DataFrame 缺少判定状态列")
    accepted = frame.loc[frame["判定状态"].eq("accepted")]
    amount = pd.to_numeric(accepted.get("标准金额", pd.Series(dtype=float)), errors="coerce").fillna(0)
    return {
        "raw_count": int(len(frame)),
        "accepted_count": int(len(accepted)),
        "accepted_amount": round(float(amount.sum()), 2),
        "accepted_unique_project_count": int(accepted["项目去重键"].nunique()) if not accepted.empty else 0,
        "review_count": int(frame["判定状态"].eq("review").sum()),
        "excluded_count": int(frame["判定状态"].eq("excluded").sum()),
        "duplicate_row_count": int(frame.get("是否重复", pd.Series(False, index=frame.index)).fillna(False).sum()),
    }


__all__ = [
    "DEFAULT_ENGINEERING_MIN_AMOUNT",
    "DERIVED_COLUMNS",
    "EVIDENCE_COLUMNS",
    "INSURANCE_TYPES",
    "NO_SERVICE_DISTRICTS",
    "OUTPUT_COLUMNS",
    "OpportunityWorkbookError",
    "STANDARD_COLUMNS",
    "amount_status",
    "assign_region",
    "assign_regions",
    "build_region_summary",
    "classify_engineering_dataframe",
    "classify_engineering_opportunity",
    "classify_insurance_dataframe",
    "classify_insurance_title",
    "clean_amount",
    "clean_date",
    "detect_excel_format",
    "make_announcement_key",
    "make_project_key",
    "parse_yifangbao_excel",
    "process_opportunity_files",
    "process_uploaded_workbook",
    "split_results",
    "summarize_opportunities",
]
