"""商机推送助手的纯内存报表层。

该模块不依赖 Streamlit，也不在本地写入临时文件。调用方只需传入一张
已完成业务判断和区域分配的 DataFrame，即可获得 Excel、成都地区长图以及
企业微信群发文字。

公开主接口：

* ``build_opportunity_excel`` -> ``io.BytesIO``
* ``build_chengdu_opportunity_png`` -> ``io.BytesIO``
* ``build_wecom_messages`` -> ``(简洁文字, 完整文字)``
* ``build_report_bundle`` -> ``ReportBundle``

输入建议至少包含：selected（或“推送”）、business_type、category、
project_name、publish_date、amount、city、district、region_group、
service_region、stage、deadline/deadlines、tenderer、contact、agent、url、
decision_reason、quality_issue 和 project_key。缺失列会自动补空，不会让
“今日无数据”的场景中断。
"""

from __future__ import annotations

import io
import math
import os
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from functools import lru_cache
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension


__all__ = [
    "ReportBundle",
    "normalize_report_dataframe",
    "build_opportunity_excel",
    "build_excel_report",
    "build_chengdu_opportunity_png",
    "build_chengdu_png",
    "build_wecom_text",
    "build_wecom_messages",
    "build_report_bundle",
]


REQUIRED_COLUMNS: tuple[str, ...] = (
    "selected",
    "business_type",
    "category",
    "project_name",
    "publish_date",
    "amount",
    "city",
    "district",
    "region_group",
    "service_region",
    "stage",
    "deadline",
    "tenderer",
    "contact",
    "agent",
    "url",
    "decision_reason",
    "quality_issue",
    "project_key",
    "note",
    "source_keyword",
    "source_platform",
    "official_source_id",
    "source_category",
    "amount_basis",
    "amount_evidence",
    "evidence_excerpt",
    "evidence_status",
    "ai_decision",
    "ai_confidence",
    "ai_model",
    "ai_reason",
    "announcement_key",
)


COLUMN_ALIASES: Mapping[str, tuple[str, ...]] = {
    "selected": ("selected", "推送", "是否推送", "纳入推送", "include"),
    "business_type": ("business_type", "业务类型", "商机类型"),
    "category": ("category", "分类", "险种", "工程类型"),
    "project_name": ("project_name", "项目名称", "商机标题", "标题"),
    "publish_date": ("publish_date", "发布日期", "发布时间", "公告日期"),
    "amount": ("amount", "招标金额（元）", "招标金额(元)", "招标金额", "预算金额"),
    "city": ("city", "地市", "城市", "市州"),
    "district": ("district", "区县", "区域"),
    "region_group": ("region_group", "区域分组", "地区分组"),
    "service_region": ("service_region", "服务区域", "分配区域", "归属机构"),
    "stage": ("stage", "项目阶段", "公告类型"),
    "deadline": (
        "deadline",
        "deadlines",
        "bid_deadline",
        "registration_deadline",
        "截止时间",
        "投标截止时间",
        "报名截止时间",
    ),
    "tenderer": ("tenderer", "招标人", "采购人", "项目业主"),
    "contact": (
        "contact",
        "tenderer_contact",
        "tenderer_phone",
        "agent_contact",
        "agent_phone",
        "联系人",
        "联系方式",
    ),
    "agent": ("agent", "代理机构", "招标代理"),
    "url": ("url", "原文链接", "项目链接", "公告链接"),
    "decision_reason": ("decision_reason", "判定依据", "推送理由", "筛选理由"),
    "quality_issue": ("quality_issue", "数据问题", "质量问题", "异常说明"),
    "project_key": ("project_key", "项目唯一键", "项目键"),
    "note": ("note", "推送备注", "业务备注", "备注"),
    "source_keyword": ("source_keyword", "关键词", "命中关键词"),
    "source_platform": ("source_platform", "来源平台", "信息来源"),
    "official_source_id": ("official_source_id", "官方来源标识", "来源ID"),
    "source_category": ("source_category", "来源分类", "公告分类"),
    "amount_basis": ("amount_basis", "金额口径"),
    "amount_evidence": ("amount_evidence", "金额提取依据", "金额证据"),
    "evidence_excerpt": ("evidence_excerpt", "证据摘录", "正文证据"),
    "evidence_status": ("evidence_status", "正文取证状态"),
    "ai_decision": ("ai_decision", "AI判定"),
    "ai_confidence": ("ai_confidence", "AI置信度"),
    "ai_model": ("ai_model", "AI复核模型"),
    "ai_reason": ("ai_reason", "AI理由"),
    "announcement_key": ("announcement_key", "公告去重键"),
}


DETAIL_COLUMNS: tuple[tuple[str, str], ...] = (
    ("selected", "推送"),
    ("business_type", "业务类型"),
    ("category", "分类/险种"),
    ("project_name", "项目名称"),
    ("publish_date", "发布日期"),
    ("amount", "招标金额（元）"),
    ("city", "地市"),
    ("district", "区县"),
    ("region_group", "区域分组"),
    ("service_region", "服务区域"),
    ("stage", "项目阶段"),
    ("deadline", "截止时间"),
    ("tenderer", "招标人/采购人"),
    ("contact", "联系人/联系方式"),
    ("agent", "代理机构"),
    ("url", "原文链接"),
    ("decision_reason", "判定依据"),
    ("quality_issue", "数据问题"),
    ("project_key", "项目唯一键"),
)


# 四个已入选业务页面向日常分发，仅保留业务员需要的信息。
# 完整的判定依据、区域分组、唯一键等仍保留在“未分区域/筛除”审计页。
BUSINESS_DETAIL_COLUMNS: tuple[tuple[str, str], ...] = (
    ("_sequence", "序号"),
    ("category", "分类/险种"),
    ("project_name", "项目名称"),
    ("publish_date", "发布日期"),
    ("amount", "招标金额（元）"),
    ("city", "地市"),
    ("district", "区县"),
    ("service_region", "服务区域"),
    ("stage", "项目阶段"),
    ("deadline", "截止时间"),
    ("tenderer", "招标人/采购人"),
    ("contact", "联系人/联系方式"),
    ("agent", "代理机构"),
    ("url", "原文链接"),
    ("_remarks", "备注"),
)


AUDIT_COLUMNS: tuple[tuple[str, str], ...] = (
    ("selected", "最终纳入"),
    ("source_keyword", "命中关键词"),
    ("business_type", "业务类型"),
    ("project_name", "项目名称"),
    ("source_platform", "来源平台"),
    ("official_source_id", "官方来源标识"),
    ("source_category", "公告分类"),
    ("amount", "判定金额（元）"),
    ("amount_basis", "金额口径"),
    ("amount_evidence", "金额提取依据"),
    ("evidence_status", "正文取证状态"),
    ("evidence_excerpt", "原始证据摘录"),
    ("ai_decision", "AI判定"),
    ("ai_confidence", "AI置信度"),
    ("ai_model", "AI模型"),
    ("ai_reason", "AI理由"),
    ("decision_reason", "规则判定依据"),
    ("announcement_key", "公告去重键"),
    ("url", "官方原文"),
)


BUSINESS_DETAIL_SHEETS: frozenset[str] = frozenset(
    {
        "成都地区（保险）",
        "川内其他地区（保险）",
        "成都地区（工程）",
        "川内其他地区（工程）",
    }
)


SHEET_NAMES: tuple[str, ...] = (
    "今日商机汇总",
    "成都地区（保险）",
    "川内其他地区（保险）",
    "成都地区（工程）",
    "川内其他地区（工程）",
    "未分区域项目",
    "筛除记录",
    "采集与判定审计",
    "处理日志",
)


DEFAULT_REGION_ORDER: tuple[str, ...] = (
    "锦江区",
    "金牛区",
    "武侯区",
    "成华区",
    "高新区",
    "青羊区",
    "龙泉驿区",
    "天府新区",
    "双流区",
    "温江区",
    "郫都区",
    "新都区",
    "青白江区",
    "简阳市",
    "金堂县",
    "大邑县",
    "蒲江县",
    "新津区",
    "都江堰市",
    "彭州市",
    "邛崃市",
    "崇州市",
    "未明确",
    "无区域类",
    "川内其他地区",
)


# 报表配色：保险为蓝，工程为橙，未分区域/需确认数据为金黄。
NAVY = "17365D"
BLUE = "2F75B5"
LIGHT_BLUE = "D9EAF7"
ORANGE = "ED7D31"
LIGHT_ORANGE = "FCE4D6"
GREEN = "70AD47"
LIGHT_GREEN = "E2F0D9"
GOLD = "FFC000"
LIGHT_GOLD = "FFF2CC"
RED = "C00000"
LIGHT_RED = "F4CCCC"
GREY = "7F8C8D"
LIGHT_GREY = "E7E6E6"
WHITE = "FFFFFF"
TEXT = "1F2937"
GRID = "CBD5E1"


THIN_SIDE = Side(style="thin", color=GRID)
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


@dataclass(frozen=True)
class ReportBundle:
    """一次生成的四项交付物。"""

    excel: io.BytesIO
    png: io.BytesIO
    concise_text: str
    full_text: str

    @property
    def excel_bytes(self) -> bytes:
        """返回便于 ``st.download_button`` 使用的二进制内容。"""

        return self.excel.getvalue()

    @property
    def png_bytes(self) -> bytes:
        """返回长图的二进制内容。"""

        return self.png.getvalue()


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    try:
        missing = pd.isna(value)
    except (TypeError, ValueError):
        return False
    return bool(missing) if isinstance(missing, (bool, type(pd.NA))) else False


def _clean_text(value: Any) -> str:
    if _is_missing(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat", "null", "<na>"}:
        return ""
    return re.sub(r"[\t\r]+", " ", text)


def _formal_text(value: Any) -> str:
    """将内部引擎术语转为可对外分发的业务表述。"""

    text = _clean_text(value)
    return (
        text.replace("待复核", "需人工确认")
        .replace("人工复核", "人工确认")
        .replace("复核原因", "确认原因")
        .replace("复核意见", "确认意见")
    )


def _coerce_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and not _is_missing(value):
        return bool(value)
    text = _clean_text(value).lower()
    if text in {"1", "true", "yes", "y", "是", "推送", "纳入", "保留", "通过", "√"}:
        return True
    if text in {"0", "false", "no", "n", "否", "筛除", "排除", "不推送", "×", ""}:
        return False
    # 对未知表述采取安全默认：不在群里直接推送。
    return False


def _coerce_amount(value: Any) -> int | float | None:
    """将常见的金额文本转换为“元”，未披露返回 ``None``。"""

    if _is_missing(value):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float, Decimal)):
        number = float(value)
        if not math.isfinite(number):
            return None
        return int(round(number)) if number.is_integer() else round(number, 2)

    raw = _clean_text(value).replace(",", "").replace("，", "")
    if raw.lower() in {"", "--", "-", "/", "未披露", "暂无", "不详"}:
        return None
    multiplier = Decimal("1")
    if "亿" in raw:
        multiplier = Decimal("100000000")
    elif "万" in raw:
        multiplier = Decimal("10000")
    elif "千" in raw:
        multiplier = Decimal("1000")

    match = re.search(r"[-+]?\d+(?:\.\d+)?", raw)
    if not match:
        return None
    try:
        number = Decimal(match.group()) * multiplier
    except InvalidOperation:
        return None
    if not number.is_finite():
        return None
    quantized = number.quantize(Decimal("0.01"))
    if quantized == quantized.to_integral():
        return int(quantized)
    return float(quantized)


def _normalise_business_type(value: Any) -> str:
    text = _clean_text(value)
    if "保险" in text or text == "险" or text.startswith("险-"):
        return "保险"
    if "工程" in text:
        return "工程"
    return text or "待分类"


def _normalise_region_group(value: Any, city: Any) -> str:
    text = _clean_text(value)
    city_text = _clean_text(city)
    if "无区域" in text:
        return "无区域类"
    if any(token in text for token in ("待复核", "未明确", "未分配")):
        return "待复核"
    if "成都" in text or "成都" in city_text:
        return "成都地区"
    if "川内" in text or "四川" in text:
        return "川内其他地区"
    return text or "待复核"


def _clean_quality_issue(value: Any) -> str:
    """去掉前端组合质量字段中没有信息量的“正常；False”。"""

    text = _clean_text(value)
    if not text:
        return ""
    ignored = {
        "false",
        "0",
        "否",
        "正常",
        "金额正常",
        "无",
        "无需复核",
        "不需复核",
        "none",
        "nan",
    }
    cleaned: list[str] = []
    for part in re.split(r"[;；|]+", text):
        token = part.strip()
        if not token or token.casefold() in ignored:
            continue
        if token.casefold() in {"true", "1"}:
            token = "需人工复核"
        if token not in cleaned:
            cleaned.append(token)
    return "；".join(cleaned)


def _coalesce_columns(frame: pd.DataFrame, aliases: Sequence[str]) -> pd.Series:
    result = pd.Series([None] * len(frame), index=frame.index, dtype="object")
    for column in aliases:
        if column not in frame.columns:
            continue
        candidate = frame[column]
        empty = result.map(_is_missing) | result.map(lambda value: _clean_text(value) == "")
        result = result.where(~empty, candidate)
    return result


def normalize_report_dataframe(data: pd.DataFrame | Iterable[Mapping[str, Any]] | None) -> pd.DataFrame:
    """将业务结果转为报表层的稳定列集。

    函数不修改调用方的 DataFrame，也不会因为空表或缺列报错。
    """

    if data is None:
        source = pd.DataFrame()
    elif isinstance(data, pd.DataFrame):
        source = data.copy(deep=True)
    else:
        source = pd.DataFrame(list(data))

    normalized = pd.DataFrame(index=source.index)
    for canonical in REQUIRED_COLUMNS:
        normalized[canonical] = _coalesce_columns(source, COLUMN_ALIASES[canonical])

    normalized["selected"] = normalized["selected"].map(_coerce_bool).astype(bool)
    normalized["business_type"] = normalized["business_type"].map(_normalise_business_type)
    normalized["amount"] = normalized["amount"].map(_coerce_amount)

    text_columns = [
        column
        for column in REQUIRED_COLUMNS
        if column not in {"selected", "amount", "business_type"}
    ]
    for column in text_columns:
        normalized[column] = normalized[column].map(_clean_text)

    # 前端会分别传入报名截止和投标截止；若两者都有值，不丢弃任何一个时点。
    if "registration_deadline" in source.columns or "bid_deadline" in source.columns:
        registration = (
            source["registration_deadline"].map(_clean_text)
            if "registration_deadline" in source.columns
            else pd.Series("", index=source.index)
        )
        bid = (
            source["bid_deadline"].map(_clean_text)
            if "bid_deadline" in source.columns
            else pd.Series("", index=source.index)
        )
        combined: list[str] = []
        for existing, registration_value, bid_value in zip(
            normalized["deadline"], registration, bid
        ):
            parts: list[str] = []
            if registration_value:
                parts.append(f"报名：{registration_value}")
            if bid_value:
                parts.append(f"投标：{bid_value}")
            combined.append("；".join(parts) if parts else _clean_text(existing))
        normalized["deadline"] = combined

    # 将招标人、代理机构的联系人/电话合并到单个展示列，同时避免重复文本。
    if any(
        column in source.columns
        for column in ("tenderer_contact", "tenderer_phone", "agent_contact", "agent_phone")
    ):
        contact_values: list[str] = []
        for index, existing in normalized["contact"].items():
            pieces: list[str] = []
            tenderer_contact = _clean_text(source.at[index, "tenderer_contact"]) if "tenderer_contact" in source else ""
            tenderer_phone = _clean_text(source.at[index, "tenderer_phone"]) if "tenderer_phone" in source else ""
            agent_contact = _clean_text(source.at[index, "agent_contact"]) if "agent_contact" in source else ""
            agent_phone = _clean_text(source.at[index, "agent_phone"]) if "agent_phone" in source else ""
            tenderer_piece = " ".join(part for part in (tenderer_contact, tenderer_phone) if part)
            agent_piece = " ".join(part for part in (agent_contact, agent_phone) if part)
            if tenderer_piece:
                pieces.append(f"招标人：{tenderer_piece}")
            if agent_piece:
                pieces.append(f"代理：{agent_piece}")
            contact_values.append("；".join(pieces) or _clean_text(existing))
        normalized["contact"] = contact_values

    normalized["quality_issue"] = normalized["quality_issue"].map(_clean_quality_issue)

    normalized["region_group"] = [
        _normalise_region_group(region, city)
        for region, city in zip(normalized["region_group"], normalized["city"])
    ]
    normalized["project_name"] = normalized["project_name"].replace("", "（未命名项目）")

    # 保留数据的原始顺序，仅重置行号，便于与上游处理日志对照。
    return normalized.reset_index(drop=True)


def _report_date(value: date | datetime | str | None) -> date:
    if value is None:
        return date.today()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    parsed = pd.to_datetime(_clean_text(value), errors="coerce")
    if pd.isna(parsed):
        return date.today()
    return parsed.date()


def _date_cn(value: date) -> str:
    return f"{value.year}年{value.month}月{value.day}日"


def _date_short(value: date) -> str:
    return f"{value.month}月{value.day}日"


def _display_date(value: Any) -> Any:
    """在 Excel 中尽可能保留可排序的真实日期值。"""

    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, (date, datetime)):
        return value
    text = _clean_text(value)
    if not text:
        return ""
    # 只对单一日期格式做类型化，时间段、“见公告”等保留原文。
    if re.fullmatch(r"\d{4}[-/.]\d{1,2}[-/.]\d{1,2}(?:[ T]\d{1,2}:\d{2}(?::\d{2})?)?", text):
        parsed = pd.to_datetime(text, errors="coerce")
        if not pd.isna(parsed):
            return parsed.to_pydatetime()
    return text


def _is_nonempty(value: Any) -> bool:
    return bool(_clean_text(value))


def _review_mask(frame: pd.DataFrame) -> pd.Series:
    if frame.empty:
        return pd.Series(dtype=bool, index=frame.index)
    allowed_groups = frame["region_group"].isin({"成都地区", "川内其他地区"})
    unassigned = frame["service_region"].str.contains(
        r"无区域|待复核|未明确|未分配", regex=True, na=False
    )
    def quality_requires_review(row: pd.Series) -> bool:
        issue = _clean_quality_issue(row["quality_issue"])
        if not issue:
            return False
        # 保险类不设金额门槛，仅“金额缺失”不应把所有保险商机打回复核。
        issue_parts = {part.strip() for part in issue.split("；") if part.strip()}
        if row["business_type"] == "保险" and issue_parts <= {"缺失", "金额缺失"}:
            return False
        return True

    has_quality_issue = frame.apply(quality_requires_review, axis=1)
    known_type = frame["business_type"].isin({"保险", "工程"})
    return frame["selected"] & (~allowed_groups | unassigned | has_quality_issue | ~known_type)


def _allocation_label(row: pd.Series) -> str:
    region_group = _clean_text(row.get("region_group"))
    service_region = _clean_text(row.get("service_region"))
    district = _clean_text(row.get("district"))
    if region_group == "成都地区":
        if service_region and service_region not in {"成都地区", "成都"}:
            return service_region
        return district or "未明确"
    if region_group == "川内其他地区":
        return "川内其他地区"
    if "无区域" in region_group or "无区域" in service_region:
        return "无区域类"
    return "未明确"


def _money_sum(values: Iterable[Any]) -> int | float:
    total = Decimal("0")
    for value in values:
        number = _coerce_amount(value)
        if number is None:
            continue
        total += Decimal(str(number))
    if total == total.to_integral():
        return int(total)
    return float(total.quantize(Decimal("0.01")))


def _format_amount(value: Any) -> str:
    amount = _coerce_amount(value)
    if amount is None:
        return "金额未披露"
    number = Decimal(str(amount))
    if abs(number) >= Decimal("100000000"):
        shown = number / Decimal("100000000")
        display = format(shown.quantize(Decimal("0.01")), "f").rstrip("0").rstrip(".")
        return f"{display}亿元"
    if abs(number) >= Decimal("10000"):
        shown = number / Decimal("10000")
        display = format(shown.quantize(Decimal("0.01")), "f").rstrip("0").rstrip(".")
        return f"{display}万元"
    if number == number.to_integral():
        return f"{int(number):,}元"
    return f"{float(number):,.2f}元"


def _partition(frame: pd.DataFrame) -> dict[str, pd.DataFrame]:
    selected = frame[frame["selected"]].copy()
    review = selected[_review_mask(selected)].copy()
    return {
        "成都地区（保险）": selected[
            (selected["business_type"] == "保险") & (selected["region_group"] == "成都地区")
        ].copy(),
        "川内其他地区（保险）": selected[
            (selected["business_type"] == "保险")
            & (selected["region_group"] == "川内其他地区")
        ].copy(),
        "成都地区（工程）": selected[
            (selected["business_type"] == "工程") & (selected["region_group"] == "成都地区")
        ].copy(),
        "川内其他地区（工程）": selected[
            (selected["business_type"] == "工程")
            & (selected["region_group"] == "川内其他地区")
        ].copy(),
        "未分区域项目": review,
        "筛除记录": frame[~frame["selected"]].copy(),
        "采集与判定审计": frame.copy(),
    }


def _set_cell_style(
    cell: Cell,
    *,
    fill: str | None = None,
    font_color: str = TEXT,
    bold: bool = False,
    size: int = 10,
    horizontal: str = "left",
    vertical: str = "center",
    wrap: bool = True,
    border: bool = True,
) -> None:
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name="Microsoft YaHei", size=size, bold=bold, color=font_color)
    cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)
    if border:
        cell.border = THIN_BORDER


def _style_title(ws: Any, title: str, subtitle: str, max_column: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
    title_cell = ws.cell(1, 1, title)
    _set_cell_style(
        title_cell,
        fill=NAVY,
        font_color=WHITE,
        bold=True,
        size=18,
        horizontal="center",
        border=False,
    )
    ws.row_dimensions[1].height = 34

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_column)
    subtitle_cell = ws.cell(2, 1, subtitle)
    _set_cell_style(
        subtitle_cell,
        fill=LIGHT_BLUE,
        font_color=NAVY,
        size=10,
        horizontal="center",
        border=False,
    )
    ws.row_dimensions[2].height = 24


def _apply_sheet_defaults(ws: Any, *, landscape: bool = True) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddFooter.center.text = "商机推送助手·系统自动生成"
    ws.oddFooter.right.text = "第 &[Page] 页 / 共 &[Pages] 页"
    ws.sheet_properties.outlinePr.summaryBelow = True


def _write_summary_sheet(ws: Any, frame: pd.DataFrame, report_day: date) -> None:
    selected = frame[frame["selected"]].copy()
    selected["_allocation"] = selected.apply(_allocation_label, axis=1)
    insurance = selected[selected["business_type"] == "保险"]
    engineering = selected[selected["business_type"] == "工程"]
    review_count = int(_review_mask(selected).sum())

    _style_title(ws, f"{_date_cn(report_day)}商机信息情况", "保险与工程商机区域分配总览", 12)

    kpis = (
        ("推送商机", len(selected), BLUE),
        ("保险商机", len(insurance), BLUE),
        ("保险金额", _money_sum(insurance["amount"]), GREEN),
        ("工程商机", len(engineering), ORANGE),
        ("工程金额", _money_sum(engineering["amount"]), ORANGE),
        ("未分区域", review_count, GOLD),
    )
    for index, (label, value, color) in enumerate(kpis):
        column = index * 2 + 1
        label_cell = ws.cell(4, column, label)
        value_cell = ws.cell(5, column, value)
        ws.merge_cells(start_row=4, start_column=column, end_row=4, end_column=column + 1)
        ws.merge_cells(start_row=5, start_column=column, end_row=5, end_column=column + 1)
        _set_cell_style(label_cell, fill=color, font_color=WHITE, bold=True, size=10, horizontal="center")
        _set_cell_style(value_cell, fill=WHITE, font_color=color, bold=True, size=16, horizontal="center")
        if "金额" in label:
            value_cell.number_format = '¥#,##0.00'
    ws.row_dimensions[4].height = 23
    ws.row_dimensions[5].height = 34

    headers = (
        "区域",
        "保险项目数",
        "保险招标金额（元）",
        "工程项目数",
        "工程招标金额（元）",
        "合计项目数",
        "合计金额（元）",
    )
    header_row = 8
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col, header)
        _set_cell_style(cell, fill=NAVY, font_color=WHITE, bold=True, horizontal="center")
    ws.row_dimensions[header_row].height = 30

    actual_labels = [_clean_text(value) for value in selected["_allocation"].unique() if _clean_text(value)]
    extra_labels = sorted(label for label in actual_labels if label not in DEFAULT_REGION_ORDER)
    region_labels = list(DEFAULT_REGION_ORDER) + extra_labels

    first_data_row = header_row + 1
    for row_index, label in enumerate(region_labels, first_data_row):
        group = selected[selected["_allocation"] == label]
        ins = group[group["business_type"] == "保险"]
        eng = group[group["business_type"] == "工程"]
        values: tuple[Any, ...] = (
            label,
            len(ins),
            _money_sum(ins["amount"]),
            len(eng),
            _money_sum(eng["amount"]),
            f"=SUM(B{row_index},D{row_index})",
            f"=SUM(C{row_index},E{row_index})",
        )
        for col, value in enumerate(values, 1):
            cell = ws.cell(row_index, col, value)
            fill = LIGHT_BLUE if row_index % 2 else WHITE
            if label in {"未明确", "无区域类"}:
                fill = LIGHT_GOLD
            _set_cell_style(cell, fill=fill, horizontal="center" if col != 1 else "left")
            if col in {3, 5, 7}:
                cell.number_format = '#,##0.00;[Red]-#,##0.00;0'

    total_row = first_data_row + len(region_labels)
    ws.cell(total_row, 1, "总计")
    for col in range(2, 8):
        letter = get_column_letter(col)
        ws.cell(total_row, col, f"=SUM({letter}{first_data_row}:{letter}{total_row - 1})")
    for col in range(1, 8):
        cell = ws.cell(total_row, col)
        _set_cell_style(cell, fill=NAVY, font_color=WHITE, bold=True, horizontal="center")
        if col in {3, 5, 7}:
            cell.number_format = '#,##0.00;[Red]-#,##0.00;0'
    ws.row_dimensions[total_row].height = 26

    note_row = total_row + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 1, end_column=7)
    note = ws.cell(
        note_row,
        1,
        "说明：金额均按元汇总；未披露金额不按 0 元伪造，仅不计入金额合计。"
        "“无区域类/未明确”项目请在推送前完成内部确认。",
    )
    _set_cell_style(note, fill=LIGHT_GOLD, font_color=TEXT, size=10, vertical="top")
    ws.row_dimensions[note_row].height = 35

    widths = {"A": 20, "B": 13, "C": 22, "D": 13, "E": 22, "F": 13, "G": 22}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    # KPI 区域使用到 H:L，适度留白。
    for column in "HIJKL":
        ws.column_dimensions[column].width = 12
    ws.freeze_panes = "A9"
    ws.auto_filter.ref = f"A{header_row}:G{total_row - 1}"
    ws.print_title_rows = f"1:{header_row}"
    _apply_sheet_defaults(ws)
    ws.freeze_panes = "A9"


def _write_detail_sheet(ws: Any, title: str, frame: pd.DataFrame, report_day: date) -> None:
    compact = title in BUSINESS_DETAIL_SHEETS
    columns = (
        BUSINESS_DETAIL_COLUMNS
        if compact
        else AUDIT_COLUMNS
        if title == "采集与判定审计"
        else DETAIL_COLUMNS
    )
    subtitle = f"报告日期：{_date_cn(report_day)}｜记录数：{len(frame)}｜金额合计：{_format_amount(_money_sum(frame['amount']))}"
    _style_title(ws, title, subtitle, len(columns))

    ws.cell(3, 1, "操作提示")
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=len(columns))
    prompt = (
        "本页为业务分发精简表；请核对截止时间和联系方式后发送。"
        if compact
        else "请先确认黄色/红色数据问题，再将该页商机发送至企业微信群。"
    )
    ws.cell(3, 2, prompt)
    _set_cell_style(ws.cell(3, 1), fill=LIGHT_GOLD, bold=True, horizontal="center")
    _set_cell_style(ws.cell(3, 2), fill=LIGHT_GOLD, font_color=TEXT)
    ws.row_dimensions[3].height = 24

    header_row = 4
    for column_index, (_, display_name) in enumerate(columns, 1):
        cell = ws.cell(header_row, column_index, display_name)
        _set_cell_style(cell, fill=NAVY, font_color=WHITE, bold=True, horizontal="center")
    ws.row_dimensions[header_row].height = 32

    if frame.empty:
        ws.merge_cells(start_row=5, start_column=1, end_row=6, end_column=len(columns))
        empty_cell = ws.cell(5, 1, "今日暂无符合条件的记录")
        _set_cell_style(empty_cell, fill=LIGHT_GREY, font_color=GREY, size=13, horizontal="center")
        ws.row_dimensions[5].height = 42
        last_row = header_row
    else:
        for output_row, (_, record) in enumerate(frame.iterrows(), header_row + 1):
            sequence = output_row - header_row
            for output_col, (field, _) in enumerate(columns, 1):
                if field == "_sequence":
                    value: Any = sequence
                elif field == "_remarks":
                    # 正式分发页只展示业务老师主动填写的推送备注；
                    # 自动判定依据和数据质量信息保留在内部审计页。
                    value = _clean_text(record["note"])
                else:
                    value = record[field]
                if field == "selected":
                    value = "是" if _coerce_bool(value) else "否"
                elif field in {"publish_date", "deadline"}:
                    value = _display_date(value)
                elif isinstance(value, str):
                    value = _formal_text(value)
                cell = ws.cell(output_row, output_col, value)

                fill = WHITE if output_row % 2 else "F8FAFC"
                if field == "business_type":
                    if record["business_type"] == "保险":
                        fill = LIGHT_BLUE
                    elif record["business_type"] == "工程":
                        fill = LIGHT_ORANGE
                if field == "quality_issue" and _is_nonempty(value):
                    fill = LIGHT_RED if re.search(r"异常|错误|失败|冲突", _clean_text(value)) else LIGHT_GOLD
                if title == "筛除记录":
                    fill = LIGHT_GREY
                if title == "未分区域项目" and field in {
                    "region_group",
                    "service_region",
                    "quality_issue",
                }:
                    fill = LIGHT_GOLD

                alignment = "center" if field in {
                    "_sequence",
                    "selected",
                    "business_type",
                    "category",
                    "publish_date",
                    "city",
                    "district",
                    "region_group",
                    "service_region",
                    "stage",
                } else "left"
                _set_cell_style(cell, fill=fill, horizontal=alignment, vertical="top")
                if field == "amount":
                    cell.number_format = '#,##0.00;[Red]-#,##0.00;0'
                    cell.alignment = Alignment(horizontal="right", vertical="top")
                elif field in {"publish_date", "deadline"} and isinstance(value, (datetime, date)):
                    if field == "publish_date":
                        cell.number_format = "yyyy-mm-dd"
                    elif isinstance(value, datetime) and value.time() != datetime.min.time():
                        cell.number_format = "yyyy-mm-dd hh:mm"
                    else:
                        cell.number_format = "yyyy-mm-dd"
                elif field == "url" and _clean_text(value).startswith(("http://", "https://")):
                    cell.hyperlink = _clean_text(value)
                    cell.font = Font(name="Microsoft YaHei", size=9, color="0563C1", underline="single")
            ws.row_dimensions[output_row].height = 50
        last_row = header_row + len(frame)

    if compact:
        widths = {
            1: 8,
            2: 16,
            3: 48,
            4: 15,
            5: 18,
            6: 13,
            7: 13,
            8: 16,
            9: 15,
            10: 23,
            11: 26,
            12: 24,
            13: 25,
            14: 34,
            15: 36,
        }
    elif title == "采集与判定审计":
        widths = {
            1: 11,
            2: 12,
            3: 12,
            4: 48,
            5: 24,
            6: 24,
            7: 20,
            8: 18,
            9: 16,
            10: 38,
            11: 16,
            12: 44,
            13: 13,
            14: 13,
            15: 20,
            16: 40,
            17: 40,
            18: 25,
            19: 34,
        }
    else:
        widths = {
            1: 8,
            2: 11,
            3: 16,
            4: 48,
            5: 15,
            6: 18,
            7: 13,
            8: 13,
            9: 17,
            10: 16,
            11: 15,
            12: 21,
            13: 25,
            14: 22,
            15: 25,
            16: 34,
            17: 34,
            18: 30,
            19: 20,
        }
    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{max(header_row, last_row)}"
    ws.freeze_panes = "D5" if compact else "E5"
    ws.print_title_rows = "1:4"
    _apply_sheet_defaults(ws)
    ws.freeze_panes = "D5" if compact else "E5"


def _normalise_logs(
    processing_log: pd.DataFrame | Iterable[str | Mapping[str, Any]] | None,
    frame: pd.DataFrame,
) -> list[dict[str, Any]]:
    logs: list[dict[str, Any]] = []
    if isinstance(processing_log, pd.DataFrame):
        items: Iterable[Any] = processing_log.to_dict("records")
    else:
        items = processing_log or []

    for item in items:
        if isinstance(item, Mapping):
            logs.append(
                {
                    "time": item.get("time") or item.get("timestamp") or item.get("时间") or "",
                    "level": item.get("level") or item.get("级别") or "INFO",
                    "stage": item.get("stage") or item.get("阶段") or item.get("step") or "处理",
                    "message": item.get("message")
                    or item.get("log")
                    or item.get("content")
                    or item.get("日志内容")
                    or "",
                }
            )
        else:
            logs.append({"time": "", "level": "INFO", "stage": "处理", "message": _clean_text(item)})

    selected_count = int(frame["selected"].sum())
    review_count = int(_review_mask(frame).sum())
    logs.extend(
        [
            {
                "time": "",
                "level": "INFO",
                "stage": "报表汇总",
                "message": f"共接收 {len(frame)} 条记录，纳入推送 {selected_count} 条，筛除 {len(frame) - selected_count} 条。",
            },
            {
                "time": "",
                "level": "WARNING" if review_count else "INFO",
                "stage": "质量校验",
                "message": f"未分区域或需人工确认记录 {review_count} 条。",
            },
        ]
    )
    return logs


def _write_log_sheet(
    ws: Any,
    frame: pd.DataFrame,
    processing_log: pd.DataFrame | Iterable[str | Mapping[str, Any]] | None,
    report_day: date,
) -> None:
    _style_title(ws, "处理日志", f"报告日期：{_date_cn(report_day)}｜用于追溯本次筛选与生成过程", 4)
    headers = ("时间", "级别", "处理阶段", "日志内容")
    for column, header in enumerate(headers, 1):
        cell = ws.cell(4, column, header)
        _set_cell_style(cell, fill=NAVY, font_color=WHITE, bold=True, horizontal="center")

    logs = _normalise_logs(processing_log, frame)
    for row_index, item in enumerate(logs, 5):
        values = (item["time"], item["level"], item["stage"], item["message"])
        level = _clean_text(item["level"]).upper()
        fill = LIGHT_RED if level in {"ERROR", "CRITICAL", "错误"} else LIGHT_GOLD if level in {
            "WARNING",
            "WARN",
            "警告",
        } else WHITE if row_index % 2 else "F8FAFC"
        for column, value in enumerate(values, 1):
            shown_value = _display_date(value) if column == 1 else value
            if isinstance(shown_value, str):
                shown_value = _formal_text(shown_value)
            cell = ws.cell(row_index, column, shown_value)
            _set_cell_style(cell, fill=fill, vertical="top")
            if column == 1 and isinstance(cell.value, (date, datetime)):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
        ws.row_dimensions[row_index].height = 30

    widths = (22, 13, 22, 100)
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.auto_filter.ref = f"A4:D{max(4, 4 + len(logs))}"
    ws.freeze_panes = "A5"
    ws.print_title_rows = "1:4"
    _apply_sheet_defaults(ws)


def build_opportunity_excel(
    data: pd.DataFrame | Iterable[Mapping[str, Any]] | None,
    *,
    report_date: date | datetime | str | None = None,
    processing_log: pd.DataFrame | Iterable[str | Mapping[str, Any]] | None = None,
) -> io.BytesIO:
    """在内存中生成可直接下载的专业 Excel 商机报告。"""

    frame = normalize_report_dataframe(data)
    report_day = _report_date(report_date)
    partitions = _partition(frame)

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.properties.title = f"{_date_cn(report_day)}商机推送报告"
    workbook.properties.subject = "保险与工程商机区域分配"
    workbook.properties.creator = "商机推送助手"

    summary = workbook.create_sheet("今日商机汇总")
    _write_summary_sheet(summary, frame, report_day)
    for sheet_name in SHEET_NAMES[1:-1]:
        sheet = workbook.create_sheet(sheet_name)
        _write_detail_sheet(sheet, sheet_name, partitions[sheet_name], report_day)
    log_sheet = workbook.create_sheet("处理日志")
    _write_log_sheet(log_sheet, frame, processing_log, report_day)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


# 兼容更短的调用名，也便于从原标书项目中平滑迁移。
build_excel_report = build_opportunity_excel


@lru_cache(maxsize=1)
def _discover_chinese_font_path() -> Path | None:
    """自动发现 Windows/Linux/macOS 上可用的中文字体。"""

    search_roots: list[Path] = []
    windows_root = os.environ.get("WINDIR") or os.environ.get("SystemRoot")
    if windows_root:
        search_roots.append(Path(windows_root) / "Fonts")
    local_app_data = os.environ.get("LOCALAPPDATA")
    if local_app_data:
        search_roots.append(Path(local_app_data) / "Microsoft" / "Windows" / "Fonts")
    search_roots.extend(
        [
            Path("/usr/share/fonts"),
            Path("/usr/local/share/fonts"),
            Path("/System/Library/Fonts"),
            Path("/Library/Fonts"),
        ]
    )
    preferred_names = (
        "msyh.ttc",
        "msyhbd.ttc",
        "simhei.ttf",
        "simsun.ttc",
        "deng.ttf",
        "dengb.ttf",
        "NotoSansCJK-Regular.ttc",
        "NotoSansCJKsc-Regular.otf",
        "SourceHanSansCN-Regular.otf",
        "PingFang.ttc",
        "Arial Unicode.ttf",
        "DejaVuSans.ttf",
    )
    for root in search_roots:
        if not root.exists():
            continue
        for name in preferred_names:
            direct = root / name
            if direct.is_file():
                return direct
        preferred_lower = {name.lower() for name in preferred_names}
        try:
            for candidate in root.rglob("*"):
                if candidate.is_file() and candidate.name.lower() in preferred_lower:
                    return candidate
        except OSError:
            continue
    return None


@lru_cache(maxsize=32)
def _load_image_font(size: int, *, bold: bool = False) -> Any:
    from PIL import ImageFont

    discovered = _discover_chinese_font_path()
    if discovered:
        candidate = discovered
        if bold:
            siblings = {
                "msyh.ttc": "msyhbd.ttc",
                "deng.ttf": "dengb.ttf",
                "NotoSansCJK-Regular.ttc": "NotoSansCJK-Bold.ttc",
                "NotoSansCJKsc-Regular.otf": "NotoSansCJKsc-Bold.otf",
                "SourceHanSansCN-Regular.otf": "SourceHanSansCN-Bold.otf",
            }
            bold_name = siblings.get(discovered.name)
            if bold_name and (discovered.parent / bold_name).is_file():
                candidate = discovered.parent / bold_name
        try:
            return ImageFont.truetype(str(candidate), size=size)
        except OSError:
            pass
    # Pillow 的新版默认字体接受 size；旧版不支持时再降级。
    try:
        return ImageFont.load_default(size=size)
    except TypeError:
        return ImageFont.load_default()


def _wrap_for_image(draw: Any, text: Any, font: Any, max_width: int) -> list[str]:
    content = _clean_text(text)
    if not content:
        return [""]
    lines: list[str] = []
    current = ""
    for char in content:
        if char == "\n":
            lines.append(current)
            current = ""
            continue
        candidate = current + char
        width = draw.textbbox((0, 0), candidate, font=font)[2]
        if current and width > max_width:
            lines.append(current)
            current = char
        else:
            current = candidate
    if current or not lines:
        lines.append(current)
    return lines


def _render_multiline(
    draw: Any,
    xy: tuple[int, int],
    lines: Sequence[str],
    font: Any,
    fill: str,
    line_height: int,
) -> int:
    x, y = xy
    for line in lines:
        draw.text((x, y), line, font=font, fill=fill)
        y += line_height
    return y


def _safe_image_text(value: Any) -> str:
    text = _clean_text(value)
    return text.replace("\u200b", "").replace("\ufeff", "")


def build_chengdu_opportunity_png(
    data: pd.DataFrame | Iterable[Mapping[str, Any]] | None,
    *,
    report_date: date | datetime | str | None = None,
) -> io.BytesIO:
    """生成企业微信可直接发送的成都地区清晰长图。"""

    from PIL import Image, ImageDraw

    frame = normalize_report_dataframe(data)
    report_day = _report_date(report_date)
    chengdu = frame[
        frame["selected"] & (frame["region_group"] == "成都地区")
    ].copy()
    if not chengdu.empty:
        chengdu["_allocation"] = chengdu.apply(_allocation_label, axis=1)
        chengdu["_type_order"] = chengdu["business_type"].map({"保险": 0, "工程": 1}).fillna(2)
        chengdu = chengdu.sort_values(
            by=["_allocation", "_type_order", "publish_date", "project_name"],
            kind="stable",
        )

    width = 1440
    margin = 56
    inner_width = width - margin * 2
    title_font = _load_image_font(48, bold=True)
    subtitle_font = _load_image_font(24)
    count_font = _load_image_font(28, bold=True)
    tag_font = _load_image_font(23, bold=True)
    name_font = _load_image_font(31, bold=True)
    meta_font = _load_image_font(23)
    small_font = _load_image_font(20)

    # 先用临时画布计算每张卡片高度，避免长标题被截断。
    probe = Image.new("RGB", (width, 200), "white")
    probe_draw = ImageDraw.Draw(probe)
    cards: list[tuple[pd.Series, list[str], list[str], int]] = []
    for _, record in chengdu.iterrows():
        title_lines = _wrap_for_image(probe_draw, record["project_name"], name_font, inner_width - 72)
        tenderer_lines = _wrap_for_image(
            probe_draw,
            f"招标人/采购人：{_safe_image_text(record['tenderer']) or '未披露'}",
            small_font,
            inner_width - 72,
        )
        card_height = 40 + 34 + 18 + len(title_lines) * 43 + 18 + 34 + len(tenderer_lines) * 30 + 30
        cards.append((record, title_lines, tenderer_lines, max(174, card_height)))

    header_height = 230
    empty_height = 250 if chengdu.empty else 0
    footer_height = 92
    card_gap = 22
    total_height = header_height + empty_height + sum(card[3] + card_gap for card in cards) + footer_height
    total_height = max(540, total_height)

    image = Image.new("RGB", (width, total_height), "#F3F6FA")
    draw = ImageDraw.Draw(image)
    draw.rectangle((0, 0, width, 176), fill="#17365D")
    draw.text((margin, 35), f"{_date_cn(report_day)} 成都地区商机", font=title_font, fill="white")
    insurance_count = int((chengdu["business_type"] == "保险").sum())
    engineering_count = int((chengdu["business_type"] == "工程").sum())
    draw.text(
        (margin, 112),
        f"共 {len(chengdu)} 条｜保险 {insurance_count} 条｜工程 {engineering_count} 条",
        font=subtitle_font,
        fill="#D9EAF7",
    )
    draw.rounded_rectangle((margin, 194, width - margin, 214), radius=10, fill="#DDE7F2")
    if len(chengdu):
        insurance_amount = _money_sum(chengdu.loc[chengdu["business_type"] == "保险", "amount"])
        engineering_amount = _money_sum(chengdu.loc[chengdu["business_type"] == "工程", "amount"])
        summary_text = f"保险金额 {_format_amount(insurance_amount)}    工程金额 {_format_amount(engineering_amount)}"
    else:
        summary_text = "今日暂无成都地区已入选商机"
    summary_width = draw.textbbox((0, 0), summary_text, font=count_font)[2]
    draw.text(((width - summary_width) // 2, 183), summary_text, font=count_font, fill="#17365D")

    y = header_height
    if chengdu.empty:
        draw.rounded_rectangle(
            (margin, y + 20, width - margin, y + 190),
            radius=20,
            fill="white",
            outline="#CBD5E1",
            width=2,
        )
        empty_text = "今日暂无成都地区商机"
        text_width = draw.textbbox((0, 0), empty_text, font=name_font)[2]
        draw.text(((width - text_width) // 2, y + 82), empty_text, font=name_font, fill="#64748B")
        y += empty_height
    else:
        for record, title_lines, tenderer_lines, card_height in cards:
            card_top = y
            card_bottom = y + card_height
            draw.rounded_rectangle(
                (margin, card_top, width - margin, card_bottom),
                radius=18,
                fill="white",
                outline="#D8E0EA",
                width=2,
            )
            type_color = "#2F75B5" if record["business_type"] == "保险" else "#ED7D31"
            draw.rounded_rectangle((margin, card_top, margin + 14, card_bottom), radius=7, fill=type_color)

            allocation = _safe_image_text(_allocation_label(record))
            category = _safe_image_text(record["category"]) or "待分类"
            tag_text = f"{record['business_type']}  ·  {category}  ·  {allocation}"
            draw.text((margin + 36, card_top + 26), tag_text, font=tag_font, fill=type_color)

            text_y = card_top + 77
            text_y = _render_multiline(
                draw,
                (margin + 36, text_y),
                title_lines,
                name_font,
                "#182230",
                43,
            )
            deadline = _safe_image_text(record["deadline"]) or "未披露"
            stage = _safe_image_text(record["stage"]) or "阶段未标注"
            meta_text = f"金额：{_format_amount(record['amount'])}    截止：{deadline}    阶段：{stage}"
            draw.text((margin + 36, text_y + 6), meta_text, font=meta_font, fill="#41546A")
            _render_multiline(
                draw,
                (margin + 36, text_y + 49),
                tenderer_lines,
                small_font,
                "#64748B",
                30,
            )
            y = card_bottom + card_gap

    footer_y = total_height - footer_height
    draw.line((margin, footer_y, width - margin, footer_y), fill="#CBD5E1", width=2)
    footer = "注：金额未披露不等于 0 元；未分区域或数据问题请先完成内部确认。"
    draw.text((margin, footer_y + 26), footer, font=small_font, fill="#64748B")

    output = io.BytesIO()
    image.save(output, format="PNG", optimize=True, dpi=(144, 144))
    output.seek(0)
    return output


build_chengdu_png = build_chengdu_opportunity_png


def _text_date(value: Any) -> str:
    shown = _display_date(value)
    if isinstance(shown, datetime):
        return shown.strftime("%Y-%m-%d %H:%M")
    if isinstance(shown, date):
        return shown.strftime("%Y-%m-%d")
    return _clean_text(shown) or "未披露"


def _sort_for_message(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame.copy()
    sorted_frame = frame.copy()
    sorted_frame["_allocation"] = sorted_frame.apply(_allocation_label, axis=1)
    sorted_frame["_type_order"] = sorted_frame["business_type"].map({"保险": 0, "工程": 1}).fillna(2)
    return sorted_frame.sort_values(
        by=["_allocation", "_type_order", "publish_date", "project_name"],
        kind="stable",
    )


def _build_concise_wecom_message(frame: pd.DataFrame, day: date) -> str:
    """生成贴合部门现有发送习惯的短文案。

    群内文字只点名需要分配的成都保险商机；数量较多的工程和川内其他
    项目交由随附的成都图片与 Excel 承载，避免几十个标题刷屏。
    """

    selected = _sort_for_message(frame[frame["selected"]])
    if selected.empty:
        return "\n".join(
            [
                f"今日商机｜{_date_short(day)}",
                "今日暂无符合条件的可推送商机。",
            ]
        )

    insurance = selected[selected["business_type"] == "保险"]
    engineering = selected[selected["business_type"] == "工程"]
    chengdu = selected[selected["region_group"] == "成都地区"]
    other = selected.drop(index=chengdu.index)
    chengdu_insurance = insurance[insurance["region_group"] == "成都地区"]
    unassigned_mask = _review_mask(chengdu_insurance)
    assigned_insurance = chengdu_insurance.drop(index=chengdu_insurance[unassigned_mask].index)
    unassigned_insurance = chengdu_insurance[unassigned_mask]

    lines = [
        f"今日商机｜{_date_short(day)}",
        (
            f"今日共纳入 {len(selected)} 条：保险 {len(insurance)} 条，工程 {len(engineering)} 条；"
            f"成都地区 {len(chengdu)} 条，川内其他地区 {len(other)} 条。"
        ),
    ]

    if not assigned_insurance.empty:
        lines.extend(["", "已分区域项目："])
        for item_no, (_, record) in enumerate(assigned_insurance.iterrows(), 1):
            lines.append(f"{item_no}. {record['project_name']}（{_allocation_label(record)}）")

    if not unassigned_insurance.empty:
        lines.extend(["", "未分区域项目："])
        for item_no, (_, record) in enumerate(unassigned_insurance.iterrows(), 1):
            lines.append(f"{item_no}. {record['project_name']}")

    lines.extend(
        [
            "",
            "烦请各机构专员推送至对应机构/团队，并确认参与情况。",
            "工程及川内其他地区项目明细见随附 Excel，成都地区汇总见随附图片。",
        ]
    )
    return "\n".join(lines)


def build_wecom_text(
    data: pd.DataFrame | Iterable[Mapping[str, Any]] | None,
    *,
    report_date: date | datetime | str | None = None,
    detailed: bool = False,
) -> str:
    """生成可直接复制到企业微信群的中文文案。"""

    frame = normalize_report_dataframe(data)
    day = _report_date(report_date)
    selected = _sort_for_message(frame[frame["selected"]])
    excluded_count = len(frame) - len(selected)
    insurance = selected[selected["business_type"] == "保险"]
    engineering = selected[selected["business_type"] == "工程"]
    review = selected[_review_mask(selected)]
    assigned = selected.drop(index=review.index)

    lines = [
        f"今日商机｜{_date_short(day)}",
        f"已筛选 {len(selected)} 条：保险 {len(insurance)} 条，工程 {len(engineering)} 条；筛除 {excluded_count} 条。",
    ]
    if len(review):
        lines.append(f"其中已分配 {len(assigned)} 条，未分区域项目 {len(review)} 条。")

    if selected.empty:
        lines.append("今日暂无符合条件的可推送商机。")
        return "\n".join(lines)

    if not detailed:
        return _build_concise_wecom_message(frame, day)

    for region_name, region_frame in (
        ("成都地区", assigned[assigned["region_group"] == "成都地区"]),
        ("川内其他地区", assigned[assigned["region_group"] == "川内其他地区"]),
    ):
        if region_frame.empty:
            continue
        lines.extend(["", f"【{region_name}】共 {len(region_frame)} 条"])
        for item_no, (_, record) in enumerate(region_frame.iterrows(), 1):
            allocation = _allocation_label(record)
            prefix = f"{item_no}. [{allocation}/{record['business_type']}]"
            if detailed:
                lines.append(f"{prefix}{record['project_name']}")
                lines.append(
                    f"   分类：{record['category'] or '待分类'}｜金额：{_format_amount(record['amount'])}"
                    f"｜阶段：{record['stage'] or '未标注'}"
                )
                lines.append(
                    f"   截止：{_text_date(record['deadline'])}｜招标人/采购人：{record['tenderer'] or '未披露'}"
                )
                if record["contact"]:
                    lines.append(f"   联系：{record['contact']}")
                if record["url"]:
                    lines.append(f"   原文：{record['url']}")
            else:
                amount = _format_amount(record["amount"])
                lines.append(f"{prefix}{record['project_name']}（{amount}）")

    if not review.empty:
        lines.extend(
            [
                "",
                f"【未分区域项目】共 {len(review)} 条",
                "注：以下项目暂未匹配服务区域，烦请各机构专员协调确认承接机构/团队。",
            ]
        )
        for item_no, (_, record) in enumerate(review.iterrows(), 1):
            if detailed:
                lines.append(f"{item_no}. [{record['business_type']}]{record['project_name']}")
                lines.append(
                    f"   分类：{record['category'] or '待分类'}｜金额：{_format_amount(record['amount'])}"
                    f"｜阶段：{record['stage'] or '未标注'}"
                )
                lines.append(
                    f"   截止：{_text_date(record['deadline'])}｜招标人/采购人：{record['tenderer'] or '未披露'}"
                )
                if record["contact"]:
                    lines.append(f"   联系：{record['contact']}")
                if record["url"]:
                    lines.append(f"   原文：{record['url']}")
            else:
                lines.append(f"{item_no}. {record['project_name']}（{_format_amount(record['amount'])}）")
    return "\n".join(lines)


def build_wecom_messages(
    data: pd.DataFrame | Iterable[Mapping[str, Any]] | None,
    *,
    report_date: date | datetime | str | None = None,
) -> tuple[str, str]:
    """同时返回简洁版和完整版群发文字。"""

    return (
        build_wecom_text(data, report_date=report_date, detailed=False),
        build_wecom_text(data, report_date=report_date, detailed=True),
    )


def build_report_bundle(
    data: pd.DataFrame | Iterable[Mapping[str, Any]] | None,
    *,
    report_date: date | datetime | str | None = None,
    processing_log: pd.DataFrame | Iterable[str | Mapping[str, Any]] | None = None,
) -> ReportBundle:
    """一次性生成前端下载/复制需要的全部交付物。"""

    excel = build_opportunity_excel(
        data,
        report_date=report_date,
        processing_log=processing_log,
    )
    png = build_chengdu_opportunity_png(data, report_date=report_date)
    concise, full = build_wecom_messages(data, report_date=report_date)
    return ReportBundle(excel=excel, png=png, concise_text=concise, full_text=full)
