from __future__ import annotations

import io
import zipfile
from datetime import date

import pandas as pd
from openpyxl import load_workbook
from PIL import Image

from opportunity_assistant.reporting import (
    REQUIRED_COLUMNS,
    SHEET_NAMES,
    build_opportunity_excel,
    build_report_bundle,
    normalize_report_dataframe,
)


def _sample_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "selected": True,
                "business_type": "保险",
                "category": "责任险",
                "project_name": "成都市金牛区雇主责任险采购项目",
                "publish_date": "2026-08-26",
                "amount": 85_000,
                "city": "成都市",
                "district": "金牛区",
                "region_group": "成都地区",
                "service_region": "金牛区",
                "stage": "招标公告",
                "deadline": "2026-09-01 10:00",
                "tenderer": "金牛区某医疗机构",
                "contact": "王老师 028-12345678",
                "agent": "四川某代理公司",
                "url": "https://example.com/insurance-1",
                "decision_reason": "命中责任险业务范围",
                "quality_issue": "",
                "project_key": "insurance-1",
            },
            {
                "selected": True,
                "business_type": "保险",
                "category": "责任险",
                "project_name": "成都武侯宾馆雇员忠诚险及现金险项目",
                "publish_date": "2026-08-26",
                "amount": 500,
                "city": "成都市",
                "district": "武侯区",
                "region_group": "成都地区",
                "service_region": "无区域类",
                "stage": "评选公告",
                "deadline": "",
                "tenderer": "成都某酒店",
                "contact": "",
                "agent": "",
                "url": "https://example.com/insurance-2",
                "decision_reason": "险种匹配，但该区无服务点",
                "quality_issue": "无区域可分配",
                "project_key": "insurance-2",
            },
            {
                "selected": True,
                "business_type": "保险",
                "category": "工程险",
                "project_name": "宜宾市某工程一切险项目",
                "publish_date": "2026-08-26",
                "amount": "150万元",
                "city": "宜宾市",
                "district": "叙州区",
                "region_group": "川内其他地区",
                "service_region": "宜宾机构",
                "stage": "竞争性磋商",
                "deadline": "2026-09-03",
                "tenderer": "宜宾某工程公司",
                "contact": "",
                "agent": "",
                "url": "https://example.com/insurance-3",
                "decision_reason": "命中工程险",
                "quality_issue": "",
                "project_key": "insurance-3",
            },
            {
                "selected": True,
                "business_type": "工程",
                "category": "市政施工",
                "project_name": "泸州市江阳区老旧街区改造工程",
                "publish_date": "2026-08-25",
                "amount": 20_000_000,
                "city": "泸州市",
                "district": "江阳区",
                "region_group": "川内其他地区",
                "service_region": "泸州机构",
                "stage": "公开招标",
                "deadline": "2026-09-15 09:30",
                "tenderer": "泸州某项目业主",
                "contact": "",
                "agent": "泸州某工程咨询公司",
                "url": "https://example.com/engineering-1",
                "decision_reason": "施工项目且金额超过一千万元",
                "quality_issue": "",
                "project_key": "engineering-1",
            },
            {
                "selected": False,
                "business_type": "保险",
                "category": "无关物料",
                "project_name": "某公司保险孔螺钉采购",
                "publish_date": "2026-08-26",
                "amount": 2_000_000,
                "city": "德阳市",
                "district": "",
                "region_group": "川内其他地区",
                "service_region": "德阳机构",
                "stage": "询价",
                "deadline": "",
                "tenderer": "某机械公司",
                "contact": "",
                "agent": "",
                "url": "https://example.com/excluded",
                "decision_reason": "保险仅是零件名称，不是险种",
                "quality_issue": "",
                "project_key": "excluded-1",
            },
        ]
    )


def _find_summary_row(sheet, label: str) -> int:
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row, 1).value == label:
            return row
    raise AssertionError(f"汇总表未找到区域：{label}")


def test_excel_contains_all_required_sheets_and_reconciles_counts_and_amounts() -> None:
    output = build_opportunity_excel(
        _sample_frame(),
        report_date=date(2026, 8, 26),
        processing_log=[
            {"level": "INFO", "stage": "导入", "message": "两份源文件读取完成"},
            {"level": "WARNING", "stage": "质量", "message": "待复核项目 1 条"},
        ],
    )

    assert isinstance(output, io.BytesIO)
    assert output.tell() == 0
    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == list(SHEET_NAMES)

    summary = workbook["今日商机汇总"]
    # 汇总页复刻部门现用的五列表：两级表头、真实区县、保险/工程各两列。
    assert summary.max_column == 5
    assert summary.merged_cells.ranges
    assert [summary.cell(2, column).value for column in range(1, 6)] == [
        "区域",
        "保险类",
        None,
        "工程类",
        None,
    ]
    assert [summary.cell(3, column).value for column in range(1, 6)] == [
        None,
        "项目数",
        "招标金额（元）",
        "项目数",
        "招标金额（元）",
    ]
    assert summary.freeze_panes == "A4"

    jin_niu_row = _find_summary_row(summary, "金牛区")
    assert summary.cell(jin_niu_row, 2).value == 1
    assert summary.cell(jin_niu_row, 3).value == 85_000

    # 武侯区虽然无我司服务点，统计仍必须落在真实项目区县，不能落入“未明确”。
    wuhou_row = _find_summary_row(summary, "武侯区")
    assert summary.cell(wuhou_row, 2).value == 1
    assert summary.cell(wuhou_row, 3).value == 500

    other_row = _find_summary_row(summary, "川内其他地区")
    assert summary.cell(other_row, 2).value == 1
    assert summary.cell(other_row, 3).value == 1_500_000
    assert summary.cell(other_row, 4).value == 1
    assert summary.cell(other_row, 5).value == 20_000_000

    total_row = _find_summary_row(summary, "总计")
    assert sum(summary.cell(row, 2).value for row in range(4, total_row)) == 3
    assert sum(summary.cell(row, 3).value for row in range(4, total_row)) == 1_585_500
    assert sum(summary.cell(row, 4).value for row in range(4, total_row)) == 1
    assert sum(summary.cell(row, 5).value for row in range(4, total_row)) == 20_000_000
    assert summary.cell(total_row, 2).value == f"=SUM(B4:B{total_row - 1})"
    assert summary.cell(total_row, 5).value == f"=SUM(E4:E{total_row - 1})"

    # 无区域项仍保留在成都保险明细中，同时进入未分区域页。
    assert workbook["成都地区（保险）"].max_row == 6
    assert workbook["未分区域项目"].max_row == 5
    assert workbook["筛除记录"].max_row == 5
    assert workbook["处理日志"]["D5"].value == "两份源文件读取完成"

    compact_headers = [
        "序号",
        "项目地区",
        "分配区域",
        "险种/项目类型",
        "项目名称",
        "招标金额（元）",
        "截止时间",
        "招标人/采购人",
        "商机关键要点",
        "联系人/联系方式",
        "原文链接",
    ]
    for sheet_name in (
        "成都地区（保险）",
        "川内其他地区（保险）",
        "成都地区（工程）",
        "川内其他地区（工程）",
    ):
        sheet = workbook[sheet_name]
        assert sheet.max_column == 11
        assert [sheet.cell(4, column).value for column in range(1, 12)] == compact_headers
        assert sheet.freeze_panes == "E5"

    # 审计信息仍保留供内部追溯，但默认全部隐藏，正式打开只看汇总与业务子表。
    assert workbook["未分区域项目"].max_column == 19
    assert workbook["筛除记录"].max_column == 19
    for sheet_name in ("未分区域项目", "筛除记录", "采集与判定审计", "处理日志"):
        assert workbook[sheet_name].sheet_state == "hidden"
    for sheet_name in (
        "今日商机汇总",
        "成都地区（保险）",
        "川内其他地区（保险）",
        "成都地区（工程）",
        "川内其他地区（工程）",
    ):
        assert workbook[sheet_name].sheet_state == "visible"

    chengdu_insurance = workbook["成都地区（保险）"]
    assert chengdu_insurance["A5"].value == 1
    assert chengdu_insurance["B5"].value == "成都市-金牛区"
    assert chengdu_insurance["F5"].number_format == '#,##0.00;[Red]-#,##0.00;0'
    assert chengdu_insurance["G5"].number_format == "yyyy-mm-dd hh:mm"

    # 正式工作簿不暴露内部引擎的“待复核”术语，包括处理日志。
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    assert "待复核" not in cell.value


def test_png_is_a_clear_long_image_and_messages_are_ready_to_copy() -> None:
    frame = _sample_frame()
    bundle = build_report_bundle(frame, report_date="2026-08-26")
    image = Image.open(bundle.png)

    assert image.format == "PNG"
    assert image.width == 1440
    assert image.height >= 700

    concise, full = bundle.concise_text, bundle.full_text
    assert "今日商机（2026年8月26日）" in concise
    assert "【已分区域项目｜1条】" in concise
    assert "【金牛区｜责任险】" in concise
    assert "成都市金牛区雇主责任险采购项目" in concise
    assert "【无我司服务点区域｜1条】" in concise
    assert "【武侯区｜责任险】" in concise
    assert "成都武侯宾馆雇员忠诚险及现金险项目" in concise
    assert "成都工程 0 条、川内其他商机 2 条，详见附件" in concise
    assert "宜宾市某工程一切险项目" in full
    assert "泸州市江阳区老旧街区改造工程" in full
    assert "截止：2026-09-01 10:00" in full
    assert "原文：https://example.com/engineering-1" in full
    assert "联系：王老师 028****5678" in full
    assert "028-12345678" not in full
    assert "待复核" not in concise
    assert "待复核" not in full

    # 成都保险项目按“一项目一图”打包；两条项目必须得到两张不同的高清 PNG。
    assert bundle.cards_zip is not None
    assert bundle.cards_zip_bytes == bundle.cards_zip.getvalue()
    with zipfile.ZipFile(bundle.cards_zip) as archive:
        card_names = archive.namelist()
        assert len(card_names) == 2
        assert len(set(card_names)) == 2
        assert all(name.endswith(".png") for name in card_names)
        assert any("金牛区" in name and "雇主责任险" in name for name in card_names)
        assert any("武侯区" in name and "忠诚险" in name for name in card_names)
        for name in card_names:
            card = Image.open(io.BytesIO(archive.read(name)))
            assert card.format == "PNG"
            assert card.width == 1440
            assert card.height >= 1_000


def test_empty_data_still_produces_openable_artifacts_and_safe_copy() -> None:
    bundle = build_report_bundle(None, report_date="2026-08-26", processing_log=None)

    workbook = load_workbook(bundle.excel)
    assert workbook.sheetnames == list(SHEET_NAMES)
    summary = workbook["今日商机汇总"]
    assert summary.max_column == 5
    total_row = _find_summary_row(summary, "总计")
    assert sum(summary.cell(row, 2).value for row in range(4, total_row)) == 0
    assert sum(summary.cell(row, 4).value for row in range(4, total_row)) == 0
    assert workbook["成都地区（保险）"]["A5"].value == "今日暂无符合条件的记录"
    assert workbook["采集与判定审计"].sheet_state == "hidden"

    image = Image.open(bundle.png)
    assert image.format == "PNG"
    assert image.width == 1440
    assert "今日暂无符合条件的可推送商机" in bundle.concise_text
    assert bundle.concise_text == bundle.full_text
    with zipfile.ZipFile(bundle.cards_zip) as archive:
        assert archive.namelist() == ["今日无成都保险商机.txt"]


def test_column_aliases_and_chinese_amount_units_are_supported() -> None:
    source = pd.DataFrame(
        [
            {
                "推送": "是",
                "业务类型": "工程类",
                "分类": "市政施工",
                "项目名称": "金堂县某市政工程",
                "发布日期": "2026-08-26",
                "招标金额": "1.25亿元",
                "地市": "成都市",
                "区县": "金堂县",
                "区域分组": "",
                "服务区域": "金堂县",
                "项目阶段": "招标公告",
                "deadlines": "2026-09-08",
            }
        ]
    )

    normalized = normalize_report_dataframe(source)
    assert list(normalized.columns) == list(REQUIRED_COLUMNS)
    assert bool(normalized.loc[0, "selected"]) is True
    assert normalized.loc[0, "business_type"] == "工程"
    assert normalized.loc[0, "amount"] == 125_000_000
    assert normalized.loc[0, "region_group"] == "成都地区"
    assert normalized.loc[0, "deadline"] == "2026-09-08"


def test_frontend_deadlines_contacts_and_benign_quality_status_are_preserved() -> None:
    """锁定 opportunity_app 当前中间表的字段契约。"""

    source = pd.DataFrame(
        [
            {
                "selected": True,
                "business_type": "保险",
                "category": "意外险",
                "project_name": "成都某团体意外险",
                "amount": None,
                "city": "成都市",
                "district": "金牛区",
                "region_group": "成都地区",
                "service_region": "金牛区",
                "registration_deadline": "2026-09-01",
                "bid_deadline": "2026-09-02 10:00",
                "tenderer_contact": "王老师",
                "tenderer_phone": "028-12345678",
                "agent_contact": "李老师",
                "agent_phone": "028-87654321",
                "quality_issue": "缺失；False；",
            }
        ]
    )
    normalized = normalize_report_dataframe(source)

    assert normalized.loc[0, "deadline"] == "报名：2026-09-01；投标：2026-09-02 10:00"
    assert normalized.loc[0, "contact"] == (
        "招标人：王老师 028-12345678；代理：李老师 028-87654321"
    )
    assert normalized.loc[0, "quality_issue"] == "缺失"

    workbook = load_workbook(build_opportunity_excel(source, report_date="2026-08-26"))
    # 保险金额不是筛选门槛，金额缺失不单独造成内部确认任务。
    assert workbook["未分区域项目"]["A5"].value == "今日暂无符合条件的记录"
