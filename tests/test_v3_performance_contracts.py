"""V3 performance and failure-isolation contract tests.

These tests are intentionally offline.  They define the small public/internal
surface that keeps the V3 workflow fast and predictable:

* three semantic review lanes run concurrently and normally issue one request
  each;
* one malformed result row is quarantined without retrying or discarding its
  valid siblings;
* a ``finish_reason=length`` failure gets one bounded 32K retry only;
* one task can reserve no more than six real model calls, even across threads.

The tests must never call a real model API.
"""

from __future__ import annotations

from collections import Counter
from concurrent.futures import ThreadPoolExecutor
from copy import deepcopy
import json
from pathlib import Path
import re
import sys
import threading
import time
from types import SimpleNamespace
from typing import Any, Callable, Mapping

import pytest
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import app


EXPECTED_LANES = {
    "fatal_compliance",
    "scoring",
    "technical_commercial",
}


@pytest.fixture(scope="module")
def sample_document_evidence() -> dict[str, Any]:
    """Parse the two supplied regression documents once for all sample tests."""

    sample_dir = PROJECT_ROOT / "招标模版及投标文件word"
    documents = sorted(sample_dir.glob("*.docx"), key=lambda path: path.stat().st_size)
    if len(documents) < 2:
        pytest.skip("The two local sample DOCX files are not present in this checkout.")

    tender_path, bid_path = documents[0], documents[-1]
    tender_text, tender_stats = app.extract_docx_text(tender_path.read_bytes(), tender_path.name)
    bid_text, bid_stats = app.extract_docx_text(bid_path.read_bytes(), bid_path.name)
    tender_context, tender_sources = app.build_full_document_context(tender_text, "招标文件")
    bid_context, bid_sources = app.build_full_document_context(bid_text, "投标文件")
    return {
        "tender_text": tender_text,
        "bid_text": bid_text,
        "tender_stats": tender_stats,
        "bid_stats": bid_stats,
        "tender_context": tender_context,
        "bid_context": bid_context,
        "tender_sources": tender_sources,
        "bid_sources": bid_sources,
    }


def _require_callable(name: str) -> Callable[..., Any]:
    value = getattr(app, name, None)
    assert callable(value), f"V3 contract requires app.{name}(...)"
    return value


def _reserve_call(state: Any, lane_name: str = "test_contract") -> bool:
    """Normalize either supported exhaustion style: ``False`` or an exception."""

    try:
        result = state.reserve_call(lane_name)
    except app.TaskBudgetError:
        return False
    return result is not False


def _empty_lane_payload(lane: str) -> dict[str, Any]:
    return {
        "schema_version": "3.0",
        "lane": lane,
        "status": "complete",
        "covered_anchor_ids": [],
        "defects_list": [],
        "scoring_list": [],
        "warnings": [],
    }


def _defect_row(sequence: int, check_point: str) -> dict[str, Any]:
    return {
        "finding_id": f"F-{sequence:03d}",
        "module": "资格审查",
        "check_point": check_point,
        "tender_source_ids": ["T-P00001"],
        "tender_quote": "投标人须提供有效营业执照。",
        "requirement": "投标人须提供有效营业执照。",
        "bid_source_ids": ["B-P00001"],
        "bid_quote": "已提供有效营业执照。",
        "evidence_type": "direct",
        "conclusion": "符合",
        "bid_status": "已提供有效营业执照。",
        "issue": "符合",
        "risk_level": "正常",
        "suggestion": "无需修改",
        "confidence": "high",
        "anchor_ids": [],
    }


class ScriptedChatClient:
    """Tiny OpenAI-compatible offline client for bounded-retry tests."""

    def __init__(self, outcomes: list[Any]) -> None:
        self._outcomes = list(outcomes)
        self.calls: list[dict[str, Any]] = []
        self.chat = SimpleNamespace(
            completions=SimpleNamespace(create=self._create),
        )

    def _create(self, **kwargs: Any) -> Any:
        self.calls.append(dict(kwargs))
        if not self._outcomes:
            raise AssertionError("request_lane_json_resilient exceeded its scripted call budget")
        outcome = self._outcomes.pop(0)
        if isinstance(outcome, BaseException):
            raise outcome
        finish_reason, payload = outcome
        content = payload if isinstance(payload, str) else json.dumps(payload, ensure_ascii=False)
        return SimpleNamespace(
            choices=[
                SimpleNamespace(
                    finish_reason=finish_reason,
                    message=SimpleNamespace(content=content, reasoning_content=""),
                )
            ],
            usage=SimpleNamespace(prompt_tokens=100, completion_tokens=50),
        )


def test_v3_contract_surface_is_explicit() -> None:
    """Fail with actionable names instead of an obscure AttributeError."""

    assert callable(getattr(app, "V3RunState", None)), "V3 requires app.V3RunState"
    for name in (
        "validate_lane_payload",
        "request_lane_json_resilient",
        "run_three_lane_review",
        "merge_lane_results",
        "analyze_documents_v3",
    ):
        _require_callable(name)


def test_v3_run_state_enforces_six_call_cap_atomically_across_threads() -> None:
    """A ContextVar-like per-thread counter is insufficient for the V3 pool."""

    run_state_type = getattr(app, "V3RunState", None)
    assert callable(run_state_type), "V3 requires app.V3RunState(max_calls=6)"
    state = run_state_type(max_calls=6)

    # Deliberately race far more reservations than the task is allowed to use.
    with ThreadPoolExecutor(max_workers=24) as executor:
        accepted = list(executor.map(lambda _: _reserve_call(state), range(96)))

    assert sum(accepted) == 6
    assert state.calls == 6
    assert _reserve_call(state) is False
    assert state.calls == 6


def test_validate_lane_payload_quarantines_only_the_bad_row() -> None:
    """One malformed item must not invalidate or re-request the whole lane."""

    validate_lane_payload = _require_callable("validate_lane_payload")
    first = _defect_row(1, "营业执照有效性")
    second = _defect_row(3, "投标有效期")
    payload = {
        "lane": "fatal_compliance",
        "status": "complete",
        "defects_list": [first, "this row is deliberately not a JSON object", second],
        "scoring_list": [],
        "warnings": [],
    }

    validated = validate_lane_payload(
        payload,
        "fatal_compliance",
        {"T-P00001": "投标人须提供有效营业执照。"},
        {"B-P00001": "已提供有效营业执照。"},
        expected_anchor_ids=(),
    )

    assert isinstance(validated, Mapping)
    valid_rows = validated.get("defects_list")
    assert isinstance(valid_rows, list)
    check_points = [row["检查要点"] for row in valid_rows]
    assert "营业执照有效性" in check_points
    assert "投标有效期" in check_points
    assert all(isinstance(row, Mapping) for row in valid_rows)
    assert sum("自动核查未能可靠完成" in value for value in check_points) == 1

    # The malformed sibling is observable, but it neither raises nor removes
    # the two valid rows.  Implementations may expose it as invalid_rows, a
    # warning, or both.
    invalid_rows = validated.get("invalid_rows", [])
    warnings = validated.get("warnings", [])
    assert invalid_rows or warnings
    assert validated.get("status") in {"partial", "complete_with_warnings", "complete"}


def test_length_gets_one_32k_retry_and_no_source_block_recursion(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A lane may retry once; it must never re-enter V2's recursive splitter."""

    request_lane = _require_callable("request_lane_json_resilient")
    run_state_type = getattr(app, "V3RunState", None)
    assert callable(run_state_type)
    state = run_state_type(max_calls=6)
    client = ScriptedChatClient(
        [
            ("length", "truncated JSON"),
            ("stop", _empty_lane_payload("fatal_compliance")),
        ]
    )

    # V3 must not fall back to either of the V2 recursive splitters.
    def forbidden_split(*args: Any, **kwargs: Any) -> Any:
        del args, kwargs
        raise AssertionError("V3 length handling must not recursively split source blocks")

    if hasattr(app, "_split_work_unit"):
        monkeypatch.setattr(app, "_split_work_unit", forbidden_split)
    if hasattr(app, "extract_requirements_resilient"):
        monkeypatch.setattr(app, "extract_requirements_resilient", forbidden_split)

    result = request_lane(
        client,
        "offline-mock",
        "fatal_compliance",
        "Tender context",
        "Bid context",
        (),
        state,
    )

    max_tokens_seen = [int(call["max_tokens"]) for call in client.calls]
    assert result["lane"] == "fatal_compliance"
    assert len(max_tokens_seen) == 2
    assert max_tokens_seen[0] < 32_768
    assert max_tokens_seen[1] == 32_768
    assert state.calls == 2


def test_second_length_is_terminal_after_two_calls() -> None:
    """Repeated length cannot create a binary retry tree."""

    request_lane = _require_callable("request_lane_json_resilient")
    run_state_type = getattr(app, "V3RunState", None)
    assert callable(run_state_type)
    state = run_state_type(max_calls=6)
    client = ScriptedChatClient(
        [
            ("length", "first truncated JSON"),
            ("length", "second truncated JSON"),
        ]
    )

    with pytest.raises(app.ResponseLengthError):
        request_lane(
            client,
            "offline-mock",
            "scoring",
            "Tender context",
            "Bid context",
            (),
            state,
        )

    max_tokens_seen = [int(call["max_tokens"]) for call in client.calls]
    assert len(max_tokens_seen) == 2
    assert max_tokens_seen[0] < 32_768
    assert max_tokens_seen[1] == 32_768
    assert state.calls == 2


def test_length_retry_cannot_cross_global_six_call_cap() -> None:
    """The seventh attempted request is rejected before any HTTP work starts."""

    request_lane = _require_callable("request_lane_json_resilient")
    run_state_type = getattr(app, "V3RunState", None)
    assert callable(run_state_type)
    state = run_state_type(max_calls=6)
    for _ in range(5):
        assert _reserve_call(state) is True

    client = ScriptedChatClient(
        [
            ("length", "sixth and final allowed call"),
            # This response must remain unused because reservation seven fails.
            ("stop", _empty_lane_payload("technical_commercial")),
        ]
    )

    with pytest.raises(app.TaskBudgetError):
        request_lane(
            client,
            "offline-mock",
            "technical_commercial",
            "Tender context",
            "Bid context",
            (),
            state,
        )

    assert len(client.calls) == 1
    assert state.calls == 6


def test_typical_review_runs_exactly_three_lanes_concurrently(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """The normal path is three parallel requests, not dozens of serial calls."""

    run_three_lane_review = _require_callable("run_three_lane_review")
    _require_callable("request_lane_json_resilient")
    run_state_type = getattr(app, "V3RunState", None)
    assert callable(run_state_type)
    state = run_state_type(max_calls=6)
    expected_state = state

    main_thread = threading.get_ident()
    barrier = threading.Barrier(3, timeout=2.0)
    lock = threading.Lock()
    active = 0
    peak_active = 0
    lane_calls: list[str] = []
    worker_threads: list[int] = []
    clients: list[object] = []

    def client_factory(*args: Any, **kwargs: Any) -> object:
        del args, kwargs
        client = object()
        with lock:
            clients.append(client)
        return client

    def fake_lane_request(
        client: object,
        model: str,
        lane: Any,
        tender_context: str,
        bid_context: str,
        anchors: Any,
        state: Any,
    ) -> dict[str, Any]:
        nonlocal active, peak_active
        del tender_context, bid_context, anchors
        assert model == "offline-mock"
        assert state is expected_state
        lane_name = getattr(lane, "name", str(lane))
        assert _reserve_call(state, lane_name) is True
        with lock:
            lane_calls.append(lane_name)
            worker_threads.append(threading.get_ident())
            active += 1
            peak_active = max(peak_active, active)
        try:
            barrier.wait()
            time.sleep(0.02)
            return _empty_lane_payload(lane_name)
        finally:
            with lock:
                active -= 1

    monkeypatch.setattr(app, "request_lane_json_resilient", fake_lane_request)

    def main_thread_logger(message: str) -> None:
        del message
        assert threading.get_ident() == main_thread

    def main_thread_progress(value: int, label: str) -> None:
        del value, label
        assert threading.get_ident() == main_thread

    lane_results = run_three_lane_review(
        client_factory,
        "offline-mock",
        "【T-P00001】Tender text",
        "【B-P00001】Bid text",
        main_thread_logger,
        main_thread_progress,
        state=state,
    )

    assert isinstance(lane_results, Mapping)
    assert set(lane_results["v3_meta"]["lane_status"]) == EXPECTED_LANES
    assert Counter(lane_calls) == Counter({lane: 1 for lane in EXPECTED_LANES})
    assert state.calls == 3
    assert peak_active == 3
    assert len(set(worker_threads)) == 3
    assert all(thread_id != main_thread for thread_id in worker_threads)
    assert len(clients) == 3
    assert len({id(client) for client in clients}) == 3


def test_sample_documents_preserve_decision_critical_source_ids_and_text(
    sample_document_evidence: dict[str, Any],
) -> None:
    """Lock facts that a concise Gemini-style report must not confuse or omit."""

    tender_sources = sample_document_evidence["tender_sources"]
    bid_sources = sample_document_evidence["bid_sources"]

    # The exact project number includes the final ``-2`` in both documents.
    project_id = "ZHH-BX-202638-2"
    assert tender_sources["P00001"] == f"项目编号:{project_id}"
    assert project_id in bid_sources["P00355"]
    assert project_id in sample_document_evidence["tender_context"]
    assert project_id in sample_document_evidence["bid_context"]

    # The tender's scoring table has exactly four top-level rows and totals 100.
    expected_scoring_rows = {
        "T004-R002": ("报价", "10分"),
        "T004-R003": ("项目要求", "30分"),
        "T004-R004": ("实施方案", "35分"),
        "T004-R005": ("履约能力", "25分"),
    }
    for source_id, expected_fragments in expected_scoring_rows.items():
        source_text = tender_sources[source_id]
        assert all(fragment in source_text for fragment in expected_fragments)
    assert sum(int(points.removesuffix("分")) for _, points in expected_scoring_rows.values()) == 100

    # The formal quotation form is blank; a later schedule separately contains
    # 699/1200.  These are distinct facts and source IDs, not interchangeable.
    formal_quote = bid_sources["T024-R006"]
    schedule_quote = bid_sources["T025-R006"]
    assert re.search(r"小写:\s*/\s*大写:\s*$", formal_quote)
    assert "699元/人" not in formal_quote and "1200元/人" not in formal_quote
    assert "699元/人" in schedule_quote and "1200元/人" in schedule_quote
    assert "【T024-R006】" in sample_document_evidence["bid_context"]
    assert "【T025-R006】" in sample_document_evidence["bid_context"]

    # Images exist, but their content is deliberately outside the text-only
    # evidence contract.
    assert sample_document_evidence["bid_stats"]["media_files"] == 90


def test_empty_model_lanes_still_emit_sample_deterministic_truths(
    sample_document_evidence: dict[str, Any],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Even three empty model lanes must still produce the safe local minimum report."""

    lane_calls: list[str] = []
    state = app.V3RunState(max_calls=6)

    def fake_lane_request(
        client: object,
        model: str,
        lane: Any,
        tender_context: str,
        bid_context: str,
        anchors: Any,
        state: Any,
    ) -> dict[str, Any]:
        del client, tender_context, bid_context, anchors
        assert model == "offline-empty-model"
        lane_name = getattr(lane, "name", str(lane))
        assert _reserve_call(state, lane_name) is True
        lane_calls.append(lane_name)
        return _empty_lane_payload(lane_name)

    monkeypatch.setattr(app, "request_lane_json_resilient", fake_lane_request)

    result = app.analyze_documents_v3(
        client_factory=lambda: object(),
        model="offline-empty-model",
        tender_text=sample_document_evidence["tender_text"],
        bid_text=sample_document_evidence["bid_text"],
        tender_name="sample-tender.docx",
        bid_name="sample-bid.docx",
        logger=lambda message: None,
        progress=lambda value, label: None,
        state=state,
    )

    assert Counter(lane_calls) == Counter({lane: 1 for lane in EXPECTED_LANES})
    assert result["v3_meta"]["api_calls"] == 3

    quotation_rows = [
        row
        for row in result["defects_list"]
        if row.get("检查要点") == "正式报价一览表文字完整性"
    ]
    assert len(quotation_rows) == 1
    quotation = quotation_rows[0]
    assert set(quotation["_bid_source_ids"]) >= {
        "T024-R004",
        "T024-R005",
        "T024-R006",
        "T025-R006",
    }
    assert quotation["_tender_source_ids"] == ["P00146"]
    assert "(空)" in quotation["投标文件现状"]
    assert "699元/人" in quotation["投标文件现状"]
    assert "1200元/人" in quotation["投标文件现状"]
    assert "不能替代正式报价表填写" in quotation["投标文件现状"]
    assert quotation["风险等级"] == "致命/废标风险"
    assert "废标高风险" in quotation["存在问题与缺陷"]

    coverage_rows = [
        row
        for row in result["defects_list"]
        if row.get("检查要点") == "保障额度文字一致性"
    ]
    assert len(coverage_rows) == 1
    coverage = coverage_rows[0]
    assert coverage["_tender_source_ids"] == ["P00242"]
    assert coverage["_bid_source_ids"]
    assert "100万元" in coverage["招标文件要求"]
    assert "50万元" in coverage["招标文件要求"]
    assert "100万元" in coverage["投标文件现状"]
    assert "50万元" in coverage["投标文件现状"]
    assert coverage["风险等级"] == "正常/符合"

    expected_scores = {
        "报价": (10, "T004-R002"),
        "项目要求": (30, "T004-R003"),
        "实施方案": (35, "T004-R004"),
        "履约能力": (25, "T004-R005"),
    }
    assert len(result["scoring_list"]) == 4
    assert {row["评分项"] for row in result["scoring_list"]} == set(expected_scores)
    for row in result["scoring_list"]:
        full_score, source_id = expected_scores[row["评分项"]]
        assert row["满分"] == full_score
        assert row["_tender_source_ids"] == [source_id]
        assert row["当前预估得分"] == "待人工复核"
    assert sum(row["满分"] for row in result["scoring_list"]) == 100

    # Source requirements may mention a seal, but the generated report must not
    # turn text-only evidence into a definite visual or legal conclusion.
    public_text = "\n".join(
        str(row.get(field, ""))
        for row in result["defects_list"]
        for field in app.DEFECT_FIELDS
    ) + "\n" + "\n".join(
        str(row.get(field, ""))
        for row in result["scoring_list"]
        for field in app.SCORING_FIELDS
    )
    for forbidden in (
        "已废标",
        "编号遗漏-2",
        "已盖公章",
        "未盖公章",
        "未发现公章",
        "缺签",
        "已签字",
        "签章齐全",
        "证明材料齐全",
    ):
        assert forbidden not in public_text


def test_text_only_safety_downgrades_visual_claims_and_rejects_false_finality(
    sample_document_evidence: dict[str, Any],
) -> None:
    result = {
        "defects_list": [
            {
                "序号": 1,
                "核查模块": "格式规范",
                "检查要点": "签字盖章",
                "招标文件出处": "【P00399】",
                "招标文件要求": "报价一览表每页均需盖章并签字。",
                "投标文件现状": "未发现公章或手写签字。",
                "存在问题与缺陷": "报价表未盖章，已废标，结论100%正确。",
                "风险等级": "致命/废标风险",
                "修改建议": "确定废标。",
            },
            {
                "序号": 2,
                "核查模块": "形式审查",
                "检查要点": "项目编号",
                "招标文件出处": "【P00001】",
                "招标文件要求": "项目编号必须一致。",
                "投标文件现状": "项目编号缺少末尾-2。",
                "存在问题与缺陷": "项目编号遗漏。",
                "风险等级": "致命/废标风险",
                "修改建议": "补全编号。",
            },
        ],
        "scoring_list": [
            {
                "评分项": "履约能力",
                "满分": 25,
                "评分标准": "合同复印件需盖章。",
                "招标文件出处": "【T004-R005】",
                "当前预估得分": 25,
                "得分依据及扣分说明": "证明材料齐全且均已盖章。",
            }
        ],
        "v3_meta": {"warnings": []},
    }

    guarded = app.enforce_v3_text_only_safety(
        result,
        sample_document_evidence["tender_sources"],
        sample_document_evidence["bid_sources"],
    )

    visual_row = guarded["defects_list"][0]
    assert visual_row["风险等级"] == "待人工复核"
    assert visual_row["存在问题与缺陷"].startswith("仅文字模式无法确认")
    assert "已废标" not in visual_row["存在问题与缺陷"]
    assert "确定废标" not in visual_row["修改建议"]
    assert "100%正确" not in visual_row["存在问题与缺陷"]

    project_row = guarded["defects_list"][1]
    assert project_row["检查要点"] == "项目编号全文一致性"
    assert "ZHH-BX-202638-2" in project_row["投标文件现状"]
    assert project_row["风险等级"] == "正常/符合"
    assert "缺少末尾-2" not in project_row["投标文件现状"]
    assert project_row["_bid_source_ids"]

    scoring_row = guarded["scoring_list"][0]
    assert scoring_row["当前预估得分"] == "待人工复核"
    assert scoring_row["_manual"] is True
    assert scoring_row["得分依据及扣分说明"].startswith("待人工复核")
    assert guarded["v3_meta"]["bid_contains_tender_project_id"] is True
    assert any("视觉事项已降级" in warning for warning in guarded["v3_meta"]["warnings"])
    assert any("项目编号缺失判断已被本地全文反证" in warning for warning in guarded["v3_meta"]["warnings"])


def test_excel_manual_review_row_uses_neutral_fill_not_risk_colors() -> None:
    manual_row = {
        "序号": 1,
        "核查模块": "图片与签章",
        "检查要点": "仅文字模式复核",
        "招标文件出处": "【P00001】",
        "招标文件要求": "需核对原件。",
        "投标文件现状": "图片内容未识别。",
        "存在问题与缺陷": "仅文字模式无法确认签章。",
        "风险等级": "待人工复核",
        "修改建议": "人工查看原件。",
    }
    report = app.build_excel_report({"defects_list": [manual_row], "scoring_list": []})
    workbook = load_workbook(report, data_only=False)
    try:
        worksheet = workbook["缺陷核查记录"]
        risk_cell = worksheet.cell(2, app.DEFECT_FIELDS.index("风险等级") + 1)
        assert str(risk_cell.fill.fgColor.rgb)[-6:].upper() == "D9EAF7"
        assert risk_cell.font.bold is True
        assert risk_cell.font.italic is not True
        for cell in worksheet[2]:
            assert cell.font.italic is not True
            assert str(cell.fill.fgColor.rgb)[-6:].upper() not in {"C00000", "F4B183", "C6E0B4"}
            if cell.coordinate != risk_cell.coordinate:
                assert str(cell.fill.fgColor.rgb)[-6:].upper() != "D9EAF7"
    finally:
        workbook.close()


def test_excel_presentation_is_compact_non_mutating_and_keeps_full_evidence_in_comments() -> None:
    long_quote = "这是完整原文证据，包含金额、期限和逐项要求。" * 35
    fatal = {
        "序号": 1,
        "核查模块": "报价",
        "检查要点": "报价一览表填写完整性",
        "招标文件出处": "、".join(f"【T005-R00{i}】" for i in range(1, 7)),
        "招标文件要求": f"报价表必须填写单价、合计和总价。 原文摘录：{long_quote}",
        "投标文件现状": f"【T024-R004】 正式报价表关键金额为空。\n原文摘录：{long_quote}",
        "存在问题与缺陷": "正式报价表关键金额为空，存在无效报价高风险。",
        "风险等级": "致命/废标风险",
        "修改建议": "完整填写各分项单价、合计金额和总价，并核对大小写。",
    }
    no_action = {
        "序号": 2,
        "核查模块": "技术",
        "检查要点": "保险条款覆盖",
        "招标文件出处": "【P00242】",
        "招标文件要求": "按要求响应。",
        "投标文件现状": long_quote,
        "存在问题与缺陷": "无",
        "风险等级": "待人工复核",
        "修改建议": "无",
    }
    visual = {
        "序号": 3,
        "核查模块": "形式审查",
        "检查要点": "签字盖章",
        "招标文件出处": "【P00146】",
        "招标文件要求": "按要求签字盖章。",
        "投标文件现状": "图片和印章未识别。",
        "存在问题与缺陷": "仅文字模式无法确认签章。",
        "风险等级": "待人工复核",
        "修改建议": "人工查看原件。",
    }
    auto_gap = {
        "序号": 4,
        "核查模块": "评分核查",
        "检查要点": "自动核查未能可靠完成",
        "招标文件出处": "【T004-R002】、【T004-R003】",
        "招标文件要求": "请对照原文。",
        "投标文件现状": "自动核查证据未确认。",
        "存在问题与缺陷": "待人工复核：模型未确认 12 个规则锚点。",
        "风险等级": "待人工复核",
        "修改建议": "对照 Word 原件人工复核。",
    }
    normal = {
        "序号": 5,
        "核查模块": "期限",
        "检查要点": "递交截止时间",
        "招标文件出处": "【P00053】",
        "招标文件要求": "在截止时间前递交。",
        "投标文件现状": "日期早于截止时间。",
        "存在问题与缺陷": "无",
        "风险等级": "正常/符合",
        "修改建议": "无",
    }
    score = {
        "评分项": "实施方案",
        "满分": 35,
        "评分标准": f"七项方案缺一项扣五分。 原文摘录：{long_quote}",
        "招标文件出处": "【T004-R004】",
        "当前预估得分": "待人工复核",
        "得分依据及扣分说明": long_quote,
    }
    result = {"defects_list": [fatal, no_action, visual, auto_gap, normal], "scoring_list": [score]}
    snapshot = deepcopy(result)

    projected = app.prepare_excel_report_data(result)
    assert result == snapshot
    assert len(projected["defects_list"]) == 4
    assert len(projected["defects_list"]) < len(result["defects_list"])
    projected_fatal = projected["defects_list"][0]
    assert "【T005-R001】" in projected_fatal["招标文件出处"]
    assert "等6处" in projected_fatal["招标文件出处"]
    assert "原文摘录" not in projected_fatal["招标文件要求"]
    for field, limit in app.EXCEL_VISIBLE_LIMITS.items():
        for row in projected["defects_list"] + projected["scoring_list"]:
            if field in row:
                assert len(str(row[field])) <= limit

    workbook = load_workbook(app.build_excel_report(result), data_only=False)
    try:
        defects = workbook["缺陷核查记录"]
        scoring = workbook["预估打分表"]
        assert defects.max_row == 5
        assert scoring.max_row == 2
        assert defects["E2"].comment is not None
        assert long_quote[:80] in defects["E2"].comment.text
        assert scoring["C2"].comment is not None
        assert long_quote[:80] in scoring["C2"].comment.text
        assert all(cell.font.italic is not True for row in defects.iter_rows(min_row=2) for cell in row)
    finally:
        workbook.close()


def test_sample_deterministic_augmentation_recovers_quote_coverage_and_all_scores(
    sample_document_evidence: dict[str, Any],
) -> None:
    """Python evidence checks remain useful even when all three model lanes are empty."""

    augmented = app.augment_v3_deterministic_findings(
        {"defects_list": [], "scoring_list": [], "v3_meta": {"warnings": []}},
        sample_document_evidence["tender_sources"],
        sample_document_evidence["bid_sources"],
    )

    defects_by_check = {row["检查要点"]: row for row in augmented["defects_list"]}
    quotation = defects_by_check["正式报价一览表文字完整性"]
    assert quotation["风险等级"] == "致命/废标风险"
    assert "废标高风险" in quotation["存在问题与缺陷"]
    assert "已废标" not in quotation["存在问题与缺陷"]
    assert "确定废标" not in quotation["存在问题与缺陷"]
    assert "T024-R006" in quotation["_bid_source_ids"]
    assert "T025-R006" in quotation["_bid_source_ids"]
    assert "699元/人" in quotation["投标文件现状"]
    assert "1200元/人" in quotation["投标文件现状"]

    coverage = defects_by_check["保障额度文字一致性"]
    assert coverage["风险等级"] == "正常/符合"
    assert "100万元" in coverage["招标文件要求"]
    assert "50万元" in coverage["招标文件要求"]
    assert "100万元" in coverage["投标文件现状"]
    assert "50万元" in coverage["投标文件现状"]
    assert coverage["_tender_source_ids"] == ["P00242"]
    assert coverage["_bid_source_ids"]

    scoring = augmented["scoring_list"]
    assert [(row["评分项"], row["满分"]) for row in scoring] == [
        ("报价", 10),
        ("项目要求", 30),
        ("实施方案", 35),
        ("履约能力", 25),
    ]
    assert sum(row["满分"] for row in scoring) == 100
    assert all(row["当前预估得分"] == "待人工复核" for row in scoring)
    assert all(row["_manual"] is True for row in scoring)


def _assert_lane_payload_not_complete(
    payload: Mapping[str, Any],
    lane: str,
    *,
    expected_anchor_ids: tuple[str, ...] = (),
) -> None:
    """Accept either explicit invalid status or a validation exception."""

    try:
        validated = app.validate_lane_payload(
            payload,
            lane,
            {"T-P00001": "投标人须提供有效营业执照。"},
            {"B-P00001": "已提供有效营业执照。"},
            expected_anchor_ids=expected_anchor_ids,
        )
    except app.ModelOutputError:
        return
    assert validated.get("status") != "complete"
    assert validated.get("warnings") or validated.get("invalid_rows")


@pytest.mark.parametrize(
    "payload,lane",
    [
        ({}, "fatal_compliance"),
        (
            {
                "schema_version": "3.0",
                "lane": "fatal_compliance",
                "defects_list": [],
                "scoring_list": [],
                "warnings": [],
            },
            "fatal_compliance",
        ),
        (
            {
                "status": "complete",
                "lane": "fatal_compliance",
                "defects_list": [],
                "scoring_list": [],
                "warnings": [],
            },
            "fatal_compliance",
        ),
        (
            {
                "schema_version": "2.0",
                "status": "complete",
                "lane": "fatal_compliance",
                "defects_list": [],
                "scoring_list": [],
                "warnings": [],
            },
            "fatal_compliance",
        ),
        (
            {
                "schema_version": "3.0",
                "status": "complete",
                "lane": "scoring",
                "defects_list": [],
                "scoring_list": [],
                "warnings": [],
            },
            "fatal_compliance",
        ),
        (
            {
                "schema_version": "3.0",
                "status": "complete",
                "lane": "fatal_compliance",
                "scoring_list": [],
                "warnings": [],
            },
            "fatal_compliance",
        ),
        (
            {
                "schema_version": "3.0",
                "status": "complete",
                "lane": "fatal_compliance",
                "defects_list": [],
                "warnings": [],
            },
            "fatal_compliance",
        ),
        (
            {
                "schema_version": "3.0",
                "status": "complete",
                "lane": "fatal_compliance",
                "defects_list": {},
                "scoring_list": [],
                "warnings": [],
            },
            "fatal_compliance",
        ),
    ],
    ids=[
        "empty-object",
        "missing-status",
        "missing-schema",
        "wrong-schema",
        "wrong-lane",
        "missing-defects-array",
        "missing-scoring-array",
        "non-array-defects",
    ],
)
def test_incomplete_or_malformed_lane_envelope_cannot_be_complete(
    payload: Mapping[str, Any],
    lane: str,
) -> None:
    _assert_lane_payload_not_complete(payload, lane)


def test_scoring_lane_with_covered_anchor_but_no_score_rows_cannot_be_complete() -> None:
    _assert_lane_payload_not_complete(
        {
            "schema_version": "3.0",
            "status": "complete",
            "lane": "scoring",
            "covered_anchor_ids": ["A-S-0001"],
            "defects_list": [],
            "scoring_list": [],
            "warnings": [],
        },
        "scoring",
        expected_anchor_ids=("A-S-0001",),
    )


def test_hallucinated_requirement_penalty_cannot_upgrade_verified_quote_to_fatal() -> None:
    """Risk severity must come from the verified tender quote, not model prose."""

    tender_quote = "投标人须提供有效营业执照。"
    bid_quote = "本页未提供营业执照。"
    validated = app.validate_lane_payload(
        {
            "schema_version": "3.0",
            "lane": "fatal_compliance",
            "status": "complete",
            "covered_anchor_ids": [],
            "defects_list": [
                {
                    "finding_id": "F-HALLUCINATED-PENALTY",
                    "module": "资格审查",
                    "check_point": "营业执照",
                    "tender_source_ids": ["T-P00001"],
                    "tender_quote": tender_quote,
                    # This penalty is absent from the source-backed quote.
                    "requirement": "投标人须提供有效营业执照，否则废标。",
                    "bid_source_ids": ["B-P00001"],
                    "bid_quote": bid_quote,
                    "evidence_type": "direct",
                    "conclusion": "不符合",
                    "bid_status": bid_quote,
                    "issue": "未提供营业执照。",
                    "risk_level": "扣分瑕疵",
                    "suggestion": "补充营业执照。",
                    "confidence": "high",
                    "anchor_ids": [],
                }
            ],
            "scoring_list": [],
            "warnings": [],
        },
        "fatal_compliance",
        {"T-P00001": tender_quote},
        {"B-P00001": bid_quote},
        expected_anchor_ids=(),
    )

    assert validated["status"] == "complete"
    row = validated["defects_list"][0]
    assert row["风险等级"] != "致命/废标风险"
    assert "废标" not in tender_quote


def test_local_project_id_omission_is_not_erased_by_correct_id_elsewhere() -> None:
    """Global presence cannot disprove a defect tied to a different cited field."""

    local_source_id = "B-LOCAL-FIELD"
    other_source_id = "B-OTHER-PAGE"
    result = {
        "defects_list": [
            {
                "序号": 1,
                "核查模块": "形式审查",
                "检查要点": "报价表项目编号",
                "招标文件出处": "【T-PROJECT】",
                "招标文件要求": "项目编号须填写为 ZHH-BX-202638-2。",
                "投标文件现状": f"【{local_source_id}】项目编号:ZHH-BX-202638，缺少末尾-2。",
                "存在问题与缺陷": "被引用的报价表字段项目编号不完整。",
                "风险等级": "扣分/瑕疵",
                "修改建议": "补全当前字段。",
                "_bid_source_ids": [local_source_id],
            }
        ],
        "scoring_list": [],
        "v3_meta": {"warnings": []},
    }

    guarded = app.enforce_v3_text_only_safety(
        result,
        {"T-PROJECT": "项目编号:ZHH-BX-202638-2"},
        {
            local_source_id: "项目编号:ZHH-BX-202638",
            other_source_id: "授权声明项目编号:ZHH-BX-202638-2",
        },
    )

    row = guarded["defects_list"][0]
    assert row["风险等级"] != "正常/符合"
    assert row["_bid_source_ids"] == [local_source_id]
    assert other_source_id not in row["投标文件现状"]
    assert "缺少末尾-2" in row["投标文件现状"]


def test_anchor_claim_is_covered_only_by_its_bound_tender_source() -> None:
    """One source-backed row cannot claim a sibling anchor from another source."""

    tender_sources = {
        "T1": "投标人须提供有效营业执照。",
        "T2": "投标人须提供财务审计报告。",
    }
    bid_sources = {"B1": "已提供有效营业执照。"}
    validated = app.validate_lane_payload(
        {
            "schema_version": "3.0",
            "lane": "fatal_compliance",
            "status": "complete",
            "covered_anchor_ids": ["A1", "A2"],
            "defects_list": [
                {
                    "finding_id": "F-ANCHOR-SOURCE-BINDING",
                    "module": "资格审查",
                    "check_point": "营业执照",
                    "tender_source_ids": ["T1"],
                    "tender_quote": tender_sources["T1"],
                    "requirement": tender_sources["T1"],
                    "bid_source_ids": ["B1"],
                    "bid_quote": bid_sources["B1"],
                    "evidence_type": "direct",
                    "conclusion": "符合",
                    "bid_status": bid_sources["B1"],
                    "issue": "未发现文字偏差。",
                    "risk_level": "正常",
                    "suggestion": "保持现状。",
                    "confidence": "high",
                    # A2 belongs to T2 and therefore cannot be covered here.
                    "anchor_ids": ["A1", "A2"],
                }
            ],
            "scoring_list": [],
            "warnings": [],
        },
        "fatal_compliance",
        tender_sources,
        bid_sources,
        expected_anchor_ids=("A1", "A2"),
        expected_anchor_sources={"A1": "T1", "A2": "T2"},
    )

    assert validated["status"] != "complete"
    assert validated["covered_anchor_ids"] == ["A1"]
    assert validated["missing_anchor_ids"] == ["A2"]
    assert validated["defects_list"][0]["_anchor_ids"] == ["A1"]


def test_scoring_row_without_reason_is_manual_and_covers_no_anchor() -> None:
    """A numeric score without a source-backed reason is not an audited score."""

    tender_sources = {"T1": "报价评分满分10分，按评审价计算。"}
    bid_sources = {"B1": "投标报价为100万元。"}
    validated = app.validate_lane_payload(
        {
            "schema_version": "3.0",
            "lane": "scoring",
            "status": "complete",
            "covered_anchor_ids": ["A1"],
            "defects_list": [],
            "scoring_list": [
                {
                    "score_id": "S-NO-REASON",
                    "score_item": "报价得分",
                    "full_score": 10,
                    "scoring_rule": tender_sources["T1"],
                    "tender_source_ids": ["T1"],
                    "tender_quote": tender_sources["T1"],
                    "bid_source_ids": ["B1"],
                    "bid_quote": bid_sources["B1"],
                    "estimated_score": 8,
                    "reason": "",
                    "confidence": "high",
                    "anchor_ids": ["A1"],
                }
            ],
            "warnings": [],
        },
        "scoring",
        tender_sources,
        bid_sources,
        expected_anchor_ids=("A1",),
        expected_anchor_sources={"A1": "T1"},
    )

    score = validated["scoring_list"][0]
    assert score["当前预估得分"] == "待人工复核"
    assert score["_manual"] is True
    assert validated["covered_anchor_ids"] == []
    assert validated["missing_anchor_ids"] == ["A1"]
    assert validated["status"] != "complete"


def test_partial_lane_with_valid_row_survives_two_other_lane_failures(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A usable non-manual row must keep the whole review reportable."""

    tender_text = "投标人须提供有效营业执照。"
    bid_text = "已提供有效营业执照。"

    def fake_lane_request(
        client: object,
        model: str,
        lane: Any,
        tender_context: str,
        bid_context: str,
        anchors: Any,
        state: Any,
    ) -> dict[str, Any]:
        del client, model, tender_context, bid_context, state
        lane_name = getattr(lane, "name", str(lane))
        if lane_name != "fatal_compliance":
            raise app.ModelOutputError(f"{lane_name} scripted failure")

        anchor_ids = [str(item["anchor_id"]) for item in anchors]
        return {
            "schema_version": "3.0",
            "lane": lane_name,
            "status": "complete",
            "covered_anchor_ids": anchor_ids,
            "defects_list": [
                {
                    "finding_id": "F-USABLE-PARTIAL",
                    "module": "资格审查",
                    "check_point": "营业执照",
                    "tender_source_ids": ["P00001"],
                    "tender_quote": tender_text,
                    "requirement": tender_text,
                    "bid_source_ids": ["P00001"],
                    "bid_quote": bid_text,
                    "evidence_type": "direct",
                    "conclusion": "符合",
                    "bid_status": bid_text,
                    "issue": "未发现文字偏差。",
                    "risk_level": "正常",
                    "suggestion": "保持现状。",
                    "confidence": "high",
                    "anchor_ids": anchor_ids,
                },
                # This sibling makes the lane partial without invalidating the row above.
                "malformed sibling row",
            ],
            "scoring_list": [],
            "warnings": [],
        }

    monkeypatch.setattr(app, "request_lane_json_resilient", fake_lane_request)

    report = app.run_three_lane_review(
        client_factory=object,
        model="offline-mock",
        tender_text=tender_text,
        bid_text=bid_text,
        logger=lambda _message: None,
        progress=lambda _value, _label: None,
        state=app.V3RunState(max_calls=6),
    )

    assert isinstance(report, Mapping)
    assert report["v3_meta"]["lane_status"]["fatal_compliance"] == "partial"
    assert report["v3_meta"]["lane_status"]["scoring"] == "failed"
    assert report["v3_meta"]["lane_status"]["technical_commercial"] == "failed"
    assert report["v3_meta"]["successful_lanes"] == 0
    assert report["v3_meta"]["usable_lanes"] == 1
    assert any(
        row.get("检查要点") == "营业执照" and not row.get("_manual")
        for row in report["defects_list"]
    )


def test_project_number_and_signature_mixed_row_cannot_be_promoted_to_normal() -> None:
    """A matching number cannot resolve an inseparable visual/signature claim."""

    result = {
        "defects_list": [
            {
                "序号": 1,
                "核查模块": "形式审查",
                "检查要点": "项目编号及签章检查",
                "招标文件出处": "【T1】",
                "招标文件要求": "项目编号为 ZHH-BX-202638-2，并按要求签章。",
                "投标文件现状": "【B1】模型称项目编号缺少末尾-2且未签章。",
                "存在问题与缺陷": "项目编号遗漏、签章缺失。",
                "风险等级": "扣分/瑕疵",
                "修改建议": "补全编号并签章。",
                "_tender_source_ids": ["T1"],
                "_bid_source_ids": ["B1"],
                "_manual": False,
            }
        ],
        "scoring_list": [],
        "v3_meta": {"warnings": []},
    }

    guarded = app.enforce_v3_text_only_safety(
        result,
        {"T1": "项目编号为 ZHH-BX-202638-2，并按要求签章。"},
        {"B1": "项目编号为 ZHH-BX-202638-2。"},
    )

    row = guarded["defects_list"][0]
    assert row["风险等级"] == "待人工复核"
    assert row["_manual"] is True
    assert row["检查要点"] == "项目编号及签章检查"
    assert row["检查要点"] != "项目编号全文一致性"


def test_anchor_coverage_uses_the_source_that_actually_backs_the_quote() -> None:
    """Listing T2 is insufficient when the quoted evidence exists only in T1."""

    tender_sources = {
        "T1": "投标人须提供有效营业执照。",
        "T2": "投标人须提供财务审计报告。",
    }
    bid_sources = {"B1": "已提供有效营业执照和财务审计报告。"}
    validated = app.validate_lane_payload(
        {
            "schema_version": "3.0",
            "lane": "fatal_compliance",
            "status": "complete",
            "covered_anchor_ids": ["A1", "A2"],
            "defects_list": [
                {
                    "finding_id": "F-QUOTE-SOURCE-BINDING",
                    "module": "资格审查",
                    "check_point": "资格材料",
                    # Both IDs are listed, but the verbatim quote is backed only by T1.
                    "tender_source_ids": ["T1", "T2"],
                    "tender_quote": tender_sources["T1"],
                    "requirement": "核对营业执照和财务审计报告。",
                    "bid_source_ids": ["B1"],
                    "bid_quote": bid_sources["B1"],
                    "evidence_type": "direct",
                    "conclusion": "符合",
                    "bid_status": bid_sources["B1"],
                    "issue": "未发现文字偏差。",
                    "risk_level": "正常",
                    "suggestion": "保持现状。",
                    "confidence": "high",
                    "anchor_ids": ["A1", "A2"],
                }
            ],
            "scoring_list": [],
            "warnings": [],
        },
        "fatal_compliance",
        tender_sources,
        bid_sources,
        expected_anchor_ids=("A1", "A2"),
        expected_anchor_sources={"A1": "T1", "A2": "T2"},
    )

    assert validated["status"] == "partial"
    assert validated["covered_anchor_ids"] == ["A1"]
    assert validated["missing_anchor_ids"] == ["A2"]
    row = validated["defects_list"][0]
    assert row["_tender_source_ids"] == ["T1", "T2"]
    assert row["_quote_backed_tender_source_ids"] == ["T1"]
    assert row["_anchor_ids"] == ["A1"]


def test_visual_only_model_conclusion_does_not_count_as_usable_lane(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Post-safety usability must exclude a deterministic signature/image claim."""

    tender_text = "投标文件须加盖公章。"
    bid_text = "投标文件页面显示已加盖公章。"

    def fake_lane_request(
        client: object,
        model: str,
        lane: Any,
        tender_context: str,
        bid_context: str,
        anchors: Any,
        state: Any,
    ) -> dict[str, Any]:
        del client, model, tender_context, bid_context, state
        lane_name = getattr(lane, "name", str(lane))
        if lane_name != "fatal_compliance":
            raise app.ModelOutputError(f"{lane_name} scripted failure")
        anchor_ids = [str(item["anchor_id"]) for item in anchors]
        return {
            "schema_version": "3.0",
            "lane": lane_name,
            "status": "complete",
            "covered_anchor_ids": anchor_ids,
            "defects_list": [
                {
                    "finding_id": "F-VISUAL-ONLY",
                    "module": "形式审查",
                    "check_point": "签章完整性",
                    "tender_source_ids": ["P00001"],
                    "tender_quote": tender_text,
                    "requirement": tender_text,
                    "bid_source_ids": ["P00001"],
                    "bid_quote": bid_text,
                    "evidence_type": "direct",
                    "conclusion": "符合",
                    "bid_status": "已加盖公章，签章完整。",
                    "issue": "图片中的公章和签字均完整。",
                    "risk_level": "正常",
                    "suggestion": "无需修改。",
                    "confidence": "high",
                    "anchor_ids": anchor_ids,
                }
            ],
            "scoring_list": [],
            "warnings": [],
        }

    monkeypatch.setattr(app, "request_lane_json_resilient", fake_lane_request)

    report = app.run_three_lane_review(
        client_factory=object,
        model="offline-mock",
        tender_text=tender_text,
        bid_text=bid_text,
        logger=lambda _message: None,
        progress=lambda _value, _label: None,
        state=app.V3RunState(max_calls=6),
    )

    visual_row = next(
        row for row in report["defects_list"] if row.get("检查要点") == "签章完整性"
    )
    assert visual_row["风险等级"] == "待人工复核"
    assert visual_row["_manual"] is True
    assert report["v3_meta"]["pre_safety_usable_lanes"] == 1
    assert report["v3_meta"]["usable_lanes"] == 0


def test_fatal_risk_uses_only_the_verified_quote_not_other_listed_sources() -> None:
    """An unrelated rejection clause in T2 cannot make a T1 finding fatal."""

    tender_sources = {
        "T1": "投标人须提供有效营业执照。",
        "T2": "未按时缴纳投标保证金的，作废标处理。",
    }
    bid_sources = {"B1": "本页未提供营业执照。"}
    validated = app.validate_lane_payload(
        {
            "schema_version": "3.0",
            "lane": "fatal_compliance",
            "status": "complete",
            "covered_anchor_ids": [],
            "defects_list": [
                {
                    "finding_id": "F-UNRELATED-FATAL-SOURCE",
                    "module": "资格审查",
                    "check_point": "营业执照",
                    # T2 contains rejection semantics but does not back this quote.
                    "tender_source_ids": ["T1", "T2"],
                    "tender_quote": tender_sources["T1"],
                    "requirement": tender_sources["T1"],
                    "bid_source_ids": ["B1"],
                    "bid_quote": bid_sources["B1"],
                    "evidence_type": "direct",
                    "conclusion": "不符合",
                    "bid_status": bid_sources["B1"],
                    "issue": "未提供营业执照。",
                    "risk_level": "致命废标",
                    "suggestion": "补充营业执照。",
                    "confidence": "high",
                    "anchor_ids": [],
                }
            ],
            "scoring_list": [],
            "warnings": [],
        },
        "fatal_compliance",
        tender_sources,
        bid_sources,
        expected_anchor_ids=(),
    )

    row = validated["defects_list"][0]
    assert row["_manual"] is False
    assert row["_quote_backed_tender_source_ids"] == ["T1"]
    assert row["风险等级"] == "扣分/瑕疵"
    assert row["风险等级"] != "致命/废标风险"


def test_one_character_quote_is_manual_and_covers_no_anchor() -> None:
    """A ubiquitous one-character match has no evidentiary value in v3."""

    tender_sources = {
        "T1": "投标人须提供有效营业执照。",
        "T2": "投标人须提交财务审计报告。",
    }
    bid_sources = {"B1": "已提交相关资格材料。"}
    validated = app.validate_lane_payload(
        {
            "schema_version": "3.0",
            "lane": "fatal_compliance",
            "status": "complete",
            "covered_anchor_ids": ["A1", "A2"],
            "defects_list": [
                {
                    "finding_id": "F-ONE-CHAR-QUOTE",
                    "module": "资格审查",
                    "check_point": "资格材料",
                    "tender_source_ids": ["T1", "T2"],
                    # This character appears in both sources but is below the
                    # minimum quote-information threshold.
                    "tender_quote": "须",
                    "requirement": "核对资格材料。",
                    "bid_source_ids": ["B1"],
                    "bid_quote": bid_sources["B1"],
                    "evidence_type": "direct",
                    "conclusion": "符合",
                    "bid_status": bid_sources["B1"],
                    "issue": "未发现文字偏差。",
                    "risk_level": "正常",
                    "suggestion": "保持现状。",
                    "confidence": "high",
                    "anchor_ids": ["A1", "A2"],
                }
            ],
            "scoring_list": [],
            "warnings": [],
        },
        "fatal_compliance",
        tender_sources,
        bid_sources,
        expected_anchor_ids=("A1", "A2"),
        expected_anchor_sources={"A1": "T1", "A2": "T2"},
    )

    row = validated["defects_list"][0]
    assert row["_manual"] is True
    assert row["风险等级"] == "待人工复核"
    assert validated["covered_anchor_ids"] == []
    assert validated["missing_anchor_ids"] == ["A1", "A2"]
    assert validated["status"] == "partial"
