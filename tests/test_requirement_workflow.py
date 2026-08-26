"""Requirement-centric audit workflow contract tests.

All tests are offline.  They deliberately exercise the public workflow helpers
instead of duplicating their implementation in test code.
"""

from __future__ import annotations

from dataclasses import asdict, is_dataclass
from pathlib import Path
import sys
from typing import Any

import pytest
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import app


REQUIREMENT_FIELDS = {
    "requirement_id",
    "kind",
    "module",
    "title",
    "requirement_text",
    "mandatory",
    "risk_hint",
    "full_score",
    "scoring_rule",
    "source_ids",
    "source_excerpt",
    "origin_chunk_id",
}

ASSESSMENT_FIELDS = {
    "requirement_id",
    "status",
    "bid_source_ids",
    "bid_excerpt",
    "issue",
    "risk_level",
    "recommendation",
    "estimated_score",
    "score_reason",
}


def record(value: Any) -> dict[str, Any]:
    """Read either the workflow's dataclass/model object or a mapping."""

    if isinstance(value, dict):
        return value
    if is_dataclass(value):
        return asdict(value)
    if hasattr(value, "model_dump"):
        return value.model_dump()
    return {name: getattr(value, name) for name in dir(value) if not name.startswith("_")}


def make_requirement(requirement_id: str, **overrides: Any) -> dict[str, Any]:
    result = {
        "requirement_id": requirement_id,
        "kind": "check",
        "module": "资格审查",
        "title": "营业执照",
        "requirement_text": "投标人须提供有效营业执照。",
        "mandatory": True,
        "risk_hint": "废标",
        "full_score": None,
        "scoring_rule": "",
        "source_ids": ["T-P00001"],
        "source_excerpt": "投标人须提供有效营业执照。",
        "origin_chunk_id": "T-C0001",
    }
    result.update(overrides)
    return result


def make_assessment(requirement_id: str, **overrides: Any) -> dict[str, Any]:
    result = {
        "requirement_id": requirement_id,
        "status": "compliant",
        "bid_source_ids": ["B-P00001"],
        "bid_excerpt": "本公司营业执照在有效期内。",
        "issue": "符合",
        "risk_level": "正常/符合",
        "recommendation": "无需修改",
        "estimated_score": None,
        "score_reason": "",
    }
    result.update(overrides)
    return result


def test_source_blocks_preserve_every_source_id_and_order() -> None:
    text = "\n".join(
        [
            "【P00001】第一条资格要求",
            "【T002-R003】C1: 评分项 | C2: 10分",
            "【TB001】文本框中的承诺",
            "【FN7】脚注不得偏离",
        ]
    )

    blocks = app.parse_source_blocks(text, "招标文件")
    values = [record(block) for block in blocks]

    assert [item["source_id"] for item in values] == [
        "P00001",
        "T002-R003",
        "TB001",
        "FN7",
    ]
    assert all(item["text"].strip() for item in values)
    assert len({item["source_id"] for item in values}) == len(values)


def test_dedupe_requirements_never_drops_distinct_requirement_ids() -> None:
    requirements = [
        make_requirement("REQ-0002", title="业绩", requirement_text="至少提供2项业绩"),
        make_requirement("REQ-0001"),
        # Semantically similar but independently traceable: must not disappear.
        make_requirement(
            "REQ-0003",
            title="业绩证明",
            requirement_text="业绩须附合同关键页",
            source_ids=["T-P00009"],
        ),
        # Exact duplicate ID may merge sources, but must remain one requirement.
        make_requirement("REQ-0001", source_ids=["T-P00001", "T-P00002"]),
    ]

    result = [record(item) for item in app.dedupe_requirements(requirements)]
    ids = [item["requirement_id"] for item in result]

    assert ids == ["REQ-0001", "REQ-0002", "REQ-0003"]
    assert set(result[0]["source_ids"]) == {"T-P00001", "T-P00002"}


def test_bid_index_retrieval_is_deterministic_and_source_backed() -> None:
    bid_text = "\n".join(
        [
            "【P00001】营业执照统一社会信用代码 91330000TEST000001",
            "【P00002】项目经理具有一级建造师证书",
            "【T001-R002】C1: 类似业绩 | C2: 三项合同",
            "【P00003】投标有效期为90日",
        ]
    )
    blocks = app.parse_source_blocks(bid_text, "投标文件")
    index = app.build_bid_index(blocks)
    requirement = make_requirement(
        "REQ-0100",
        title="一级建造师",
        requirement_text="项目经理须具有一级建造师证书",
    )

    first = [record(hit) for hit in app.retrieve_bid_evidence(requirement, index, top_k=2)]
    second = [record(hit) for hit in app.retrieve_bid_evidence(requirement, index, top_k=2)]

    assert first == second
    assert 1 <= len(first) <= 2
    assert first[0]["source_id"] == "P00002"
    assert "一级建造师" in first[0]["text"]
    assert all(hit["source_id"] in {"P00001", "P00002", "T001-R002", "P00003"} for hit in first)
    scores = [float(hit["score"]) for hit in first]
    assert scores == sorted(scores, reverse=True)


def test_merge_preserves_all_requirement_ids_sorts_and_fills_missing() -> None:
    requirements = [
        make_requirement("REQ-0003", title="服务方案"),
        make_requirement("REQ-0001", title="营业执照"),
        make_requirement(
            "REQ-0002",
            kind="scoring",
            title="业绩评分",
            full_score=10,
            scoring_rule="每项2分，最高10分",
        ),
    ]
    assessments = [
        make_assessment("REQ-0003", status="noncompliant", issue="未响应", risk_level="扣分"),
        make_assessment("REQ-0001"),
        # REQ-0002 deliberately absent: merge must create an uncertain placeholder.
    ]

    result = app.merge_assessments(requirements, assessments)
    rows = result.get("requirement_assessments") or result.get("assessments")
    assert rows is not None
    values = [record(row) for row in rows]

    assert [item["requirement_id"] for item in values] == ["REQ-0001", "REQ-0002", "REQ-0003"]
    assert len(values) == len(requirements)
    assert len({item["requirement_id"] for item in values}) == len(requirements)
    missing = next(item for item in values if item["requirement_id"] == "REQ-0002")
    assert missing["status"] == "uncertain"
    assert "人工" in (missing.get("issue", "") + missing.get("recommendation", ""))


@pytest.mark.parametrize("bad_score", [-0.01, 10.01])
def test_merge_rejects_out_of_range_score_as_uncertain(bad_score: float) -> None:
    requirement = make_requirement(
        "REQ-S001",
        kind="scoring",
        title="技术评分",
        full_score=10,
        scoring_rule="最高10分",
    )
    assessment = make_assessment(
        "REQ-S001",
        estimated_score=bad_score,
        score_reason="模型估分",
    )

    result = app.merge_assessments([requirement], [assessment])
    rows = result.get("requirement_assessments") or result.get("assessments")
    item = record(rows[0])

    assert item["status"] == "uncertain"
    assert item["estimated_score"] in (None, "", "待人工复核")


def test_requirement_result_can_generate_and_reopen_excel() -> None:
    requirements = [
        make_requirement("REQ-0001"),
        make_requirement(
            "REQ-0002",
            kind="scoring",
            title="业绩评分",
            full_score=10,
            scoring_rule="每项2分，最高10分",
        ),
    ]
    assessments = [
        make_assessment("REQ-0001"),
        make_assessment("REQ-0002", estimated_score=8, score_reason="提供四项业绩"),
    ]
    merged = app.merge_assessments(requirements, assessments)

    report = app.build_excel_report(merged)
    assert report.getbuffer().nbytes > 0
    workbook = load_workbook(report, data_only=False)

    assert set(workbook.sheetnames) == {"缺陷核查记录", "预估打分表"}
    assert workbook["缺陷核查记录"].max_row >= 2
    assert workbook["预估打分表"].max_row >= 2
    workbook.close()


def test_requirement_extraction_bisects_failed_batch_without_losing_source_blocks(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A permanently bad leaf becomes a review item; sibling work still succeeds."""

    blocks = [
        {
            "source_id": f"P{ordinal:05d}",
            "text": f"Tender atomic clause number {ordinal:02d}.",
            "block_type": "paragraph",
            "ordinal": ordinal,
        }
        for ordinal in range(1, 5)
    ]
    poison_source_id = "P00002"
    calls: list[tuple[str, ...]] = []
    logs: list[str] = []
    fake_client = object()

    def fake_extract_unit(
        client: object,
        model: str,
        work_blocks: list[dict[str, Any]],
        logger: Any,
        batch_id: str,
        table_headers: dict[str, str] | None = None,
    ) -> list[dict[str, Any]]:
        assert client is fake_client
        assert model == "offline-mock"
        assert table_headers is None or isinstance(table_headers, dict)
        source_ids = tuple(str(block["source_id"]) for block in work_blocks)
        calls.append(source_ids)

        # Simulate an over-large response, plus one source block that keeps
        # failing even after recursive bisection reaches a single-item leaf.
        if len(work_blocks) > 2 or poison_source_id in source_ids:
            raise app.ModelOutputError("synthetic malformed JSON")

        return [
            make_requirement(
                f"REQ-{block['source_id']}",
                title=str(block["text"]),
                requirement_text=str(block["text"]),
                source_ids=[str(block["source_id"])],
                source_excerpt=str(block["text"]),
                origin_chunk_id=batch_id,
                ordinal=int(block["ordinal"]),
            )
            for block in work_blocks
        ]

    monkeypatch.setattr(app, "_extract_requirement_unit", fake_extract_unit)

    extracted = [
        record(item)
        for item in app.extract_requirements_resilient(
            fake_client,
            "offline-mock",
            blocks,
            logs.append,
        )
    ]

    expected_sources = [block["source_id"] for block in blocks]
    covered_sources = [source_id for item in extracted for source_id in item["source_ids"]]
    assert covered_sources == expected_sources
    assert len(extracted) == len(blocks)
    assert len({item["requirement_id"] for item in extracted}) == len(blocks)

    failed_leaf = next(item for item in extracted if item["source_ids"] == [poison_source_id])
    assert failed_leaf["risk_hint"] == "待人工复核"
    assert failed_leaf["requirement_id"]

    # The call trace proves that the original batch was retried as smaller
    # independent work units instead of discarding the whole result.
    assert calls[0] == tuple(expected_sources)
    assert (poison_source_id,) in calls
    assert any(1 < len(call) < len(blocks) for call in calls)
    assert logs


def test_assessment_bisection_preserves_every_requirement_and_marks_failed_leaf(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Every accepted requirement gets exactly one result despite local failure."""

    requirements = [
        make_requirement(
            f"REQ-{ordinal:04d}",
            title=f"Atomic requirement {ordinal:02d}",
            requirement_text=f"Atomic requirement {ordinal:02d}",
            source_ids=[f"P{ordinal:05d}"],
            ordinal=ordinal,
        )
        for ordinal in range(1, 5)
    ]
    poison_requirement_id = "REQ-0002"
    evidence_map = {
        requirement["requirement_id"]: [
            {
                "source_id": f"B-P{ordinal:05d}",
                "text": f"Bid evidence {ordinal:02d}",
                "score": 1.0,
            }
        ]
        for ordinal, requirement in enumerate(requirements, start=1)
    }
    calls: list[tuple[str, ...]] = []
    logs: list[str] = []
    fake_client = object()

    def fake_assess_unit(
        client: object,
        model: str,
        work_requirements: list[dict[str, Any]],
        work_evidence_map: dict[str, list[dict[str, Any]]],
        logger: Any,
    ) -> list[dict[str, Any]]:
        assert client is fake_client
        assert model == "offline-mock"
        assert work_evidence_map is evidence_map
        requirement_ids = tuple(str(item["requirement_id"]) for item in work_requirements)
        calls.append(requirement_ids)

        if len(work_requirements) > 2 or poison_requirement_id in requirement_ids:
            raise app.ModelOutputError("synthetic assessment failure")

        return [make_assessment(requirement_id) for requirement_id in requirement_ids]

    monkeypatch.setattr(app, "_assess_requirement_unit", fake_assess_unit)

    assessments = [
        record(item)
        for item in app.assess_requirement_batch_resilient(
            fake_client,
            "offline-mock",
            requirements,
            evidence_map,
            logs.append,
        )
    ]

    expected_ids = [item["requirement_id"] for item in requirements]
    actual_ids = [item["requirement_id"] for item in assessments]
    assert actual_ids == expected_ids
    assert len(actual_ids) == len(set(actual_ids))

    failed_leaf = next(item for item in assessments if item["requirement_id"] == poison_requirement_id)
    assert failed_leaf["status"] == "uncertain"
    assert failed_leaf["risk_level"] == "待人工复核"
    assert failed_leaf["estimated_score"] is None

    assert calls[0] == tuple(expected_ids)
    assert (poison_requirement_id,) in calls
    assert any(1 < len(call) < len(requirements) for call in calls)
    assert logs


def test_normalize_requirement_rejects_excerpt_not_found_in_cited_source() -> None:
    blocks = [
        {
            "source_id": "P00001",
            "text": "The tender requires an ISO 9001 certificate valid through 2028.",
            "block_type": "paragraph",
            "ordinal": 1,
        }
    ]
    payload = {
        "status": "complete",
        "block_reviews": [
            {
                "source_id": "P00001",
                "disposition": "extracted",
                "local_requirement_ids": ["q1"],
                "note": "",
            }
        ],
        "requirements": [
            {
                "local_id": "q1",
                "kind": "check",
                "module": "qualification",
                "title": "ISO certificate",
                "requirement_text": "Provide a valid ISO 9001 certificate.",
                "mandatory": True,
                "risk_hint": "rejection",
                "full_score": None,
                "scoring_rule": "",
                "source_ids": ["P00001"],
                # Plausible model prose, but it is not a verbatim substring of
                # the only cited source and therefore must not become evidence.
                "source_excerpt": "ISO 27001 certification is mandatory.",
                "search_terms": ["ISO 9001"],
            }
        ],
    }

    with pytest.raises(app.ModelOutputError, match="原文摘录"):
        app._normalize_requirement_payload(payload, blocks, "TB-OFFLINE")


@pytest.mark.parametrize("disposition", ["extracted", "uncertain"])
def test_normalize_requirement_rejects_review_without_required_item(
    disposition: str,
) -> None:
    blocks = [{"source_id": "P00001", "text": "A clause requiring review.", "ordinal": 1}]
    payload = {
        "status": "complete",
        "block_reviews": [
            {
                "source_id": "P00001",
                "disposition": disposition,
                "local_requirement_ids": [],
                "note": "",
            }
        ],
        "requirements": [],
    }
    with pytest.raises(app.ModelOutputError):
        app._normalize_requirement_payload(payload, blocks, "TB-EMPTY")


def test_normalize_requirement_rejects_source_edge_mismatch() -> None:
    blocks = [
        {"source_id": "P00001", "text": "Provide certificate A.", "ordinal": 1},
        {"source_id": "P00002", "text": "Provide certificate B.", "ordinal": 2},
    ]
    payload = {
        "status": "complete",
        "block_reviews": [
            {"source_id": "P00001", "disposition": "extracted", "local_requirement_ids": ["q1"]},
            {"source_id": "P00002", "disposition": "no_requirement", "local_requirement_ids": []},
        ],
        "requirements": [
            {
                "local_id": "q1",
                "kind": "check",
                "module": "qualification",
                "title": "certificate B",
                "requirement_text": "Provide certificate B.",
                "mandatory": True,
                "risk_hint": "rejection",
                "full_score": None,
                "scoring_rule": "",
                "source_ids": ["P00002"],
                "source_excerpt": "Provide certificate B.",
                "search_terms": ["certificate B"],
            }
        ],
    }
    with pytest.raises(app.ModelOutputError, match="逐项闭合"):
        app._normalize_requirement_payload(payload, blocks, "TB-EDGE")


def test_uncertain_requirement_is_forced_to_manual_assessment(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    blocks = [{"source_id": "P00001", "text": "A clause requiring review.", "ordinal": 1}]
    payload = {
        "status": "complete",
        "block_reviews": [
            {
                "source_id": "P00001",
                "disposition": "uncertain",
                "local_requirement_ids": ["q1"],
            }
        ],
        "requirements": [
            {
                "local_id": "q1",
                "kind": "check",
                "module": "qualification",
                "title": "review clause",
                "requirement_text": "A clause requiring review.",
                "mandatory": False,
                "risk_hint": "",
                "full_score": None,
                "scoring_rule": "",
                "source_ids": ["P00001"],
                "source_excerpt": "A clause requiring review.",
                "search_terms": [],
            }
        ],
    }
    requirement = app._normalize_requirement_payload(payload, blocks, "TB-UNCERTAIN")[0]
    assert requirement["extraction_uncertain"] is True

    monkeypatch.setattr(
        app,
        "_assess_requirement_unit",
        lambda *args, **kwargs: [make_assessment(requirement["requirement_id"])],
    )
    result = app.assess_requirement_batch_resilient(
        object(),
        "offline-mock",
        [requirement],
        {requirement["requirement_id"]: []},
        lambda _: None,
    )[0]
    assert result["status"] == "uncertain"
    assert result["risk_level"] == "待人工复核"


def test_normalize_assessment_requires_source_backed_excerpt_for_definite_result() -> None:
    requirements = [make_requirement("REQ-EVIDENCE")]
    evidence_map = {
        "REQ-EVIDENCE": [
            {
                "source_id": "B-P00001",
                "text": "The proposed project manager holds a Class I constructor certificate valid through 2028.",
                "score": 8.5,
                "ordinal": 1,
            }
        ]
    }
    assessment = {
        "requirement_id": "REQ-EVIDENCE",
        "status": "compliant",
        "bid_source_ids": ["B-P00001"],
        "bid_excerpt": "The proposed project manager holds a senior engineer certificate.",
        "issue": "compliant",
        "risk_level": "normal",
        "recommendation": "none",
        "estimated_score": None,
        "score_reason": "",
    }

    with pytest.raises(app.ModelOutputError, match="原文摘录"):
        app._normalize_assessment_payload(
            {"status": "complete", "assessments": [assessment]},
            requirements,
            evidence_map,
        )

    assessment["bid_excerpt"] = "Class I constructor certificate valid through 2028"
    normalized = app._normalize_assessment_payload(
        {"status": "complete", "assessments": [assessment]},
        requirements,
        evidence_map,
    )

    assert len(normalized) == 1
    assert normalized[0]["requirement_id"] == "REQ-EVIDENCE"
    assert normalized[0]["status"] == "compliant"
    assert normalized[0]["bid_source_ids"] == ["B-P00001"]
    assert normalized[0]["bid_excerpt"] == assessment["bid_excerpt"]


def test_retrieve_bid_evidence_returns_empty_when_no_term_or_number_matches() -> None:
    bid_blocks = [
        {
            "source_id": "B-P00001",
            "text": "General company registration and organization profile.",
            "block_type": "paragraph",
            "ordinal": 1,
        },
        {
            "source_id": "B-P00002",
            "text": "Service team responsibilities and delivery methodology.",
            "block_type": "paragraph",
            "ordinal": 2,
        },
    ]
    index = app.build_bid_index(bid_blocks)
    unmatched_requirement = {
        "module": "",
        "title": "",
        "requirement_text": "ZXQVVV-991",
        "scoring_rule": "",
        "source_excerpt": "",
    }

    assert app.retrieve_bid_evidence(unmatched_requirement, index, top_k=2) == []


@pytest.mark.parametrize(
    ("case", "assessments", "error_fragment"),
    [
        (
            "unknown requirement id",
            [make_assessment("REQ-UNKNOWN")],
            "未知 requirement_id",
        ),
        (
            "duplicate requirement id",
            [make_assessment("REQ-0001"), make_assessment("REQ-0001")],
            "重复返回 requirement_id",
        ),
    ],
    ids=lambda value: value if isinstance(value, str) and " id" in value else None,
)
def test_merge_assessments_rejects_unknown_or_duplicate_requirement_ids(
    case: str,
    assessments: list[dict[str, Any]],
    error_fragment: str,
) -> None:
    del case  # Used only to provide an explicit pytest case label.
    requirements = [make_requirement("REQ-0001")]

    with pytest.raises(app.ModelOutputError, match=error_fragment):
        app.merge_assessments(requirements, assessments)


def test_full_scan_bisects_failed_batch_and_keeps_successful_sibling_hits(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    requirement = make_requirement(
        "REQ-FULL-SCAN",
        requirement_text="Locate the exact bid response anywhere in the document.",
    )
    blocks = [
        {
            "source_id": f"B-P{ordinal:05d}",
            "text": f"Bid source block {ordinal:02d} contains reviewable evidence.",
            "block_type": "paragraph",
            "ordinal": ordinal,
        }
        for ordinal in range(1, 5)
    ]
    poison_source_id = "B-P00002"
    calls: list[tuple[str, ...]] = []
    logs: list[str] = []
    fake_client = object()

    def fake_scan_unit(
        client: object,
        model: str,
        work_requirement: dict[str, Any],
        work_blocks: list[dict[str, Any]],
        logger: Any,
        table_headers: dict[str, str] | None = None,
    ) -> list[dict[str, Any]]:
        assert client is fake_client
        assert model == "offline-mock"
        assert table_headers is None or isinstance(table_headers, dict)
        assert work_requirement["requirement_id"] == "REQ-FULL-SCAN"
        source_ids = tuple(str(block["source_id"]) for block in work_blocks)
        calls.append(source_ids)

        # The first request is too large. One leaf remains malformed after
        # bisection, while every normal sibling can still return useful hits.
        if len(work_blocks) > 2 or poison_source_id in source_ids:
            raise app.ModelOutputError("synthetic full-scan failure")

        return [
            {
                "source_id": str(block["source_id"]),
                "text": str(block["text"]),
                "score": 100.0,
                "ordinal": int(block["ordinal"]),
                "scan_reason": "offline verified hit",
                "verified_for": "REQ-FULL-SCAN",
            }
            for block in work_blocks
        ]

    monkeypatch.setattr(app, "_scan_bid_unit_for_requirement", fake_scan_unit)

    hits, complete = app.scan_bid_evidence_resilient(
        fake_client,
        "offline-mock",
        requirement,
        blocks,
        logs.append,
    )

    expected_surviving_ids = ["B-P00001", "B-P00003", "B-P00004"]
    assert complete is False
    assert [hit["source_id"] for hit in hits] == expected_surviving_ids
    assert all(hit["verified_for"] == "REQ-FULL-SCAN" for hit in hits)
    assert calls[0] == tuple(block["source_id"] for block in blocks)
    assert (poison_source_id,) in calls
    assert any(1 < len(call) < len(blocks) for call in calls)
    assert logs


@pytest.mark.parametrize(
    "invalid_hit",
    [
        {
            "source_id": "B-P00001",
            "excerpt": "A fabricated certificate that does not occur in the source.",
            "reason": "looks relevant",
        },
        {
            "source_id": "B-P99999",
            "excerpt": "Class I constructor certificate valid through 2028",
            "reason": "unknown source",
        },
    ],
    ids=["fabricated-excerpt", "unknown-source-id"],
)
def test_normalize_full_scan_rejects_unverifiable_hits(
    invalid_hit: dict[str, Any],
) -> None:
    blocks = [
        {
            "source_id": "B-P00001",
            "text": "The project manager has a Class I constructor certificate valid through 2028.",
            "block_type": "paragraph",
            "ordinal": 1,
        }
    ]

    with pytest.raises(app.ModelOutputError):
        app._normalize_full_scan_payload(
            {
                "status": "complete",
                "reviewed_source_ids": ["B-P00001"],
                "reviewed_requirement_ids": ["REQ-FULL-SCAN"],
                "hits": [invalid_hit],
            },
            "REQ-FULL-SCAN",
            blocks,
        )


def test_normalize_full_scan_accepts_verbatim_excerpt_and_restores_full_block() -> None:
    source_text = "The project manager has a Class I constructor certificate valid through 2028."
    blocks = [
        {
            "source_id": "B-P00001",
            "text": source_text,
            "block_type": "paragraph",
            "ordinal": 7,
        }
    ]
    payload = {
        "status": "complete",
        "reviewed_source_ids": ["B-P00001"],
        "reviewed_requirement_ids": ["REQ-FULL-SCAN"],
        "hits": [
            {
                "source_id": "B-P00001",
                "excerpt": "Class I constructor certificate valid through 2028",
                "reason": "direct qualification evidence",
            }
        ],
    }

    normalized = app._normalize_full_scan_payload(
        payload,
        "REQ-FULL-SCAN",
        blocks,
    )

    assert normalized == [
        {
            "source_id": "B-P00001",
            "text": source_text,
            "score": 100.0,
            "ordinal": 7,
            "scan_reason": "direct qualification evidence",
            "verified_for": "REQ-FULL-SCAN",
        }
    ]


@pytest.mark.parametrize(
    "payload",
    [
        {
            "status": "complete",
            "reviewed_source_ids": [],
            "reviewed_requirement_ids": ["REQ-FULL-SCAN"],
            "hits": [],
        },
        {
            "status": "complete",
            "reviewed_source_ids": ["B-P00001"],
            "reviewed_requirement_ids": [],
            "hits": [],
        },
    ],
    ids=["missing-source-coverage", "missing-requirement-coverage"],
)
def test_normalize_full_scan_rejects_incomplete_coverage_ledgers(
    payload: dict[str, Any],
) -> None:
    blocks = [{"source_id": "B-P00001", "text": "Bid evidence", "ordinal": 1}]
    with pytest.raises(app.ModelOutputError):
        app._normalize_full_scan_payload(payload, "REQ-FULL-SCAN", blocks)


def test_merge_evidence_prefers_scan_hit_deduplicates_and_respects_char_limit() -> None:
    scan_hits = [
        {
            "source_id": "B-P00001",
            "text": "SCAN-A",
            "score": 1.0,
            "ordinal": 1,
            "verified_for": "REQ-FULL-SCAN",
        },
        {
            "source_id": "B-P00002",
            "text": "SCAN-B",
            "score": 0.5,
            "ordinal": 2,
            "verified_for": "REQ-FULL-SCAN",
        },
    ]
    lexical_hits = [
        {
            # Same source with a much higher lexical score must not overwrite
            # the already verified full-scan representation.
            "source_id": "B-P00001",
            "text": "LEXICAL-DUPLICATE",
            "score": 999.0,
            "ordinal": 1,
        },
        {
            "source_id": "B-P00003",
            "text": "LEXICAL",
            "score": 500.0,
            "ordinal": 3,
        },
    ]

    merged = app._merge_evidence_hits(lexical_hits, scan_hits, max_chars=12)

    assert [hit["source_id"] for hit in merged] == ["B-P00001", "B-P00002"]
    assert merged[0]["text"] == "SCAN-A"
    assert len({hit["source_id"] for hit in merged}) == len(merged)
    assert sum(len(str(hit["text"])) for hit in merged) <= 12
    assert all(hit.get("verified_for") == "REQ-FULL-SCAN" for hit in merged)

    assert app._merge_evidence_hits(
        [{"source_id": "B-LONG", "text": "X" * 20, "score": 1, "ordinal": 1}],
        [],
        max_chars=12,
    ) == []


def test_shared_full_scan_splits_requirements_before_blocks_and_isolates_poison_leaf(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A bad (requirement, block) leaf must not erase unrelated shared-scan work."""

    requirements = [
        make_requirement(
            requirement_id,
            title=f"Shared scan requirement {ordinal}",
            requirement_text=f"Find evidence for shared scan requirement {ordinal}.",
            ordinal=ordinal,
        )
        for ordinal, requirement_id in enumerate(
            ["REQ-NORMAL-A", "REQ-POISON", "REQ-NORMAL-B", "REQ-NORMAL-C"],
            start=1,
        )
    ]
    blocks = [
        {
            "source_id": f"B-P{ordinal:05d}",
            "text": f"Shared bid block {ordinal} contains independently reviewable evidence.",
            "block_type": "paragraph",
            "ordinal": ordinal,
        }
        for ordinal in range(1, 5)
    ]
    poison_requirement_id = "REQ-POISON"
    poison_source_id = "B-P00002"
    all_requirement_ids = tuple(item["requirement_id"] for item in requirements)
    all_source_ids = tuple(item["source_id"] for item in blocks)
    calls: list[tuple[tuple[str, ...], tuple[str, ...]]] = []
    logs: list[str] = []
    fake_client = object()

    def fake_group_scan_unit(
        client: object,
        model: str,
        work_requirements: list[dict[str, Any]],
        work_blocks: list[dict[str, Any]],
        logger: Any,
        table_headers: dict[str, str] | None = None,
    ) -> dict[str, list[dict[str, Any]]]:
        assert client is fake_client
        assert model == "offline-mock"
        assert table_headers is None or isinstance(table_headers, dict)
        requirement_ids = tuple(str(item["requirement_id"]) for item in work_requirements)
        source_ids = tuple(str(item["source_id"]) for item in work_blocks)
        calls.append((requirement_ids, source_ids))

        # Any multi-requirement response fails, proving that the resilient
        # wrapper first splits the requirement axis while retaining all blocks.
        if len(work_requirements) > 1:
            raise app.ModelOutputError("synthetic shared-response failure")

        requirement_id = requirement_ids[0]
        if requirement_id == poison_requirement_id and poison_source_id in source_ids:
            raise app.ModelOutputError("synthetic poison pair")

        return {
            requirement_id: [
                {
                    "source_id": str(block["source_id"]),
                    "text": str(block["text"]),
                    "score": 100.0,
                    "ordinal": int(block["ordinal"]),
                    "scan_reason": "offline shared-scan hit",
                    "verified_for": requirement_id,
                }
                for block in work_blocks
            ]
        }

    monkeypatch.setattr(app, "_scan_bid_unit_for_requirements", fake_group_scan_unit)

    hits_by_requirement, complete_by_requirement = app.scan_bid_evidence_group_resilient(
        fake_client,
        "offline-mock",
        requirements,
        blocks,
        logs.append,
    )

    assert tuple(hits_by_requirement) == all_requirement_ids
    assert tuple(complete_by_requirement) == all_requirement_ids
    assert complete_by_requirement == {
        "REQ-NORMAL-A": True,
        "REQ-POISON": False,
        "REQ-NORMAL-B": True,
        "REQ-NORMAL-C": True,
    }
    for requirement_id in ("REQ-NORMAL-A", "REQ-NORMAL-B", "REQ-NORMAL-C"):
        assert [hit["source_id"] for hit in hits_by_requirement[requirement_id]] == list(all_source_ids)
        assert all(hit["verified_for"] == requirement_id for hit in hits_by_requirement[requirement_id])
    assert [hit["source_id"] for hit in hits_by_requirement[poison_requirement_id]] == [
        "B-P00001",
        "B-P00003",
        "B-P00004",
    ]

    assert calls[0] == (all_requirement_ids, all_source_ids)
    # Requirement groups must never be split on the block axis. Only after a
    # single requirement remains may the wrapper recursively split blocks.
    assert all(
        source_ids == all_source_ids
        for requirement_ids, source_ids in calls
        if len(requirement_ids) > 1
    )
    assert all(
        len(requirement_ids) == 1
        for requirement_ids, source_ids in calls
        if source_ids != all_source_ids
    )
    assert ((poison_requirement_id,), (poison_source_id,)) in calls
    assert logs


@pytest.mark.parametrize(
    ("reviewed_source_ids", "reviewed_requirement_ids", "hits"),
    [
        (["B-P00001"], ["REQ-A", "REQ-B", "REQ-C"], []),
        (["B-P00002", "B-P00001"], ["REQ-A", "REQ-B", "REQ-C"], []),
        (["B-P00001", "B-P00002"], ["REQ-A", "REQ-C"], []),
        (["B-P00001", "B-P00002"], ["REQ-B", "REQ-A", "REQ-C"], []),
        (["B-P00001", "B-P00002"], ["REQ-A", "REQ-B", "REQ-UNKNOWN"], []),
        (
            ["B-P00001", "B-P00002"],
            ["REQ-A", "REQ-B", "REQ-C"],
            [
                {
                    "requirement_id": "REQ-UNKNOWN",
                    "source_id": "B-P00001",
                    "excerpt": "alpha evidence",
                    "reason": "unknown requirement hit",
                }
            ],
        ),
    ],
    ids=[
        "missing-reviewed-source",
        "reordered-reviewed-sources",
        "missing-reviewed-requirement",
        "reordered-reviewed-requirements",
        "unknown-reviewed-requirement",
        "unknown-hit-requirement",
    ],
)
def test_normalize_shared_full_scan_rejects_inexact_coverage_or_unknown_requirement(
    reviewed_source_ids: list[str],
    reviewed_requirement_ids: list[str],
    hits: list[dict[str, Any]],
) -> None:
    blocks = [
        {"source_id": "B-P00001", "text": "alpha evidence", "ordinal": 1},
        {"source_id": "B-P00002", "text": "gamma evidence", "ordinal": 2},
    ]
    payload = {
        "status": "complete",
        "reviewed_source_ids": reviewed_source_ids,
        "reviewed_requirement_ids": reviewed_requirement_ids,
        "hits": hits,
    }

    with pytest.raises(app.ModelOutputError):
        app._normalize_full_scan_payload(
            payload,
            ["REQ-A", "REQ-B", "REQ-C"],
            blocks,
        )


def test_normalize_shared_full_scan_accepts_sparse_hits_and_keeps_empty_requirements() -> None:
    blocks = [
        {
            "source_id": "B-P00001",
            "text": "The alpha certificate remains valid through 2028.",
            "ordinal": 1,
        },
        {
            "source_id": "B-P00002",
            "text": "The gamma response includes three completed contracts.",
            "ordinal": 2,
        },
    ]
    requirement_ids = ["REQ-A", "REQ-B", "REQ-C"]
    payload = {
        "status": "complete",
        "reviewed_source_ids": ["B-P00001", "B-P00002"],
        "reviewed_requirement_ids": requirement_ids,
        # Sparse output is valid: reviewed requirements with no evidence do
        # not need fake placeholder hit objects.
        "hits": [
            {
                "requirement_id": "REQ-A",
                "source_id": "B-P00001",
                "excerpt": "alpha certificate remains valid through 2028",
                "reason": "direct certificate evidence",
            },
            {
                "requirement_id": "REQ-C",
                "source_id": "B-P00002",
                "excerpt": "three completed contracts",
                "reason": "direct experience evidence",
            },
        ],
    }

    normalized = app._normalize_full_scan_payload(payload, requirement_ids, blocks)

    assert list(normalized) == requirement_ids
    assert normalized["REQ-B"] == []
    assert [hit["source_id"] for hit in normalized["REQ-A"]] == ["B-P00001"]
    assert [hit["source_id"] for hit in normalized["REQ-C"]] == ["B-P00002"]
    assert normalized["REQ-A"][0]["verified_for"] == "REQ-A"
    assert normalized["REQ-C"][0]["verified_for"] == "REQ-C"
    assert normalized["REQ-A"][0]["text"] == blocks[0]["text"]
    assert normalized["REQ-C"][0]["text"] == blocks[1]["text"]


def test_table_batch_injects_row_one_header_as_context_without_polluting_reviewed_ids() -> None:
    table_blocks = [
        {
            "source_id": "T007-R001",
            "text": "评分项目 | 评分标准 | 满分",
            "block_type": "table_row",
            "ordinal": 1,
        },
        {
            "source_id": "T007-R002",
            "text": "项目业绩 | 每项得2分 | 10分",
            "block_type": "table_row",
            "ordinal": 2,
        },
        {
            "source_id": "T007-R003",
            "text": "项目团队 | 每证得1分 | 5分",
            "block_type": "table_row",
            "ordinal": 3,
        },
    ]
    table_headers = app._build_table_header_context(table_blocks)
    batches = app.make_structure_batches(table_blocks, max_chars=10_000, max_blocks=1)
    second_batch = batches[1]

    assert table_headers == {"T007": "评分项目 | 评分标准 | 满分"}
    assert [block["source_id"] for block in second_batch] == ["T007-R002"]

    rendered = app._render_blocks_for_prompt(second_batch, table_headers)
    assert "<CONTEXT-ONLY table=T007 header> 评分项目 | 评分标准 | 满分" in rendered
    assert "仅供理解列含义，不得引用为 source_id" in rendered
    assert "<T007-R002> 项目业绩 | 每项得2分 | 10分" in rendered
    assert "<T007-R001>" not in rendered

    # Although R001 text crosses the batch boundary as context, the coverage
    # ledger must contain only the primary R002 source actually being scanned.
    valid_payload = {
        "status": "complete",
        "reviewed_source_ids": ["T007-R002"],
        "reviewed_requirement_ids": ["REQ-TABLE"],
        "hits": [],
    }
    assert app._normalize_full_scan_payload(
        valid_payload,
        "REQ-TABLE",
        second_batch,
    ) == []

    polluted_payload = {
        **valid_payload,
        "reviewed_source_ids": ["T007-R001", "T007-R002"],
    }
    with pytest.raises(app.ModelOutputError):
        app._normalize_full_scan_payload(
            polluted_payload,
            "REQ-TABLE",
            second_batch,
        )


def test_excel_report_applies_risk_rows_and_global_professional_styles() -> None:
    def defect(sequence: int, risk_level: str) -> dict[str, Any]:
        return {
            "序号": sequence,
            "核查模块": "资格审查",
            "检查要点": f"风险样例 {sequence}",
            "招标文件出处": "【T-P00001】",
            "招标文件要求": "提供有效证明材料。",
            "投标文件现状": "【B-P00001】 已提供证明材料。",
            "存在问题与缺陷": risk_level,
            "风险等级": risk_level,
            "修改建议": "按风险等级处理。",
        }

    report = app.build_excel_report(
        {
            "defects_list": [
                defect(1, "致命/废标风险"),
                defect(2, "扣分/瑕疵"),
                defect(3, "正常/符合"),
            ],
            "scoring_list": [],
        }
    )
    workbook = load_workbook(report, data_only=False)

    def rgb_suffix(color: Any) -> str:
        return str(color.rgb)[-6:].upper()

    try:
        defects_sheet = workbook["缺陷核查记录"]
        scoring_sheet = workbook["预估打分表"]

        for worksheet in (defects_sheet, scoring_sheet):
            assert worksheet.freeze_panes == "A2"
            assert worksheet.auto_filter.ref == worksheet.dimensions
            for header_cell in worksheet[1]:
                assert header_cell.fill.fill_type == "solid"
                assert rgb_suffix(header_cell.fill.fgColor) == "1F4E78"
                assert header_cell.font.bold is True
                assert rgb_suffix(header_cell.font.color) == "FFFFFF"
                assert header_cell.alignment.horizontal == "center"
                assert header_cell.alignment.vertical == "center"
                assert header_cell.alignment.wrap_text is True
                assert all(
                    getattr(header_cell.border, side).style == "thin"
                    for side in ("left", "right", "top", "bottom")
                )

        expected_risk_styles = {
            2: ("C00000", "FFFFFF", True),
            3: ("F4B183", "000000", True),
            4: ("C6E0B4", "000000", False),
        }
        for row_index, (fill_color, font_color, bold) in expected_risk_styles.items():
            row = list(defects_sheet[row_index])
            assert len(row) == len(app.DEFECT_FIELDS)
            for cell in row:
                assert cell.fill.fill_type == "solid"
                assert rgb_suffix(cell.fill.fgColor) == fill_color
                assert rgb_suffix(cell.font.color) == font_color
                assert bool(cell.font.bold) is bold
                assert cell.alignment.wrap_text is True
                assert all(
                    getattr(cell.border, side).style == "thin"
                    for side in ("left", "right", "top", "bottom")
                )
    finally:
        workbook.close()


def test_scoring_without_full_score_cannot_keep_model_score_after_normalize_and_merge() -> None:
    requirement = make_requirement(
        "REQ-SCORE-NO-LIMIT",
        kind="scoring",
        full_score=None,
        scoring_rule="The model did not reliably extract a maximum score.",
    )
    evidence_text = "The bidder supplied five completed project contracts."
    evidence_map = {
        "REQ-SCORE-NO-LIMIT": [
            {
                "source_id": "B-P00001",
                "text": evidence_text,
                "score": 5.0,
                "ordinal": 1,
            }
        ]
    }
    payload = {
        "status": "complete",
        "assessments": [
            {
                "requirement_id": "REQ-SCORE-NO-LIMIT",
                "status": "compliant",
                "bid_source_ids": ["B-P00001"],
                "bid_excerpt": "five completed project contracts",
                "issue": "The model claimed a score despite an unknown ceiling.",
                "risk_level": "正常/符合",
                "recommendation": "none",
                "estimated_score": 999,
                "score_reason": "unbounded model estimate",
            }
        ],
    }

    normalized = app._normalize_assessment_payload(payload, [requirement], evidence_map)
    merged = app.merge_assessments([requirement], normalized)
    row = merged["requirement_assessments"][0]

    assert row["status"] == "uncertain"
    assert row["risk_level"] == "待人工复核"
    assert row["estimated_score"] is None
    assert "满分" in row["issue"]
    assert merged["scoring_list"][0]["当前预估得分"] == "待人工复核"


@pytest.mark.parametrize("status", ["not_found", "uncertain"])
def test_nondefinite_assessment_evidence_is_verified_or_replaced_with_fixed_placeholder(
    status: str,
) -> None:
    requirement = make_requirement(
        "REQ-NONDEFINITE",
        kind="scoring",
        full_score=10,
        scoring_rule="Up to ten points.",
    )
    evidence_text = "The submitted license remains valid through 2028."
    evidence_map = {
        "REQ-NONDEFINITE": [
            {
                "source_id": "B-P00001",
                "text": evidence_text,
                "score": 4.0,
                "ordinal": 1,
            }
        ]
    }

    def assessment(*, source_ids: list[str], excerpt: str) -> dict[str, Any]:
        return {
            "requirement_id": "REQ-NONDEFINITE",
            "status": status,
            "bid_source_ids": source_ids,
            "bid_excerpt": excerpt,
            "issue": "Evidence is not sufficient for a definite conclusion.",
            "risk_level": "待人工复核",
            "recommendation": "Review manually.",
            "estimated_score": 9,
            "score_reason": "must not survive a non-definite status",
        }

    with pytest.raises(app.ModelOutputError):
        app._normalize_assessment_payload(
            {
                "status": "complete",
                "assessments": [
                    assessment(
                        source_ids=["B-P00001"],
                        excerpt="A fabricated ISO certificate appears here.",
                    )
                ],
            },
            [requirement],
            evidence_map,
        )

    verified = app._normalize_assessment_payload(
        {
            "status": "complete",
            "assessments": [
                assessment(
                    source_ids=["B-P00001"],
                    excerpt="license remains valid through 2028",
                )
            ],
        },
        [requirement],
        evidence_map,
    )[0]
    assert verified["bid_source_ids"] == ["B-P00001"]
    assert verified["bid_excerpt"] == "license remains valid through 2028"
    assert verified["estimated_score"] is None

    no_source = app._normalize_assessment_payload(
        {
            "status": "complete",
            "assessments": [
                assessment(
                    source_ids=[],
                    excerpt="Model-authored prose must not masquerade as bid text.",
                )
            ],
        },
        [requirement],
        evidence_map,
    )[0]
    assert no_source["bid_source_ids"] == []
    assert no_source["bid_excerpt"] == "候选证据未可靠确认"
    assert no_source["estimated_score"] is None


def test_extraction_fallback_remains_manual_after_merge(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source_block = {
        "source_id": "T-P00001",
        "text": "An unreadable tender clause that could not be reliably extracted.",
        "block_type": "paragraph",
        "ordinal": 1,
    }

    def always_fail_extract(
        client: object,
        model: str,
        blocks: list[dict[str, Any]],
        logger: Any,
        batch_id: str,
        table_headers: dict[str, str] | None = None,
    ) -> list[dict[str, Any]]:
        del client, model, blocks, logger, batch_id, table_headers
        raise app.ModelOutputError("synthetic extraction leaf failure")

    monkeypatch.setattr(app, "_extract_requirement_unit", always_fail_extract)
    requirements = app.extract_requirements_resilient(
        object(),
        "offline-mock",
        [source_block],
        lambda _: None,
    )

    assert len(requirements) == 1
    fallback = requirements[0]
    assert fallback["extraction_uncertain"] is True
    assert fallback["risk_hint"] == "待人工复核"

    # Even a later model claim of definite compliance must not overwrite the
    # extraction failure marker at the final merge boundary.
    merged = app.merge_assessments(
        requirements,
        [
            make_assessment(
                fallback["requirement_id"],
                status="compliant",
                risk_level="正常/符合",
            )
        ],
    )
    row = merged["requirement_assessments"][0]
    assert row["status"] == "uncertain"
    assert row["risk_level"] == "待人工复核"
    assert row["estimated_score"] is None
    assert "抽取未可靠完成" in row["issue"]
    assert merged["defects_list"][0]["风险等级"] == "待人工复核"


def test_noncompliant_model_normal_risk_is_overridden_and_never_colored_green() -> None:
    requirement = make_requirement(
        "REQ-NONCOMPLIANT",
        mandatory=False,
        risk_hint="一般",
    )
    assessment = make_assessment(
        "REQ-NONCOMPLIANT",
        status="noncompliant",
        risk_level="正常/符合",
        issue="The response does not meet the tender requirement.",
    )

    merged = app.merge_assessments([requirement], [assessment])
    row = merged["requirement_assessments"][0]
    assert row["status"] == "noncompliant"
    assert row["risk_level"] == "扣分/瑕疵"
    assert merged["defects_list"][0]["风险等级"] == "扣分/瑕疵"

    workbook = load_workbook(app.build_excel_report(merged), data_only=False)
    try:
        first_data_row = workbook["缺陷核查记录"][2]
        assert all(str(cell.fill.fgColor.rgb)[-6:].upper() == "F4B183" for cell in first_data_row)
        assert all(str(cell.fill.fgColor.rgb)[-6:].upper() != "C6E0B4" for cell in first_data_row)
    finally:
        workbook.close()


@pytest.mark.parametrize(
    "run_context",
    [
        {
            "started_at": 0.0,
            "logical_calls": app.MAX_LOGICAL_API_CALLS,
            "api_disabled": False,
        },
        {
            "started_at": 0.0,
            "logical_calls": 0,
            "api_disabled": True,
        },
    ],
    ids=["logical-call-limit-reached", "api-already-disabled"],
)
def test_extraction_precheck_never_calls_api_and_falls_back_for_all_sources(
    monkeypatch: pytest.MonkeyPatch,
    run_context: dict[str, Any],
) -> None:
    blocks = [
        {
            "source_id": f"T-P{ordinal:05d}",
            "text": f"Tender source block {ordinal} awaiting extraction.",
            "block_type": "paragraph",
            "ordinal": ordinal,
        }
        for ordinal in range(1, 4)
    ]
    calls = 0

    def forbidden_extract_unit(*args: Any, **kwargs: Any) -> list[dict[str, Any]]:
        nonlocal calls
        calls += 1
        raise AssertionError("The API unit must not run after the budget precheck fails.")

    monkeypatch.setattr(app, "_extract_requirement_unit", forbidden_extract_unit)
    token = app._API_RUN_CONTEXT.set(dict(run_context))
    try:
        requirements = app.extract_requirements_resilient(
            object(),
            "offline-mock",
            blocks,
            lambda _: None,
        )
    finally:
        app._API_RUN_CONTEXT.reset(token)

    assert calls == 0
    assert len(requirements) == 1
    fallback = requirements[0]
    assert fallback["source_ids"] == [block["source_id"] for block in blocks]
    assert fallback["extraction_uncertain"] is True
    assert fallback["risk_hint"] == "待人工复核"
    assert fallback["requirement_id"]


def test_extraction_task_budget_error_disables_api_and_returns_group_fallback(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    blocks = [
        {
            "source_id": f"T-P{ordinal:05d}",
            "text": f"Tender source block {ordinal}.",
            "block_type": "paragraph",
            "ordinal": ordinal,
        }
        for ordinal in range(1, 3)
    ]
    logs: list[str] = []
    run_context = {"started_at": app.monotonic(), "logical_calls": 0, "api_disabled": False}

    def budget_exhausted_extract(
        client: object,
        model: str,
        work_blocks: list[dict[str, Any]],
        logger: Any,
        batch_id: str,
        table_headers: dict[str, str] | None = None,
    ) -> list[dict[str, Any]]:
        del client, model, work_blocks, logger, batch_id, table_headers
        raise app.TaskBudgetError("synthetic extraction budget exhaustion")

    monkeypatch.setattr(app, "_extract_requirement_unit", budget_exhausted_extract)
    token = app._API_RUN_CONTEXT.set(run_context)
    try:
        requirements = app.extract_requirements_resilient(
            object(),
            "offline-mock",
            blocks,
            logs.append,
        )
        assert run_context["api_disabled"] is True
    finally:
        app._API_RUN_CONTEXT.reset(token)

    assert len(requirements) == 1
    assert requirements[0]["source_ids"] == [block["source_id"] for block in blocks]
    assert requirements[0]["extraction_uncertain"] is True
    assert logs


def test_assessment_task_budget_error_becomes_placeholders_and_disables_api(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    requirements = [make_requirement("REQ-BUDGET-A"), make_requirement("REQ-BUDGET-B")]
    logs: list[str] = []
    run_context = {"started_at": app.monotonic(), "logical_calls": 0, "api_disabled": False}

    def budget_exhausted_assessment(*args: Any, **kwargs: Any) -> list[dict[str, Any]]:
        del args, kwargs
        raise app.TaskBudgetError("synthetic assessment budget exhaustion")

    monkeypatch.setattr(app, "_assess_requirement_unit", budget_exhausted_assessment)
    token = app._API_RUN_CONTEXT.set(run_context)
    try:
        assessments = app.assess_requirement_batch_resilient(
            object(),
            "offline-mock",
            requirements,
            {item["requirement_id"]: [] for item in requirements},
            logs.append,
        )
        assert run_context["api_disabled"] is True
    finally:
        app._API_RUN_CONTEXT.reset(token)

    assert [item["requirement_id"] for item in assessments] == [
        "REQ-BUDGET-A",
        "REQ-BUDGET-B",
    ]
    assert all(item["status"] == "uncertain" for item in assessments)
    assert all(item["risk_level"] == "待人工复核" for item in assessments)
    assert all(item["estimated_score"] is None for item in assessments)
    assert logs


def test_single_and_group_scan_task_budget_errors_become_incomplete(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    requirements = [make_requirement("REQ-SCAN-A"), make_requirement("REQ-SCAN-B")]
    blocks = [
        {
            "source_id": "B-P00001",
            "text": "Bid source text for budget fallback testing.",
            "block_type": "paragraph",
            "ordinal": 1,
        }
    ]

    def budget_exhausted_scan(*args: Any, **kwargs: Any) -> Any:
        del args, kwargs
        raise app.TaskBudgetError("synthetic scan budget exhaustion")

    monkeypatch.setattr(app, "_scan_bid_unit_for_requirement", budget_exhausted_scan)
    monkeypatch.setattr(app, "_scan_bid_unit_for_requirements", budget_exhausted_scan)

    single_context = {"started_at": app.monotonic(), "logical_calls": 0, "api_disabled": False}
    token = app._API_RUN_CONTEXT.set(single_context)
    try:
        single_hits, single_complete = app.scan_bid_evidence_resilient(
            object(),
            "offline-mock",
            requirements[0],
            blocks,
            lambda _: None,
        )
        assert single_context["api_disabled"] is True
    finally:
        app._API_RUN_CONTEXT.reset(token)

    assert single_hits == []
    assert single_complete is False

    group_context = {"started_at": app.monotonic(), "logical_calls": 0, "api_disabled": False}
    token = app._API_RUN_CONTEXT.set(group_context)
    try:
        group_hits, group_complete = app.scan_bid_evidence_group_resilient(
            object(),
            "offline-mock",
            requirements,
            blocks,
            lambda _: None,
        )
        assert group_context["api_disabled"] is True
    finally:
        app._API_RUN_CONTEXT.reset(token)

    assert group_hits == {"REQ-SCAN-A": [], "REQ-SCAN-B": []}
    assert group_complete == {"REQ-SCAN-A": False, "REQ-SCAN-B": False}


def test_elapsed_time_guard_disables_all_three_api_stages_without_calling_units(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    requirement = make_requirement("REQ-TIME-GUARD")
    tender_blocks = [
        {
            "source_id": "T-P00001",
            "text": "Tender text protected by the elapsed-time guard.",
            "block_type": "paragraph",
            "ordinal": 1,
        }
    ]
    bid_blocks = [
        {
            "source_id": "B-P00001",
            "text": "Bid text protected by the elapsed-time guard.",
            "block_type": "paragraph",
            "ordinal": 1,
        }
    ]
    calls = {"extract": 0, "scan": 0, "assess": 0}

    def forbidden_extract(*args: Any, **kwargs: Any) -> Any:
        del args, kwargs
        calls["extract"] += 1
        raise AssertionError("elapsed-time extraction guard failed")

    def forbidden_scan(*args: Any, **kwargs: Any) -> Any:
        del args, kwargs
        calls["scan"] += 1
        raise AssertionError("elapsed-time scan guard failed")

    def forbidden_assess(*args: Any, **kwargs: Any) -> Any:
        del args, kwargs
        calls["assess"] += 1
        raise AssertionError("elapsed-time assessment guard failed")

    monkeypatch.setattr(app, "_extract_requirement_unit", forbidden_extract)
    monkeypatch.setattr(app, "_scan_bid_unit_for_requirements", forbidden_scan)
    monkeypatch.setattr(app, "_scan_bid_unit_for_requirement", forbidden_scan)
    monkeypatch.setattr(app, "_assess_requirement_unit", forbidden_assess)

    expired_context = {
        "started_at": app.monotonic() - app.MAX_TASK_SECONDS,
        "logical_calls": 0,
        "api_disabled": False,
    }
    token = app._API_RUN_CONTEXT.set(expired_context)
    try:
        assert app._extraction_budget_available() is False
        assert app._scan_budget_available() is False
        assert app._assessment_budget_available() is False

        extracted = app.extract_requirements_resilient(
            object(), "offline-mock", tender_blocks, lambda _: None
        )
        scan_hits, scan_complete = app.scan_bid_evidence_group_resilient(
            object(), "offline-mock", [requirement], bid_blocks, lambda _: None
        )
        single_scan_hits, single_scan_complete = app.scan_bid_evidence_resilient(
            object(), "offline-mock", requirement, bid_blocks, lambda _: None
        )
        assessments = app.assess_requirement_batch_resilient(
            object(),
            "offline-mock",
            [requirement],
            {requirement["requirement_id"]: []},
            lambda _: None,
        )
    finally:
        app._API_RUN_CONTEXT.reset(token)

    assert calls == {"extract": 0, "scan": 0, "assess": 0}
    assert extracted[0]["source_ids"] == ["T-P00001"]
    assert extracted[0]["extraction_uncertain"] is True
    assert scan_hits == {"REQ-TIME-GUARD": []}
    assert scan_complete == {"REQ-TIME-GUARD": False}
    assert single_scan_hits == []
    assert single_scan_complete is False
    assert assessments[0]["requirement_id"] == "REQ-TIME-GUARD"
    assert assessments[0]["status"] == "uncertain"
    assert assessments[0]["estimated_score"] is None


def test_analyze_preflight_budget_skip_still_returns_downloadable_excel(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    requirement = make_requirement(
        "REQ-PREFLIGHT",
        kind="check",
        module="qualification",
        title="business license",
        requirement_text="A valid business license is mandatory.",
        mandatory=True,
        risk_hint="rejection",
        full_score=None,
        source_ids=["P00001"],
        source_excerpt="A valid business license is mandatory.",
        origin_chunk_id="TB-PREFLIGHT",
        ordinal=1,
        extraction_uncertain=False,
    )
    extraction_calls = 0

    def fake_extract(
        client: object,
        model: str,
        blocks: list[dict[str, Any]],
        logger: Any,
        depth: int = 0,
        table_headers: dict[str, str] | None = None,
    ) -> list[dict[str, Any]]:
        nonlocal extraction_calls
        del client, model, blocks, logger, depth, table_headers
        extraction_calls += 1
        return [dict(requirement)]

    def forbidden_api_unit(*args: Any, **kwargs: Any) -> Any:
        del args, kwargs
        raise AssertionError("Preflight skip/manual fallback must avoid scan and assessment APIs.")

    monkeypatch.setattr(app, "extract_requirements_resilient", fake_extract)
    monkeypatch.setattr(app, "_scan_bid_unit_for_requirements", forbidden_api_unit)
    monkeypatch.setattr(app, "_assess_requirement_unit", forbidden_api_unit)

    logs: list[str] = []
    progress_events: list[tuple[int, str]] = []
    # 500 calls leaves extraction mocked as successful, but makes the planned
    # scan + reserved assessment budget exceed the 600-call hard limit.
    run_context = {"started_at": 0.0, "logical_calls": 500, "api_disabled": False}
    token = app._API_RUN_CONTEXT.set(run_context)
    try:
        result = app.analyze_long_documents(
            client=object(),
            model="offline-mock",
            tender_text="【P00001】 A valid business license is mandatory.",
            bid_text="【P00001】 The bidder supplied its business license.",
            tender_name="tender.docx",
            bid_name="bid.docx",
            logger=logs.append,
            progress=lambda value, label: progress_events.append((value, label)),
        )
    finally:
        app._API_RUN_CONTEXT.reset(token)

    assert extraction_calls == 1
    assert any("跳过付费全文补扫" in message for message in logs)
    assert len(result["requirements"]) == 1
    assert len(result["requirement_assessments"]) == 1
    assert result["requirement_assessments"][0]["status"] == "uncertain"
    assert result["requirement_assessments"][0]["risk_level"] == "待人工复核"
    assert progress_events

    report = app.build_excel_report(result)
    assert report.getbuffer().nbytes > 0
    workbook = load_workbook(report, data_only=False)
    try:
        assert workbook["缺陷核查记录"].max_row == 2
        assert workbook["预估打分表"].max_row == 1
    finally:
        workbook.close()
